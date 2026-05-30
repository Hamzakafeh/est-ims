import openpyxl
import os

# Read Other+.xlsm
f = 'zones\zone3\2026\05-May\Other+.xlsm'
print('File exists:', os.path.exists(f))

wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
ws = wb['Log']
rows = list(ws.iter_rows(values_only=True))[:15]
print('\n=== Other+.xlsm - Log Sheet ===')
for i,r in enumerate(rows):
    print(f'Row{i+1}: {r}')

ws = wb['Stocktaking']
rows = list(ws.iter_rows(values_only=True))[:15]
print('\n=== Other+.xlsm - Stocktaking Sheet ===')
for i,r in enumerate(rows):
    print(f'Row{i+1}: {r}')
wb.close()

# Read Sacks.xlsm
f = 'zones\zone3\2026\05-May\Sacks.xlsm'
wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
ws = wb['Log']
rows = list(ws.iter_rows(values_only=True))[:15]
print('\n=== Sacks.xlsm - Log Sheet ===')
for i,r in enumerate(rows):
    print(f'Row{i+1}: {r}')

ws = wb['Stocktaking']
rows = list(ws.iter_rows(values_only=True))[:15]
print('\n=== Sacks.xlsm - Stocktaking Sheet ===')
for i,r in enumerate(rows):
    print(f'Row{i+1}: {r}')
wb.close()
