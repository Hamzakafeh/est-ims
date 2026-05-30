"""Zone 1 — standalone Excel import / viewer / editor."""
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from flask import Blueprint, render_template, jsonify, request, session, redirect, url_for, send_from_directory

zone1_bp = Blueprint('zone1', __name__)

_Z1_DIR = os.path.join(
    os.getenv('RENDER_DISK_PATH',
              os.path.dirname(os.path.abspath(__file__))),
    'zone1_imports'
)


def _z1_allowed():
    return session.get('zone') == 'zone1' or session.get('is_super')


def _safe(name):
    name = re.sub(r'[^\w\-. ]', '', str(name)).strip()
    return name or 'file.xlsx'


def _fmt(val):
    """Serialize a cell value to a JSON-safe type."""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, float):
        return int(val) if val == int(val) else round(val, 6)
    return val


def _hex(argb):
    """Convert openpyxl ARGB string to #RRGGBB, or None if default/transparent."""
    if not argb or len(argb) < 6:
        return None
    rgb = argb[-6:] if len(argb) >= 6 else argb
    full = argb if len(argb) == 8 else ('FF' + rgb)
    # Skip transparent or white (default background)
    if full.upper() in ('00000000', 'FFFFFFFF', 'FF000000'):
        return None
    return '#' + full[-6:].upper()


# ── Pages ─────────────────────────────────────────────────────────

@zone1_bp.route('/zone1')
def zone1_page():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login_page'))
    if not _z1_allowed():
        return redirect(url_for('pages.index'))
    os.makedirs(_Z1_DIR, exist_ok=True)
    from core import get_firebase_config
    return render_template('zone1.html',
                           username=session.get('username', ''),
                           firebase_config=get_firebase_config())


# ── File list ─────────────────────────────────────────────────────

@zone1_bp.route('/api/zone1/files')
def api_z1_files():
    if not _z1_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    os.makedirs(_Z1_DIR, exist_ok=True)
    files = []
    for f in sorted(os.listdir(_Z1_DIR)):
        if f.lower().endswith(('.xlsx', '.xlsm', '.xls')) and not f.startswith('~$'):
            p = os.path.join(_Z1_DIR, f)
            try:
                files.append({
                    'name': f,
                    'size_kb': round(os.path.getsize(p) / 1024, 1),
                    'modified': datetime.fromtimestamp(
                        os.path.getmtime(p)).strftime('%Y-%m-%d %H:%M'),
                })
            except Exception:
                pass
    return jsonify({'files': files})


# ── Sheet list ────────────────────────────────────────────────────

@zone1_bp.route('/api/zone1/sheets')
def api_z1_sheets():
    if not _z1_allowed():
        return jsonify({'sheets': []}), 403
    fname = request.args.get('file', '').strip()
    path  = os.path.join(_Z1_DIR, _safe(fname))
    if not os.path.isfile(path):
        return jsonify({'sheets': []}), 404
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
        return jsonify({'sheets': sheets})
    except Exception as e:
        return jsonify({'sheets': [], 'error': str(e)})


# ── Sheet data (with full styling) ───────────────────────────────

@zone1_bp.route('/api/zone1/data')
def api_z1_data():
    if not _z1_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    fname = request.args.get('file',  '').strip()
    sname = request.args.get('sheet', '').strip()
    if not fname:
        return jsonify({'error': 'No file'}), 400
    path = os.path.join(_Z1_DIR, _safe(fname))
    if not os.path.isfile(path):
        return jsonify({'error': 'File not found'}), 404

    try:
        # data_only=True reads computed values not formulas
        # NOT read_only → we get cell styles
        wb = openpyxl.load_workbook(path, data_only=True)
        if sname not in wb.sheetnames:
            sname = wb.sheetnames[0]
        ws = wb[sname]

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0:
            wb.close()
            return jsonify({'headers': [], 'rows': [], 'total': 0,
                            'col_widths': [], 'row_heights': [], 'sheet': sname})

        # ── Merged cells ──────────────────────────────────────────
        merged_master = {}   # (row, col) → (rowspan, colspan)
        merged_skip   = set()  # cells to skip entirely
        for rng in ws.merged_cells.ranges:
            r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
            rs = r2 - r1 + 1
            cs = c2 - c1 + 1
            merged_master[(r1, c1)] = (rs, cs)
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    if (r, c) != (r1, c1):
                        merged_skip.add((r, c))

        # ── Column widths (px) ────────────────────────────────────
        col_widths = []
        for i in range(1, max_col + 1):
            letter = get_column_letter(i)
            dim = ws.column_dimensions.get(letter)
            w = (dim.width if dim and dim.width else 8.43)
            col_widths.append(max(40, int(w * 7.5)))

        # ── Row heights (px) ──────────────────────────────────────
        row_heights = []
        for i in range(1, max_row + 1):
            dim = ws.row_dimensions.get(i)
            h = (dim.height if dim and dim.height else 15)
            row_heights.append(max(20, int(h * 1.34)))

        # ── Rows + styles ─────────────────────────────────────────
        rows_out = []
        for r in range(1, max_row + 1):
            row_cells = []
            for c in range(1, max_col + 1):
                if (r, c) in merged_skip:
                    row_cells.append(None)   # frontend skips this
                    continue

                cell = ws.cell(row=r, column=c)
                cd   = {'v': _fmt(cell.value)}

                # Merge span
                if (r, c) in merged_master:
                    rs, cs = merged_master[(r, c)]
                    if rs > 1: cd['rs'] = rs
                    if cs > 1: cd['cs'] = cs

                # Background fill
                try:
                    fill = cell.fill
                    ft = fill.fill_type
                    if ft and ft != 'none':
                        fg = fill.fgColor
                        if fg.type == 'rgb':
                            h = _hex(fg.rgb)
                            if h: cd['bg'] = h
                        elif fg.type == 'indexed' and fg.indexed not in (0, 64):
                            # Indexed colours — map common ones
                            _IDXMAP = {1:'#FFFFFF',2:'#FF0000',3:'#00FF00',4:'#0000FF',
                                       5:'#FFFF00',6:'#FF00FF',7:'#00FFFF',
                                       40:'#FFFF99',41:'#99CCFF',42:'#00CCFF',
                                       43:'#CCFFCC',44:'#FFCC99',45:'#FF99CC',
                                       46:'#CC99FF',47:'#99FFFF'}
                            h = _IDXMAP.get(fg.indexed)
                            if h: cd['bg'] = h
                except Exception:
                    pass

                # Font
                try:
                    font = cell.font
                    if font.bold:   cd['b'] = 1
                    if font.italic: cd['i'] = 1
                    sz = font.size
                    if sz and sz != 11: cd['sz'] = int(sz)
                    fc = font.color
                    if fc and fc.type == 'rgb':
                        h = _hex(fc.rgb)
                        if h: cd['c'] = h
                except Exception:
                    pass

                # Alignment
                try:
                    al = cell.alignment.horizontal
                    if al and al not in ('general', None):
                        cd['al'] = al
                    vl = cell.alignment.vertical
                    if vl and vl not in ('bottom', None):
                        cd['vl'] = vl
                    if cell.alignment.wrap_text:
                        cd['wrap'] = 1
                except Exception:
                    pass

                row_cells.append(cd)
            rows_out.append(row_cells)

        wb.close()

        # First row as column headers (just values)
        headers = [str(cd.get('v', '') or f'Col {i+1}')
                   for i, cd in enumerate(rows_out[0]) if cd is not None]

        return jsonify({
            'headers':     headers,
            'rows':        rows_out,
            'total':       len(rows_out),
            'col_widths':  col_widths,
            'row_heights': row_heights,
            'sheet':       sname,
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Import ────────────────────────────────────────────────────────

@zone1_bp.route('/api/zone1/import', methods=['POST'])
def api_z1_import():
    if not _z1_allowed():
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xlsm', '.xls'):
        return jsonify({'success': False,
                        'message': 'Only .xlsx / .xlsm / .xls allowed'}), 400
    os.makedirs(_Z1_DIR, exist_ok=True)
    safe = _safe(file.filename)
    path = os.path.join(_Z1_DIR, safe)
    file.save(path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
    except Exception as e:
        os.remove(path)
        return jsonify({'success': False, 'message': f'Cannot read: {e}'}), 400
    return jsonify({'success': True, 'name': safe, 'sheets': sheets})


# ── Edit cell ─────────────────────────────────────────────────────

@zone1_bp.route('/api/zone1/edit', methods=['POST'])
def api_z1_edit():
    if not _z1_allowed():
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    body  = request.get_json(silent=True) or {}
    fname = body.get('file',  '').strip()
    sname = body.get('sheet', '').strip()
    row   = int(body.get('row', -1))   # 0-based data row (after header)
    col   = int(body.get('col', -1))   # 0-based column
    value = body.get('value', '')
    if not fname or not sname or row < 0 or col < 0:
        return jsonify({'success': False, 'message': 'Missing params'}), 400
    path = os.path.join(_Z1_DIR, _safe(fname))
    if not os.path.isfile(path):
        return jsonify({'success': False, 'message': 'File not found'}), 404
    try:
        keep_vba = fname.lower().endswith('.xlsm')
        wb = openpyxl.load_workbook(path, keep_vba=keep_vba)
        if sname not in wb.sheetnames:
            return jsonify({'success': False, 'message': 'Sheet not found'}), 404
        ws = wb[sname]
        # row+2 → skip header row (1) + convert to 1-based (+1)
        ws.cell(row=row + 2, column=col + 1).value = value if value != '' else None
        wb.save(path)
        wb.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ── Delete file ───────────────────────────────────────────────────

@zone1_bp.route('/api/zone1/file/<path:filename>', methods=['DELETE'])
def api_z1_delete(filename):
    if not _z1_allowed():
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    path = os.path.join(_Z1_DIR, _safe(filename))
    if not os.path.isfile(path):
        return jsonify({'success': False, 'message': 'Not found'}), 404
    try:
        os.remove(path)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ── Download file ─────────────────────────────────────────────────

@zone1_bp.route('/api/zone1/download/<path:filename>')
def api_z1_download(filename):
    if not _z1_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    safe = _safe(filename)
    if not os.path.isfile(os.path.join(_Z1_DIR, safe)):
        return jsonify({'error': 'Not found'}), 404
    return send_from_directory(_Z1_DIR, safe, as_attachment=True)


# ── Create new Excel file ─────────────────────────────────────────

@zone1_bp.route('/api/zone1/create', methods=['POST'])
def api_z1_create():
    if not _z1_allowed():
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    body       = request.get_json(silent=True) or {}
    name       = body.get('name', 'New File').strip()
    sheet_name = body.get('sheet', 'Sheet1').strip() or 'Sheet1'
    headers    = body.get('headers', [])   # list of column header strings

    if not name:
        return jsonify({'success': False, 'message': 'File name required'}), 400
    if not name.lower().endswith(('.xlsx', '.xlsm')):
        name += '.xlsx'
    safe = _safe(name)
    os.makedirs(_Z1_DIR, exist_ok=True)
    path = os.path.join(_Z1_DIR, safe)

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Write header row if provided
        if headers:
            for c_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c_idx, value=str(h))
                cell.font   = openpyxl.styles.Font(bold=True)
                cell.fill   = openpyxl.styles.PatternFill(
                    fill_type='solid', fgColor='0F2137')
                cell.font   = openpyxl.styles.Font(bold=True, color='FFFFFF')
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        wb.save(path)
        wb.close()
        return jsonify({'success': True, 'name': safe, 'sheets': [sheet_name]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
