import os
import openpyxl
from datetime import datetime
from flask import Blueprint, jsonify, request, session
from core import (
    zone_required,
    _validate_filepath,
    read_sheet_data,
    get_structure,
    get_years_root,
    get_available_years,
    _get_cell_val,
    _find_last_balance,
    _append_log,
    _recalc_stocktaking,
    _parse_dv_formula,
    _col_letter_to_index,
    COL_COLOR, COL_SIZE, COL_TYPE, COL_CATEGORY,
    COL_BASIC, COL_CURRENT, DATA_START_ROW,
    ZONES, SUPER_ZONES
)

excel_bp = Blueprint('excel', __name__)


@excel_bp.route('/api/structure')
@zone_required
def api_structure():
    zone = session.get('zone', '')
    is_super = session.get('is_super', False)
    if is_super:
        view_zone = request.args.get('zone') or session.get('active_view_zone', 'zone1')
    else:
        view_zone = zone
    return jsonify(get_structure(view_zone))


@excel_bp.route('/api/years')
@zone_required
def api_years():
    return jsonify({'years': get_available_years()})


@excel_bp.route('/api/sheets')
@zone_required
def api_sheets():
    filepath = request.args.get('path')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'error': 'Access denied'}), 403
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames; wb.close()
        return jsonify({'sheets': sheets})
    except Exception:
        return jsonify({'error': 'Failed to read file'}), 500


@excel_bp.route('/api/data')
@zone_required
def api_data():
    filepath = request.args.get('path')
    sheet = request.args.get('sheet','')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'error': 'Access denied'}), 403
    headers, rows = read_sheet_data(filepath, sheet)
    if headers is None:
        return jsonify({'headers':[],'rows':[],'count':0})
    return jsonify({'headers': headers, 'rows': rows, 'count': len(rows)})


@excel_bp.route('/api/transaction', methods=['POST'])
@zone_required
def api_transaction():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح — يجب تفعيل وضع التعديل'}), 403
    data = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet = data.get('sheet', '')
    row = data.get('row')
    operation = data.get('operation', '').upper()
    qty_raw = data.get('qty')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    if operation not in ('IN', 'OUT'):
        return jsonify({'success': False, 'error': 'operation must be IN or OUT'}), 400
    try:
        row = int(row)
        qty = float(qty_raw)
        if qty < 0:
            raise ValueError()
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid row or qty'}), 400
    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404
        ws = wb[sheet]
        ws_log = wb['Log'] if 'Log' in wb.sheetnames else None
        color_value = _get_cell_val(ws, row, COL_COLOR)
        size_value = _get_cell_val(ws, row, COL_SIZE)
        type_value = _get_cell_val(ws, row, COL_TYPE)
        category_value = _get_cell_val(ws, row, COL_CATEGORY)
        basic_balance = ws.cell(row=row, column=COL_BASIC).value or 0
        if not color_value or str(color_value).strip() in ('', 'None', 'null'):
            wb.close()
            return jsonify({'success': False, 'error': 'يجب تحديد اللون (Color) أولاً قبل إجراء أي عملية'}), 400
        last_balance, found = _find_last_balance(ws, row, color_value)
        if not found:
            try:
                last_balance = float(basic_balance)
            except Exception:
                last_balance = 0.0
        if operation == 'IN':
            new_balance = last_balance + qty
        else:
            new_balance = last_balance - qty
        ws.cell(row=row, column=COL_CURRENT).value = new_balance
        if ws_log:
            _append_log(ws_log, operation, qty, new_balance, color_value, size_value, type_value, category_value)
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()
        return jsonify({
            'success': True,
            'new_balance': new_balance,
            'operation': operation,
            'qty': qty,
            'color': color_value,
            'size': size_value,
        })
    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء تنفيذ العملية'}), 500


@excel_bp.route('/api/update_cell', methods=['POST'])
@zone_required
def api_update_cell():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet = data.get('sheet', '')
    row = data.get('row')
    col_name = data.get('col_name', '')
    value = data.get('value', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    try:
        row = int(row)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid row'}), 400
    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404
        ws = wb[sheet]
        header_row_idx = None
        for i, row_cells in enumerate(ws.iter_rows(values_only=True), start=1):
            rv = [str(v).strip() if v else '' for v in row_cells]
            if 'Date' in rv or 'التاريخ' in rv:
                header_row_idx = i; break
        if header_row_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': 'Header row not found'}), 400
        col_idx = None
        for ci, cell in enumerate(ws[header_row_idx], start=1):
            if cell.value and str(cell.value).strip() == col_name:
                col_idx = ci; break
        if col_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': f'Column "{col_name}" not found'}), 400
        cast_value = value
        try:
            if '.' in str(value):
                cast_value = float(value)
            else:
                cast_value = int(value)
        except Exception:
            cast_value = value if value != '' else None
        ws.cell(row=row, column=col_idx).value = cast_value
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()
        return jsonify({'success': True, 'row': row, 'col': col_name, 'value': cast_value})
    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء تحديث الخلية'}), 500


@excel_bp.route('/api/color_balance')
@zone_required
def api_color_balance():
    filepath = request.args.get('path', '')
    sheet = request.args.get('sheet', '')
    color = request.args.get('color', '')
    before_row = request.args.get('before_row', None)
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if not color or color.strip().lower() in ('', 'null', 'none'):
        return jsonify({'balance': None, 'found': False})
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'error': 'Access denied'}), 403
    try:
        before_row = int(before_row) if before_row else None
    except Exception:
        before_row = None
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'balance': None, 'found': False})
        ws = wb[sheet]
        last_balance = None
        last_row = None
        rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))
        for idx, row_vals in enumerate(rows):
            actual_row = DATA_START_ROW + idx
            if before_row and actual_row >= before_row:
                break
            try:
                cell_color = row_vals[COL_COLOR - 1]
                cell_current = row_vals[COL_CURRENT - 1]
            except IndexError:
                continue
            if cell_color and str(cell_color).strip() == color.strip():
                try:
                    val = float(cell_current) if cell_current is not None else None
                    if val is not None:
                        last_balance = val
                        last_row = actual_row
                except Exception:
                    pass
        wb.close()
        return jsonify({'balance': last_balance, 'found': last_balance is not None, 'row': last_row})
    except Exception:
        return jsonify({'error': 'حدث خطأ أثناء قراءة الرصيد'}), 500


@excel_bp.route('/api/set_opening_balance', methods=['POST'])
@zone_required
def api_set_opening_balance():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet = data.get('sheet', '')
    row = data.get('row')
    balance = data.get('balance')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    try:
        row = int(row)
        balance = float(balance)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid row or balance'}), 400
    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        ws = wb[sheet]
        ws.cell(row=row, column=COL_BASIC).value = balance
        ws.cell(row=row, column=COL_CURRENT).value = balance
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()
        return jsonify({'success': True, 'balance': balance})
    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء تعيين الرصيد الافتتاحي'}), 500


@excel_bp.route('/api/options')
@zone_required
def api_options():
    filepath = request.args.get('path', '')
    sheet = request.args.get('sheet', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'error': 'Access denied'}), 403
    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'colors': [], 'types': [], 'sizes': [], 'categories': []})
        ws = wb[sheet]
        header_col_map = {}
        for ci, cell in enumerate(ws[6], start=1):
            if cell.value and str(cell.value).strip():
                header_col_map[ci] = str(cell.value).strip()
        name_to_col = {v: k for k, v in header_col_map.items()}
        col_color = name_to_col.get('Color', COL_COLOR)
        col_type = name_to_col.get('Type', COL_TYPE)
        col_size = name_to_col.get('Size', COL_SIZE)
        col_category = name_to_col.get('Category', COL_CATEGORY)
        options = {'colors': [], 'types': [], 'sizes': [], 'categories': []}
        col_target = {col_color: 'colors', col_type: 'types', col_size: 'sizes', col_category: 'categories'}
        for dv in ws.data_validations.dataValidation:
            if dv.type != 'list' or not dv.formula1:
                continue
            try:
                first_ref = str(dv.sqref).split()[0]
                col_letters = ''.join(c for c in first_ref.split(':')[0] if c.isalpha())
                ci = _col_letter_to_index(col_letters)
            except Exception:
                continue
            key = col_target.get(ci)
            if key:
                vals = _parse_dv_formula(dv.formula1)
                options[key] = vals
        wb.close()
        return jsonify(options)
    except Exception:
        return jsonify({'error': 'حدث خطأ أثناء قراءة الخيارات'}), 500


@excel_bp.route('/api/clear_row', methods=['POST'])
@zone_required
def api_clear_row():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet = data.get('sheet', '')
    row = data.get('row')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    try:
        row = int(row)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid row'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        ws = wb[sheet]
        for col in range(1, 14):
            ws.cell(row=row, column=col).value = None
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()
        return jsonify({'success': True})
    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء مسح الصف'}), 500


@excel_bp.route('/api/add_row', methods=['POST'])
@zone_required
def api_add_row():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403

    data = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet = data.get('sheet', '')
    fields = data.get('fields', {})

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    if not fields:
        return jsonify({'success': False, 'error': 'No fields provided'}), 400

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404
        ws = wb[sheet]

        header_row_idx = None
        for i, row_cells in enumerate(ws.iter_rows(values_only=True), start=1):
            rv = [str(v).strip() if v else '' for v in row_cells]
            if 'Date' in rv or 'التاريخ' in rv:
                header_row_idx = i; break

        if header_row_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': 'Header row not found'}), 400

        col_map = {}
        for ci, cell in enumerate(ws[header_row_idx], start=1):
            if cell.value:
                col_map[str(cell.value).strip()] = ci

        new_row = DATA_START_ROW
        for r in range(DATA_START_ROW, ws.max_row + 2):
            row_empty = True
            for ci in range(1, 14):
                if ws.cell(row=r, column=ci).value not in (None, ''):
                    row_empty = False
                    break
            if row_empty:
                new_row = r
                break

        for col_name, value in fields.items():
            if col_name in col_map:
                cast_value = value
                try:
                    if '.' in str(value):
                        cast_value = float(value)
                    else:
                        cast_value = int(value)
                except Exception:
                    cast_value = value if value != '' else None
                ws.cell(row=new_row, column=col_map[col_name]).value = cast_value

        if 'Date' not in fields and 'التاريخ' not in fields:
            date_col = col_map.get('Date') or col_map.get('التاريخ')
            if date_col:
                ws.cell(row=new_row, column=date_col).value = datetime.now().strftime('%Y-%m-%d')

        wb.save(filepath)
        wb.close()
        return jsonify({'success': True, 'new_row': new_row})

    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء إضافة الصف'}), 500