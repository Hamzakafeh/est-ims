import io
import json
import os
import openpyxl
from datetime import datetime, timedelta
from flask import Blueprint, jsonify, request, session, send_file
from core import (
    zone_required, _db_connect, _normalize_username, _hash_secret,
    _is_single_login_exempt, _clear_active_session, AUTH_DB_FILE,
    CONTACT_MESSAGES_FILE, _read_json_list, _write_json_list, _data_lock,
    _firebase_set_user_status, _firebase_clear_user_status, DATA_STORE_DIR,
)

USER_ZONES_FILE = os.path.join(DATA_STORE_DIR, 'user_zones.json')


def _read_user_zones():
    try:
        with open(USER_ZONES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _write_user_zones(data):
    os.makedirs(DATA_STORE_DIR, exist_ok=True)
    with open(USER_ZONES_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

admin_bp = Blueprint('admin', __name__)


@admin_bp.route('/api/admin/pending_requests_count')
@zone_required
def api_admin_pending_requests_count():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    # _pending_request_count is in app.py; compute via DB
    with _db_connect() as conn:
        row = conn.execute("SELECT COUNT(1) as c FROM registration_requests WHERE status = 'pending'").fetchone()
        return jsonify({'count': int(row['c'] if row and 'c' in row.keys() else 0)})


@admin_bp.route('/api/admin/registration_requests')
@zone_required
def api_admin_registration_requests():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        rows = conn.execute(
            "SELECT * FROM registration_requests WHERE status = 'pending' ORDER BY created_at DESC, id DESC"
        ).fetchall()
    return jsonify({
        'count': len(rows),
        'requests': [
            {
                'id': row['id'],
                'full_name': row['full_name'],
                'username': row['username'],
                'email': row['email'],
                'phone': row['phone'],
                'job_title': row['job_title'],
                'gender': row['gender'] if 'gender' in row.keys() else '',
                'birth_date': row['birth_date'] if 'birth_date' in row.keys() else '',
                'privacy_accepted': bool(row['privacy_accepted'] if 'privacy_accepted' in row.keys() else 0),
                'security_question': row['security_question'],
                'created_at': row['created_at'],
            }
            for row in rows
        ],
    })


@admin_bp.route('/api/admin/registration_requests/<int:request_id>/approve', methods=['POST'])
@zone_required
def api_admin_approve_registration(request_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    admin_user = session.get('username', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _db_connect() as conn:
        row = conn.execute(
            "SELECT * FROM registration_requests WHERE id = ? AND status = 'pending'",
            (request_id,),
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
        exists = conn.execute(
            "SELECT 1 FROM users WHERE lower(username) = ?",
            (_normalize_username(row['username']),),
        ).fetchone()
        if exists:
            conn.execute(
                "UPDATE registration_requests SET status = 'approved', reviewed_at = ?, reviewed_by = ? WHERE id = ?",
                (now, admin_user, request_id),
            )
            return jsonify({'success': False, 'message': 'اسم المستخدم موجود مسبقاً'}), 409
        conn.execute(
            """
            INSERT INTO users
            (full_name, username, email, phone, job_title, gender, birth_date, privacy_accepted, password_hash, security_question, security_answer_hash, approved, is_admin, created_at, approved_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 0, ?, ?, ?)
            """,
            (
                row['full_name'],
                row['username'],
                row['email'],
                row['phone'],
                row['job_title'],
                row['gender'] if 'gender' in row.keys() else '',
                row['birth_date'] if 'birth_date' in row.keys() else '',
                row['privacy_accepted'] if 'privacy_accepted' in row.keys() else 0,
                row['password_hash'],
                row['security_question'],
                row['security_answer_hash'],
                row['created_at'],
                now,
                admin_user,
            ),
        )
        conn.execute(
            "UPDATE registration_requests SET status = 'approved', reviewed_at = ?, reviewed_by = ? WHERE id = ?",
            (now, admin_user, request_id),
        )
    return jsonify({'success': True, 'message': 'تمت الموافقة على الطلب'})


@admin_bp.route('/api/admin/registration_requests/<int:request_id>/reject', methods=['POST'])
@zone_required
def api_admin_reject_registration(request_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    admin_user = session.get('username', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _db_connect() as conn:
        row = conn.execute(
            "SELECT id FROM registration_requests WHERE id = ? AND status = 'pending'",
            (request_id,),
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
        conn.execute(
            "UPDATE registration_requests SET status = 'rejected', reviewed_at = ?, reviewed_by = ? WHERE id = ?",
            (now, admin_user, request_id),
        )
    return jsonify({'success': True, 'message': 'تم رفض الطلب'})


@admin_bp.route('/api/admin/registered_users')
@zone_required
def api_admin_registered_users():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        rows = conn.execute(
            "SELECT id, full_name, username, email, phone, job_title, gender, birth_date, privacy_accepted, security_question, password_hash, security_answer_hash, approved_at, created_at, suspended_until, suspended_by, suspended_at, is_verified, perm_switch_zones, perm_manage_permissions, perm_can_edit FROM users WHERE approved = 1 ORDER BY approved_at DESC, id DESC"
        ).fetchall()
    keys = lambda row, k, default=None: row[k] if k in row.keys() else default
    return jsonify({
        'count': len(rows),
        'db_file': AUTH_DB_FILE,
        'users': [
            {
                'id': row['id'],
                'full_name': row['full_name'],
                'username': row['username'],
                'email': row['email'],
                'phone': row['phone'],
                'job_title': row['job_title'],
                'gender': keys(row, 'gender', ''),
                'birth_date': keys(row, 'birth_date', ''),
                'privacy_accepted': bool(keys(row, 'privacy_accepted', 0)),
                'security_question': row['security_question'],
                'password_stored_as': 'one_way_hash' if row['password_hash'] else '',
                'security_answer_stored_as': 'one_way_hash' if row['security_answer_hash'] else '',
                'approved_at': row['approved_at'],
                'created_at': row['created_at'],
                'suspended_until': keys(row, 'suspended_until'),
                'suspended_by': keys(row, 'suspended_by'),
                'suspended_at': keys(row, 'suspended_at'),
                'is_verified': bool(keys(row, 'is_verified', 0)),
                'perm_switch_zones': bool(keys(row, 'perm_switch_zones', 0)),
                'perm_manage_permissions': bool(keys(row, 'perm_manage_permissions', 0)),
                'can_edit': bool(keys(row, 'perm_can_edit', 0)),
            }
            for row in rows
        ],
    })


@admin_bp.route('/api/admin/registered_users/export.xlsx')
@zone_required
def api_admin_export_registered_users():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        rows = conn.execute("""
            SELECT id, full_name, username, email, phone, job_title, gender, birth_date, privacy_accepted,
                   security_question, password_hash, security_answer_hash,
                   approved, is_admin, created_at, approved_at, created_by,
                   suspended_until, suspended_by, suspended_at
            FROM users
            WHERE approved = 1
            ORDER BY approved_at DESC, id DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registered Users'
    headers = [
        'ID', 'Full Name', 'Username', 'Email', 'Phone', 'Job Title', 'Gender', 'Birth Date', 'Privacy Accepted',
        'Password', 'Security Question', 'Security Answer',
        'Password Hash', 'Security Answer Hash',
        'Approved', 'Is Admin', 'Created At', 'Approved At', 'Created By',
        'Suspended Until', 'Suspended By', 'Suspended At',
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row['id'], row['full_name'], row['username'], row['email'], row['phone'], row['job_title'],
            row['gender'] if 'gender' in row.keys() else '',
            row['birth_date'] if 'birth_date' in row.keys() else '',
            'Yes' if (row['privacy_accepted'] if 'privacy_accepted' in row.keys() else 0) else 'No',
            'Not recoverable - stored as one-way hash',
            row['security_question'],
            'Not recoverable - stored as one-way hash',
            row['password_hash'], row['security_answer_hash'],
            row['approved'], row['is_admin'], row['created_at'], row['approved_at'], row['created_by'],
            row['suspended_until'], row['suspended_by'], row['suspended_at'],
        ])
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        try:
            from openpyxl.styles import Font, PatternFill
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9EAF7')
        except Exception:
            pass
    for col in ws.columns:
        try:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                max_len = max(max_len, len(str(cell.value or '')))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)
        except Exception:
            pass

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = 'registered_users_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@admin_bp.route('/api/admin/registered_users/<int:user_id>/suspend', methods=['POST'])
@zone_required
def api_admin_suspend_user(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    minutes = int(data.get('minutes') or 0)
    if minutes <= 0 or minutes > 43200:
        return jsonify({'success': False, 'message': 'مدة الإيقاف يجب أن تكون بين دقيقة و 30 يوم'}), 400
    now = datetime.now()
    until = now + timedelta(minutes=minutes)
    admin_user = session.get('username', '')
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        if _is_single_login_exempt(row['username']):
            return jsonify({'success': False, 'message': 'لا يمكن إيقاف حساب الأدمن أو الديف'}), 400
        conn.execute(
            "UPDATE users SET suspended_until = ?, suspended_by = ?, suspended_at = ? WHERE id = ?",
            (until.strftime('%Y-%m-%d %H:%M:%S'), admin_user, now.strftime('%Y-%m-%d %H:%M:%S'), user_id),
        )
    _clear_active_session(row['username'])
    _firebase_set_user_status(
        row['username'], 'suspended',
        f'تم إيقاف حسابك مؤقتاً حتى {until.strftime("%Y-%m-%d %H:%M")} بواسطة الإدارة'
    )
    return jsonify({'success': True, 'message': 'تم إيقاف المستخدم مؤقتاً', 'suspended_until': until.strftime('%Y-%m-%d %H:%M:%S')})


@admin_bp.route('/api/admin/registered_users/<int:user_id>/unsuspend', methods=['POST'])
@zone_required
def api_admin_unsuspend_user(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        cur = conn.execute(
            "UPDATE users SET suspended_until = NULL, suspended_by = NULL, suspended_at = NULL WHERE id = ?",
            (user_id,),
        )
    if cur.rowcount == 0:
        return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
    _firebase_clear_user_status(row['username'])
    return jsonify({'success': True, 'message': 'تم إلغاء الإيقاف'})


@admin_bp.route('/api/admin/registered_users/<int:user_id>/password', methods=['POST'])
@zone_required
def api_admin_reset_user_password(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    new_password = str(data.get('new_password', '')).strip()
    confirm_password = str(data.get('confirm_password', '')).strip()
    if not new_password:
        return jsonify({'success': False, 'message': 'يرجى إدخال كلمة المرور الجديدة'}), 400
    if new_password != confirm_password:
        return jsonify({'success': False, 'message': 'كلمة المرور وتأكيدها غير متطابقين'}), 400
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (_hash_secret(new_password), user_id))
    _clear_active_session(row['username'])
    return jsonify({'success': True, 'message': 'تم تغيير كلمة مرور المستخدم'})


@admin_bp.route('/api/admin/registered_users/<int:user_id>', methods=['DELETE'])
@zone_required
def api_admin_delete_user(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        if _is_single_login_exempt(row['username']):
            return jsonify({'success': False, 'message': 'لا يمكن حذف حساب الأدمن أو الديف'}), 400
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    _clear_active_session(row['username'])
    _firebase_set_user_status(row['username'], 'deleted', 'تم حذف حسابك من النظام بواسطة الإدارة')
    return jsonify({'success': True, 'message': 'تم حذف المستخدم'})


@admin_bp.route('/api/admin/contact_messages')
@zone_required
def api_admin_contact_messages():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _data_lock:
        messages = _read_json_list(CONTACT_MESSAGES_FILE)
    unread = sum(1 for m in messages if m.get('status') == 'new')
    return jsonify({'messages': list(reversed(messages)), 'count': unread})


@admin_bp.route('/api/admin/contact_messages/<int:msg_id>/read', methods=['POST'])
@zone_required
def api_admin_contact_message_read(msg_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _data_lock:
        messages = _read_json_list(CONTACT_MESSAGES_FILE)
        for m in messages:
            if int(m.get('id', -1)) == msg_id:
                m['status'] = 'read'
                break
        _write_json_list(CONTACT_MESSAGES_FILE, messages)
    return jsonify({'success': True})


@admin_bp.route('/api/admin/contact_messages/<int:msg_id>', methods=['DELETE'])
@zone_required
def api_admin_contact_message_delete(msg_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _data_lock:
        messages = _read_json_list(CONTACT_MESSAGES_FILE)
        messages = [m for m in messages if int(m.get('id', -1)) != msg_id]
        _write_json_list(CONTACT_MESSAGES_FILE, messages)
    return jsonify({'success': True})


@admin_bp.route('/api/admin/registered_users/<int:user_id>/permissions', methods=['POST'])
@zone_required
def api_admin_set_user_permissions(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    switch_zones = int(bool(data.get('switch_zones', False)))
    can_edit = int(bool(data.get('can_edit', False)))
    manage_perms = int(bool(data.get('manage_permissions', False)))
    with _db_connect() as conn:
        row = conn.execute("SELECT id FROM users WHERE id = ? AND approved = 1", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        conn.execute(
            "UPDATE users SET perm_switch_zones=?, perm_can_edit=?, perm_manage_permissions=? WHERE id=?",
            (switch_zones, can_edit, manage_perms, user_id)
        )
    return jsonify({'success': True})


@admin_bp.route('/api/admin/registered_users/<int:user_id>/toggle_verified', methods=['POST'])
@zone_required
def api_admin_toggle_user_verified(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    is_verified = int(bool(data.get('is_verified', False)))
    with _db_connect() as conn:
        row = conn.execute("SELECT id FROM users WHERE id = ? AND approved = 1", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        conn.execute("UPDATE users SET is_verified=? WHERE id=?", (is_verified, user_id))
    return jsonify({'success': True})


@admin_bp.route('/api/admin/registered_users/<int:user_id>/zones', methods=['GET'])
@zone_required
def api_admin_get_user_zones(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        return jsonify({'error': 'المستخدم غير موجود'}), 404
    username = row['username'].lower()
    zones = _read_user_zones()
    return jsonify({'zones': zones.get(username, None)})


@admin_bp.route('/api/admin/registered_users/<int:user_id>/zones', methods=['POST'])
@zone_required
def api_admin_set_user_zones(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    allowed = data.get('zones')  # None = all zones, [] = no zones, [...] = specific
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        return jsonify({'error': 'المستخدم غير موجود'}), 404
    username = row['username'].lower()
    with _data_lock:
        zones = _read_user_zones()
        if allowed is None:
            zones.pop(username, None)  # remove restriction = all zones
        else:
            zones[username] = [str(z) for z in allowed]
        _write_user_zones(zones)
    return jsonify({'success': True})
