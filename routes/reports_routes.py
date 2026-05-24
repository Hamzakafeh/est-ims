import os
import openpyxl
from datetime import datetime
from flask import Blueprint, jsonify, send_file, Response

reports_bp = Blueprint('reports', __name__)

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reports')
REPORTS_ALLOWED = ('.xlsx', '.xlsm', '.xls', '.csv', '.docx', '.doc', '.dotx', '.pptx', '.ppt', '.pps', '.pdf')


@reports_bp.route('/api/reports')
def api_reports():
    if not os.path.isdir(REPORTS_DIR):
        return jsonify({'files': []})
    files = [f for f in sorted(os.listdir(REPORTS_DIR)) if f.lower().endswith(REPORTS_ALLOWED)]
    return jsonify({'files': files})


@reports_bp.route('/reports/file/<path:filename>')
def download_report(filename):
    safe_name = os.path.basename(filename)
    filepath = os.path.join(REPORTS_DIR, safe_name)
    if not os.path.isfile(filepath):
        return 'File not found', 404
    if not safe_name.lower().endswith(REPORTS_ALLOWED):
        return 'File type not allowed', 403
    return send_file(filepath, as_attachment=False)


@reports_bp.route('/reports/print/<path:filename>')
def print_report(filename):
    safe_name = os.path.basename(filename)
    filepath = os.path.join(REPORTS_DIR, safe_name)
    if not os.path.isfile(filepath):
        return 'File not found', 404
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return f'Could not open file: {e}', 500
    sheets_html = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data:
            continue
        table = f'<h2 class="sheet-title">{sheet_name}</h2><table>'
        for ri, row in enumerate(rows_data):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            tag = 'th' if ri == 0 else 'td'
            cells = ''.join(f'<{tag}>{("" if (c is None or str(c).strip() == "") else str(c))}</{tag}>' for c in row)
            table += f'<tr>{cells}</tr>'
        table += '</table>'
        sheets_html.append(table)
    wb.close()
    html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <title>{safe_name}</title>
  <style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family: Arial, sans-serif; font-size: 11px; padding: 16px; background:#fff; color:#000; }}
    .sheet-title {{ font-size:14px; font-weight:700; margin: 18px 0 6px; color:#1a3a5c; border-bottom:2px solid #1a3a5c; padding-bottom:4px; }}
    table {{ border-collapse: collapse; width:100%; margin-bottom: 24px; page-break-inside: avoid; }}
    th, td {{ border: 1px solid #bbb; padding: 5px 8px; text-align: center; white-space: nowrap; }}
    th {{ background: #1a3a5c; color: #fff; font-size:11px; }}
    tr:nth-child(even) td {{ background: #f5f7fa; }}
    @media print {{ body {{ padding:8px; }} .sheet-title {{ margin-top:10px; }} @page {{ margin: 1cm; size: landscape; }} }}
  </style>
</head>
<body>
  <div style="text-align:center;margin-bottom:14px;">
    <strong style="font-size:15px;">{safe_name.rsplit('.',1)[0]}</strong>
    <span style="font-size:11px;color:#666;margin-left:10px;">Printed: {datetime.now().strftime('%Y-%m-%d %H:%M')}</span>
  </div>
  {''.join(sheets_html)}
  <script>window.onload = function(){{ window.print(); }};</script>
</body>
</html>"""
    return Response(html, mimetype='text/html')
