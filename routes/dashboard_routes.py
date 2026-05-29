"""Dashboard and inventory scan metrics."""
import os
import re
import openpyxl
from datetime import datetime
from flask import Blueprint, jsonify, request, session
from core import (
    zone_required,
    get_years_root,
    ZONES,
    SUPER_ZONES,
    COL_COLOR,
    COL_TYPE,
    COL_SIZE,
    COL_CATEGORY,
    COL_CURRENT,
    _log_lock,
    _read_login_log,
    _LOGIN_LOG_FILE,
)

dashboard_bp = Blueprint('dashboard', __name__)

@dashboard_bp.route('/api/dashboard')
@zone_required
def api_dashboard():
    """Scan Excel files for the current dashboard scope and return live metrics."""
    zone_id = session.get('active_view_zone') or session.get('zone', '')
    is_super = session.get('is_super', False)
    requested_zone = request.args.get('zone', '').strip()

    root = get_years_root()
    if not root:
        return jsonify({'error': 'No data directory'}), 404

    files_to_scan = []
    available_dashboard_zones = [z for z in ZONES if z['id'] not in SUPER_ZONES]
    valid_zone_ids = {z['id'] for z in available_dashboard_zones}
    if is_super and requested_zone and requested_zone != 'all':
        if requested_zone not in valid_zone_ids:
            return jsonify({'error': 'Invalid dashboard zone'}), 400
        scan_zones = [requested_zone]
    elif is_super:
        scan_zones = [z['id'] for z in available_dashboard_zones]
    else:
        scan_zones = [zone_id]

    for zid in scan_zones:
        zone_path = os.path.join(root, zid)
        if not os.path.isdir(zone_path):
            continue
        for root_dir, dirs, files in os.walk(zone_path):
            for f in files:
                lf = f.lower()
                if f.startswith('~$'):
                    continue
                if lf in ('other+.xlsm', 'sacks.xlsm'):
                    files_to_scan.append((zid, os.path.join(root_dir, f)))

    total_items = 0
    total_in    = 0
    total_out   = 0
    zero_stock  = 0
    low_stock   = 0
    LOW_THRESHOLD = 10
    alerts      = []
    item_out    = {}   # name -> total out
    zone_in     = {}   # zone -> total in
    zone_out    = {}   # zone -> total out
    zone_items  = {}   # zone -> row count
    zone_files  = {}   # zone -> file count
    zone_item_out = {}  # zone -> item -> total out
    zone_low_zero = {}  # zone -> low/zero counters
    latest_files = []
    unreadable  = 0
    latest_mtime = None
    # Log sheet data: IN/OUT operations
    log_in_ops  = []   # list of {'time','qty','item','category','file','zone'}
    log_out_ops = []
    log_item_out = {}  # item_name -> total out from Log

    def to_number(value):
        if value in (None, ''):
            return 0
        try:
            return float(value)
        except Exception:
            try:
                text = str(value).replace(',', '').strip()
                return float(text) if text else 0
            except Exception:
                return 0

    def clean_header(value):
        return re.sub(r'\s+', ' ', str(value or '').strip()).lower()

    def build_header_map(row):
        aliases = {
            'color': {'color', 'colour', 'item', 'item name', 'name', 'الصنف', 'اللون'},
            'type': {'type', 'category', 'item type', 'النوع', 'الفئة'},
            'size': {'size', 'الحجم', 'المقاس'},
            'in': {'in', 'in qty', 'in quantity', 'qty in', 'الوارد'},
            'out': {'out', 'out qty', 'out quantity', 'qty out', 'الصادر'},
            'balance': {'current balance', 'current', 'balance', 'stock', 'الرصيد الحالي', 'الرصيد'},
        }
        header_map = {}
        normalized = [clean_header(c) for c in row]
        for key, names in aliases.items():
            for idx, header in enumerate(normalized):
                if header in names:
                    header_map[key] = idx
                    break
        return header_map

    def get_cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    for zid, fpath in files_to_scan:
        zone_meta = next((z for z in ZONES if z['id'] == zid), None)
        zone_label = zone_meta['name'] if zone_meta else zid
        zone_files[zone_label] = zone_files.get(zone_label, 0) + 1
        try:
            mtime = os.path.getmtime(fpath)
            latest_mtime = mtime if latest_mtime is None else max(latest_mtime, mtime)
            latest_files.append({
                'zone': zone_label,
                'file': os.path.basename(fpath),
                'path': os.path.relpath(fpath, root),
                'updated': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),
                'mtime': mtime,
            })
        except Exception:
            pass

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception:
            unreadable += 1
            continue

        # Read Log sheet for actual IN/OUT operations
        for log_sheet in wb.sheetnames:
            if log_sheet.lower() != 'log':
                continue
            try:
                ws_log = wb[log_sheet]
                log_rows_all = list(ws_log.iter_rows(values_only=True))
                if not log_rows_all:
                    continue
                log_headers = [str(v).strip() if v else '' for v in log_rows_all[0]]
                # Find column indices: Time(0), Type(1), Qty(2), Balance(3), Color(4/name), Size(5), ItemType(6), Category(7)
                # Based on LOG_COL_* constants
                lh = {h.lower(): i for i, h in enumerate(log_headers) if h}
                # Try by position fallback or header name
                ci_type = lh.get('type', lh.get('operation', 1))
                ci_qty  = lh.get('qty', lh.get('quantity', 2))
                ci_color = lh.get('color', lh.get('item', lh.get('name', 4)))
                ci_cat  = lh.get('category', lh.get('item type', 7))
                ci_size = lh.get('size', 5)
                for lrow in log_rows_all[1:]:
                    if not lrow or all(v is None for v in lrow):
                        continue
                    def lget(idx):
                        try: return lrow[idx] if idx < len(lrow) else None
                        except: return None
                    op  = str(lget(ci_type) or '').strip().upper()
                    qty = to_number(lget(ci_qty))
                    color = str(lget(ci_color) or '').strip()
                    cat   = str(lget(ci_cat) or '').strip()
                    size  = str(lget(ci_size) or '').strip()
                    item_label = ' - '.join(p for p in [cat, color, size] if p) or color or '—'
                    if op == 'IN' and qty:
                        log_in_ops.append({'qty': qty, 'item': item_label, 'zone': zone_label})
                    elif op == 'OUT' and qty:
                        log_out_ops.append({'qty': qty, 'item': item_label, 'zone': zone_label})
                        log_item_out[item_label] = log_item_out.get(item_label, 0) + qty
            except Exception as _le:
                pass

        for sheet_name in wb.sheetnames:
            if 'log' in sheet_name.lower():
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = None
            header_map = {}
            for i, row in enumerate(rows):
                candidate = build_header_map(row)
                if 'in' in candidate and 'out' in candidate and ('balance' in candidate or 'color' in candidate):
                    header_idx = i
                    header_map = candidate
                    break

            if header_idx is None:
                continue

            for row in rows[header_idx + 1:]:
                color_val = get_cell(row, header_map.get('color'))
                type_val = get_cell(row, header_map.get('type'))
                size_val = get_cell(row, header_map.get('size'))

                # Fallback to fixed column indices for merged cells
                # (openpyxl read_only=True returns None for non-master merged cells)
                def _fb(val, col_idx):
                    if val is None or str(val).strip() in ('', 'None'):
                        try: return row[col_idx - 1] if col_idx - 1 < len(row) else None
                        except: return None
                    return val
                color_val = _fb(color_val, COL_COLOR)
                type_val  = _fb(type_val,  COL_TYPE)
                size_val  = _fb(size_val,  COL_SIZE)
                # Category column as extra fallback for type
                if type_val is None or str(type_val).strip() in ('', 'None'):
                    try: type_val = row[COL_CATEGORY - 1] if COL_CATEGORY - 1 < len(row) else None
                    except: pass

                in_val = to_number(get_cell(row, header_map.get('in')))
                out_val = to_number(get_cell(row, header_map.get('out')))
                bal_cell = get_cell(row, header_map.get('balance'))
                # Fallback balance to fixed column
                if bal_cell is None or str(bal_cell).strip() in ('', 'None'):
                    try: bal_cell = row[COL_CURRENT - 1] if COL_CURRENT - 1 < len(row) else None
                    except: pass
                bal_val = None if bal_cell in (None, '') else to_number(bal_cell)

                meaningful = any(str(v or '').strip() for v in (color_val, type_val, size_val)) or in_val or out_val or bal_val not in (None, 0)
                if not meaningful:
                    continue
                total_items += 1
                zone_items[zone_label] = zone_items.get(zone_label, 0) + 1

                total_in  += in_val
                total_out += out_val
                zone_in[zone_label]  = zone_in.get(zone_label,  0) + in_val
                zone_out[zone_label] = zone_out.get(zone_label, 0) + out_val

                # Color is the primary identifier; size next, then type/category
                name_parts = [str(v).strip() for v in (color_val, size_val, type_val) if str(v or '').strip()]
                item_name = ' - '.join(name_parts) if name_parts else ''

                if item_name and out_val > 0:
                    item_out[item_name] = item_out.get(item_name, 0) + out_val
                    zone_bucket = zone_item_out.setdefault(zone_label, {})
                    zone_bucket[item_name] = zone_bucket.get(item_name, 0) + out_val

                if bal_val is not None:
                    if bal_val == 0:
                        zero_stock += 1
                        zone_low_zero.setdefault(zone_label, {'low': 0, 'zero': 0})['zero'] += 1
                        alerts.append({'name': item_name or '—', 'sheet': sheet_name,
                                       'balance': 0, 'level': 'danger'})
                    elif bal_val < LOW_THRESHOLD:
                        low_stock += 1
                        zone_low_zero.setdefault(zone_label, {'low': 0, 'zero': 0})['low'] += 1
                        alerts.append({'name': item_name or '—', 'sheet': sheet_name,
                                       'balance': bal_val, 'level': 'warn'})

        wb.close()

    top_items = sorted([{'name': k, 'out': v} for k, v in item_out.items()],
                       key=lambda x: x['out'], reverse=True)[:10]
    zone_consumption = {}
    for zname, items in zone_item_out.items():
        ranked = sorted(items.items(), key=lambda kv: kv[1], reverse=True)
        positive = [kv for kv in ranked if kv[1] > 0]
        zone_consumption[zname] = {
            'top': {'name': ranked[0][0], 'out': round(ranked[0][1], 2)} if ranked else None,
            'lowest': {'name': positive[-1][0], 'out': round(positive[-1][1], 2)} if positive else None,
            'top5': [{'name': name, 'out': round(value, 2)} for name, value in ranked[:5]],
            'moving_items': len(positive),
        }
    most_active_zone = None
    for zname in set(list(zone_in.keys()) + list(zone_out.keys())):
        movement = zone_in.get(zname, 0) + zone_out.get(zname, 0)
        if most_active_zone is None or movement > most_active_zone['movement']:
            most_active_zone = {'zone': zname, 'movement': round(movement, 2)}
    latest_files = sorted(latest_files, key=lambda f: f.get('mtime', 0), reverse=True)[:8]
    high_usage_threshold = max(100, (total_out / max(len(item_out), 1)) * 2) if item_out else 100
    high_usage_items = [
        {'name': item['name'], 'out': round(item['out'], 2)}
        for item in top_items
        if item.get('out', 0) >= high_usage_threshold
    ]
    latest_file_update = datetime.fromtimestamp(latest_mtime).strftime('%Y-%m-%d %H:%M:%S') if latest_mtime else ''
    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    scope = 'All zones' if len(scan_zones) > 1 else (next((z['name'] for z in ZONES if z['id'] == scan_zones[0]), scan_zones[0]) if scan_zones else zone_id)

    # Build log-based top items (from actual Log sheet operations)
    log_top_items = sorted(
        [{'name': k, 'out': round(v, 2)} for k, v in log_item_out.items()],
        key=lambda x: x['out'], reverse=True
    )[:10]
    log_total_in  = round(sum(op['qty'] for op in log_in_ops), 2)
    log_total_out = round(sum(op['qty'] for op in log_out_ops), 2)

    return jsonify({
        'total_items': total_items,
        'total_in':    round(total_in,  2),
        'total_out':   round(total_out, 2),
        'zero_stock':  zero_stock,
        'low_stock':   low_stock,
        'alerts':      alerts[:50],
        'top_items':   [{'name': item['name'], 'out': round(item['out'], 2)} for item in top_items],
        'zone_in':     {k: round(v, 2) for k, v in zone_in.items()},
        'zone_out':    {k: round(v, 2) for k, v in zone_out.items()},
        'zone_items':  zone_items,
        'zone_files':  zone_files,
        'zone_consumption': zone_consumption,
        'zone_low_zero': zone_low_zero,
        'latest_files': [{k: v for k, v in item.items() if k != 'mtime'} for item in latest_files],
        'most_active_zone': most_active_zone,
        'high_usage_items': high_usage_items,
        'high_usage_threshold': round(high_usage_threshold, 2),
        'file_count':  len(files_to_scan),
        'unreadable_files': unreadable,
        'latest_file_update': latest_file_update,
        'generated_at': generated_at,
        'scope': scope,
        'selected_zone': requested_zone if is_super and requested_zone else ('all' if is_super else zone_id),
        'dashboard_zones': [{'id': 'all', 'name': 'All zones'}] + [{'id': z['id'], 'name': z['name'], 'label': z['label']} for z in available_dashboard_zones] if is_super else [],
        # Log sheet data
        'log_top_items': log_top_items,
        'log_total_in':  log_total_in,
        'log_total_out': log_total_out,
        'log_ops_count': len(log_in_ops) + len(log_out_ops),
    })

@dashboard_bp.route('/api/login_log')
@zone_required
def api_login_log():
    """Return login history — dev only."""
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _log_lock:
        entries = _read_login_log()
    return jsonify({
        'entries': list(reversed(entries)),
        'total': len(entries),
        'log_file': _LOGIN_LOG_FILE,
    })

@dashboard_bp.route('/api/alert_count')
@zone_required
def api_alert_count():
    """Quick scan: return number of zero-stock items for the badge."""
    zone_id  = session.get('active_view_zone') or session.get('zone', '')
    is_super = session.get('is_super', False)
    root = get_years_root()
    if not root:
        return jsonify({'zero': 0})
    scan_zones = [z['id'] for z in ZONES if z['id'] not in SUPER_ZONES] if is_super else [zone_id]
    zero = 0
    for zid in scan_zones:
        zone_path = os.path.join(root, zid)
        if not os.path.isdir(zone_path):
            continue
        for rdir, dirs, files in os.walk(zone_path):
            for f in files:
                if not f.lower().endswith(('.xlsx','.xlsm','.xls')):
                    continue
                try:
                    wb = openpyxl.load_workbook(os.path.join(rdir,f), data_only=True, read_only=True)
                    for sname in wb.sheetnames:
                        if 'log' in sname.lower(): continue
                        ws = wb[sname]
                        rows = list(ws.iter_rows(values_only=True))
                        hdr_idx = None
                        headers = []
                        for i,row in enumerate(rows):
                            rv=[str(c).strip() if c else '' for c in row]
                            if 'Current Balance' in rv:
                                hdr_idx=i; headers=rv; break
                        if hdr_idx is None: continue
                        ci_bal = headers.index('Current Balance')
                        for row in rows[hdr_idx+1:]:
                            if all(c is None or str(c).strip()=='' for c in row): continue
                            try:
                                if ci_bal < len(row) and float(row[ci_bal] or 1) == 0:
                                    zero += 1
                            except: pass
                    wb.close()
                except: pass
    return jsonify({'zero': zero})
@dashboard_bp.route('/api/dashboard/excel_status')
@zone_required
def api_dashboard_excel_status():
    root = get_years_root()
    if not root:
        return jsonify({'connected': False, 'message': 'فشل قراءة الملف', 'last_update': None})
    latest_path = None
    latest_mtime = 0
    for rdir, dirs, files in os.walk(root):
        for f in files:
            if f.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                p = os.path.join(rdir, f)
                try:
                    m = os.path.getmtime(p)
                    if m > latest_mtime:
                        latest_mtime = m
                        latest_path = p
                except Exception:
                    pass
    if not latest_path:
        return jsonify({'connected': False, 'message': 'فشل قراءة الملف', 'last_update': None})
    minutes = max(0, int((datetime.now() - datetime.fromtimestamp(latest_mtime)).total_seconds() // 60))
    return jsonify({'connected': True, 'message': 'متصل', 'file': os.path.basename(latest_path), 'minutes_ago': minutes, 'last_update': datetime.fromtimestamp(latest_mtime).strftime('%Y-%m-%d %H:%M:%S')})


@dashboard_bp.route('/api/dashboard/stocktaking')
@zone_required
def api_dashboard_stocktaking():
    """Read 'Stocktaking' sheets from Excel files and return inventory count rows."""
    zone_id = session.get('active_view_zone') or session.get('zone', '')
    is_super = session.get('is_super', False)
    requested_zone = request.args.get('zone', '').strip()
    root = get_years_root()
    if not root:
        return jsonify({'items': [], 'total': 0, 'files_scanned': 0})

    available_zones = [z for z in ZONES if z['id'] not in SUPER_ZONES]
    valid_zone_ids = {z['id'] for z in available_zones}
    if is_super and requested_zone and requested_zone != 'all':
        scan_zones = [requested_zone] if requested_zone in valid_zone_ids else []
    elif is_super:
        scan_zones = [z['id'] for z in available_zones]
    else:
        scan_zones = [zone_id]

    items = []
    files_scanned = 0
    STOCKTAKING_SHEET = 'Stocktaking'

    for zid in scan_zones:
        zone_path = os.path.join(root, zid)
        if not os.path.isdir(zone_path):
            continue
        for fname in os.listdir(zone_path):
            if not fname.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                continue
            fpath = os.path.join(zone_path, fname)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                if STOCKTAKING_SHEET not in wb.sheetnames:
                    wb.close()
                    continue
                ws = wb[STOCKTAKING_SHEET]
                headers = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [str(c).strip() if c is not None else '' for c in row]
                        continue
                    if not any(c for c in row if c is not None):
                        continue
                    row_dict = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
                    name = row_dict.get(COL_TYPE) or row_dict.get('Item') or row_dict.get('Name') or row_dict.get('اسم الصنف') or ''
                    balance = row_dict.get(COL_CURRENT) or row_dict.get('Current Balance') or row_dict.get('Balance') or row_dict.get('الرصيد') or 0
                    category = row_dict.get(COL_CATEGORY) or row_dict.get('Category') or row_dict.get('الفئة') or ''
                    color = row_dict.get(COL_COLOR) or row_dict.get('Color') or row_dict.get('اللون') or ''
                    size = row_dict.get(COL_SIZE) or row_dict.get('Size') or row_dict.get('الحجم') or ''
                    if not name:
                        continue
                    try:
                        balance = float(balance) if balance not in (None, '') else 0
                    except (ValueError, TypeError):
                        balance = 0
                    items.append({
                        'name': str(name),
                        'balance': balance,
                        'category': str(category) if category else '',
                        'color': str(color) if color else '',
                        'size': str(size) if size else '',
                        'zone': zid,
                        'file': fname,
                    })
                files_scanned += 1
                wb.close()
            except Exception:
                pass

    items.sort(key=lambda x: x['name'].lower())
    return jsonify({'items': items, 'total': len(items), 'files_scanned': files_scanned})

