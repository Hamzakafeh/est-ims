"""
EST Inventory System - Full Read/Write
Alestesharia Animal Nutrition
"""
from dotenv import load_dotenv
load_dotenv()
import os
import io
import sys
import json
import re
import sqlite3
import hashlib
import secrets
import warnings
import threading
import shutil
import requests as _requests
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, jsonify, request, session, redirect, url_for, send_file

warnings.filterwarnings('ignore')

# ── Brute-force protection (in-memory, resets on restart) ──────────
from collections import defaultdict
import time as _time

_login_attempts  = defaultdict(list)   # ip -> [timestamp, ...]
_MAX_ATTEMPTS    = 5
_LOCKOUT_SECONDS = 300                 # 5 دقائق

def _get_ip():
    return request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip()

_lockout_until = {}  # ip -> timestamp when lockout expires

def _is_locked_out(ip):
    """Returns True only if IP is actively locked out (enough FAILED attempts, no correct login)."""
    now = _time.time()
    # Remove expired lockout
    if ip in _lockout_until and now >= _lockout_until[ip]:
        _lockout_until.pop(ip, None)
        _login_attempts.pop(ip, None)
    if ip in _lockout_until:
        return True
    _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < _LOCKOUT_SECONDS]
    if len(_login_attempts[ip]) >= _MAX_ATTEMPTS:
        # Set explicit lockout timer
        _lockout_until[ip] = now + _LOCKOUT_SECONDS
        return True
    return False

def _lockout_remaining(ip):
    """Seconds remaining in lockout, 0 if not locked."""
    now = _time.time()
    until = _lockout_until.get(ip, 0)
    return max(0, int(until - now))

def _record_failed_attempt(ip):
    _login_attempts[ip].append(_time.time())

def _clear_attempts(ip):
    _login_attempts.pop(ip, None)
    _lockout_until.pop(ip, None)

_LOGIN_LOG_FILE = os.getenv(
    'LOGIN_LOG_FILE',
    os.path.join(os.getenv('RENDER_DISK_PATH', os.path.dirname(os.path.abspath(__file__))), 'login_log.json')
)
_log_lock = threading.Lock()
_country_cache = {}
VERIFIED_USERS = {'mlo5'}
_active_user_sessions = {}
_active_sessions_lock = threading.Lock()
_ACTIVE_SESSION_TTL = 8 * 60 * 60


def _is_single_login_exempt(username):
    uname = _normalize_username(username)
    privileged = {_normalize_username(v) for v in ZONE_USER_RESTRICTIONS.values() if v}
    return uname in privileged


def _active_session_alive(record):
    return bool(record) and (_time.time() - float(record.get('last_seen', 0))) < _ACTIVE_SESSION_TTL


def _register_active_session(username):
    token = secrets.token_hex(16)
    uname = _normalize_username(username)
    with _active_sessions_lock:
        _active_user_sessions[uname] = {'token': token, 'last_seen': _time.time()}
    session['session_token'] = token


def _clear_active_session(username=None):
    explicit_username = username is not None
    uname = _normalize_username(username or session.get('username', ''))
    token = session.get('session_token')
    if not uname:
        return
    with _active_sessions_lock:
        record = _active_user_sessions.get(uname)
        if explicit_username or not token or (record and record.get('token') == token):
            _active_user_sessions.pop(uname, None)


def _validate_active_session():
    if not session.get('logged_in'):
        return True
    username = session.get('username', '')
    if _is_single_login_exempt(username):
        return True
    uname = _normalize_username(username)
    token = session.get('session_token')
    with _active_sessions_lock:
        record = _active_user_sessions.get(uname)
        if not record or record.get('token') != token:
            return False
        record['last_seen'] = _time.time()
    return True

def _read_login_log():
    try:
        with open(_LOGIN_LOG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []

def _write_login_log(entries):
    log_dir = os.path.dirname(_LOGIN_LOG_FILE)
    if log_dir:
        os.makedirs(log_dir, exist_ok=True)
    with open(_LOGIN_LOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(entries[-500:], f, ensure_ascii=False)  # keep last 500

def _ip_country(ip):
    if not ip or ip.startswith(('127.', '10.', '192.168.', '172.16.', '172.17.', '172.18.', '172.19.', '172.20.', '172.21.', '172.22.', '172.23.', '172.24.', '172.25.', '172.26.', '172.27.', '172.28.', '172.29.', '172.30.', '172.31.')):
        return 'Local / Private'
    if ip in _country_cache:
        return _country_cache[ip]
    try:
        res = _requests.get(f'https://ipapi.co/{ip}/json/', timeout=2)
        data = res.json() if res.ok else {}
        country = data.get('country_name') or data.get('country') or 'Unknown'
    except Exception:
        country = 'Unknown'
    _country_cache[ip] = country
    return country

def _record_login(username, zone_id, zone_label, ip):
    with _log_lock:
        entries = _read_login_log()
        entries.append({
            'username':   username,
            'zone_id':    zone_id,
            'zone_label': zone_label,
            'ip':         ip,
            'country':    _ip_country(ip),
            'time':       datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })
        _write_login_log(entries)


APP_DIR = os.path.dirname(os.path.abspath(__file__))
LEGACY_AUTH_DB_FILE = os.path.join(APP_DIR, 'auth.sqlite3')
AUTH_DB_FILE = os.getenv(
    'AUTH_DB_FILE',
    os.path.join(os.getenv('RENDER_DISK_PATH', APP_DIR), 'auth.sqlite3')
)

DATA_STORE_DIR = os.getenv('RENDER_DISK_PATH', APP_DIR)
CONTACT_MESSAGES_FILE = os.path.join(DATA_STORE_DIR, 'contact_messages.json')
QC_SUBMISSIONS_FILE = os.path.join(DATA_STORE_DIR, 'qc_submissions.json')
QC_UPLOAD_DIR = os.path.join(APP_DIR, 'static', 'qc_uploads')
_data_lock = threading.Lock()


def _read_json_list(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _write_json_list(path, items):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def _next_json_id(items):
    return (max([int(x.get('id', 0) or 0) for x in items] or [0]) + 1)


def _prepare_auth_db_file():
    db_dir = os.path.dirname(AUTH_DB_FILE)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    if os.path.abspath(AUTH_DB_FILE) != os.path.abspath(LEGACY_AUTH_DB_FILE):
        if not os.path.exists(AUTH_DB_FILE) and os.path.exists(LEGACY_AUTH_DB_FILE):
            shutil.copy2(LEGACY_AUTH_DB_FILE, AUTH_DB_FILE)



def _normalize_username(value):
    return str(value or '').strip().lower()


def _normalize_text(value):
    return re.sub(r'\s+', ' ', str(value or '').strip()).lower()


def _hash_secret(value, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        'sha256',
        str(value or '').encode('utf-8'),
        bytes.fromhex(salt),
        200_000,
    ).hex()
    return f'pbkdf2_sha256${salt}${digest}'


def _verify_secret(value, stored):
    if not stored:
        return False
    try:
        algo, salt, digest = str(stored).split('$', 2)
    except ValueError:
        return False
    if algo != 'pbkdf2_sha256':
        return False
    test = hashlib.pbkdf2_hmac(
        'sha256',
        str(value or '').encode('utf-8'),
        bytes.fromhex(salt),
        200_000,
    ).hex()
    return test == digest


def _db_connect():
    conn = sqlite3.connect(AUTH_DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def _init_auth_db():
    _prepare_auth_db_file()
    with _db_connect() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                username TEXT NOT NULL UNIQUE,
                email TEXT NOT NULL,
                phone TEXT NOT NULL,
                job_title TEXT NOT NULL,
                gender TEXT,
                birth_date TEXT,
                privacy_accepted INTEGER NOT NULL DEFAULT 0,
                password_hash TEXT NOT NULL,
                security_question TEXT NOT NULL,
                security_answer_hash TEXT NOT NULL,
                approved INTEGER NOT NULL DEFAULT 0,
                is_admin INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                approved_at TEXT,
                created_by TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS registration_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                username TEXT NOT NULL,
                email TEXT NOT NULL,
                phone TEXT NOT NULL,
                job_title TEXT NOT NULL,
                gender TEXT,
                birth_date TEXT,
                privacy_accepted INTEGER NOT NULL DEFAULT 0,
                password_hash TEXT NOT NULL,
                security_question TEXT NOT NULL,
                security_answer_hash TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL,
                reviewed_at TEXT,
                reviewed_by TEXT,
                review_note TEXT
            )
        """)
        for column_sql in (
            "ALTER TABLE users ADD COLUMN suspended_until TEXT",
            "ALTER TABLE users ADD COLUMN suspended_by TEXT",
            "ALTER TABLE users ADD COLUMN suspended_at TEXT",
            "ALTER TABLE users ADD COLUMN gender TEXT",
            "ALTER TABLE users ADD COLUMN birth_date TEXT",
            "ALTER TABLE users ADD COLUMN privacy_accepted INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE registration_requests ADD COLUMN gender TEXT",
            "ALTER TABLE registration_requests ADD COLUMN birth_date TEXT",
            "ALTER TABLE registration_requests ADD COLUMN privacy_accepted INTEGER NOT NULL DEFAULT 0",
        ):
            try:
                conn.execute(column_sql)
            except sqlite3.OperationalError:
                pass
def _username_in_env(username):
    uname = _normalize_username(username)
    return uname in ENV_USERS


def _get_db_user(username):
    uname = _normalize_username(username)
    if not uname:
        return None
    with _db_connect() as conn:
        return conn.execute(
            "SELECT * FROM users WHERE lower(username) = ?",
            (uname,),
        ).fetchone()


def _username_exists_everywhere(username):
    uname = _normalize_username(username)
    if not uname:
        return False
    if uname in ENV_USERS:
        return True
    with _db_connect() as conn:
        user = conn.execute("SELECT 1 FROM users WHERE lower(username) = ?", (uname,)).fetchone()
        if user:
            return True
        req = conn.execute(
            "SELECT 1 FROM registration_requests WHERE lower(username) = ? AND status = 'pending'",
            (uname,),
        ).fetchone()
        return bool(req)


def _approved_db_user(username):
    user = _get_db_user(username)
    if user and int(user['approved'] or 0) == 1:
        return user
    return None


def _user_suspension_remaining(user):
    if not user:
        return 0
    until = user['suspended_until'] if 'suspended_until' in user.keys() else None
    if not until:
        return 0
    try:
        dt = datetime.strptime(until, '%Y-%m-%d %H:%M:%S')
    except Exception:
        return 0
    remaining = int((dt - datetime.now()).total_seconds())
    return max(0, remaining)


def _pending_request_count():
    with _db_connect() as conn:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM registration_requests WHERE status = 'pending'"
        ).fetchone()
    return int(row['c'] if row else 0)


_init_auth_db()
ENV_USERS = {
    _normalize_username(k): v
    for k, v in {
        os.getenv("USER1"): os.getenv("PASS1"),
        os.getenv("USER2"): os.getenv("PASS2"),
        os.getenv("USER3"): os.getenv("PASS3"),
        os.getenv("USER4"): os.getenv("PASS4"),
    }.items()
    if k
}


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl not installed.")
    sys.exit(1)

app = Flask(__name__, static_folder='static')
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# ── Session Security ───────────────────────────────────────────────
app.config['SESSION_COOKIE_HTTPONLY']  = True   # لا يقرأها JavaScript
app.config['SESSION_COOKIE_SAMESITE']  = 'Lax'  # حماية CSRF جزئية
app.config['PERMANENT_SESSION_LIFETIME'] = 28800 # 8 ساعات

import logging
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s [%(levelname)s] %(message)s'
)
_logger = logging.getLogger('est_ims')

@app.after_request
def add_security_headers(response):
    response.headers['X-Frame-Options']        = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection']       = '1; mode=block'
    response.headers['Referrer-Policy']        = 'strict-origin-when-cross-origin'
    if response.direct_passthrough:
        return response
    if 'text/html' in (response.content_type or ''):
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response

@app.before_request
def enforce_single_active_session():
    if request.endpoint in {'static', 'login_page', 'do_login', 'logout', 'api_lockout_status', 'api_captcha', 'api_register', 'register_page', 'forgot_password_page', 'privacy_page', 'terms_page', 'for_more_page', 'about_page', 'welcome', 'api_password_reset_verify', 'api_password_reset_complete'}:
        return None
    if not _validate_active_session():
        session.clear()
        if request.path.startswith('/api/'):
            return jsonify({'error': 'تم تسجيل الدخول من جهاز آخر، يرجى الدخول من جديد'}), 409
        return redirect(url_for('login_page'))
    return None
_secret_key = os.getenv("SECRET_KEY")
if not _secret_key:
    raise RuntimeError("SECRET_KEY environment variable is not set. Refusing to start.")
app.secret_key = _secret_key

@app.route("/ping")
def ping():
    return {"status": "ok"}, 200

@app.route('/api/lockout_status')
def api_lockout_status():
    """Check if current IP is locked out and how many seconds remain."""
    ip = _get_ip()
    locked = _is_locked_out(ip)
    remaining = _lockout_remaining(ip) if locked else 0
    return jsonify({'locked': locked, 'remaining': remaining})
# ── بيانات الدخول ──────────────────────────────────────────────────
USERS = {
    os.getenv("USER1"): os.getenv("PASS1"),
    os.getenv("USER2"): os.getenv("PASS2"),
    os.getenv("USER3"): os.getenv("PASS3"),
    os.getenv("USER4"): os.getenv("PASS4"),
}
EDIT_PASSWORD = os.getenv("EDIT_PASSWORD")

# ── Zone passwords ──────────────────────────────────────────────────
ZONES = [
    {'id': 'zone1', 'name': 'Zone 1', 'label': 'زون 1',        'icon': '🏭'},
    {'id': 'zone2', 'name': 'Zone 2', 'label': 'زون 2',        'icon': '🏭'},
    {'id': 'zone3', 'name': 'Zone 3', 'label': 'Packaging',    'icon': '🏭'},
    {'id': 'zone4', 'name': 'Zone 4', 'label': 'زون 4',        'icon': '🏭'},
    {'id': 'zone5', 'name': 'Zone 5', 'label': 'زون 5',        'icon': '🏭'},
    {'id': 'qc',    'name': 'QC',     'label': 'Quality Control','icon': '🔬'},
    {'id': 'admin', 'name': 'Admin',  'label': 'Administration', 'icon': '🏢'},
    {'id': 'dev',   'name': 'Dev',    'label': 'Dev',            'icon': '💻'},
]

ZONE_PASSWORDS = {
    'zone1': os.getenv("ZONE1_PASSWORD"),
    'zone2': os.getenv("ZONE2_PASSWORD"),
    'zone3': os.getenv("ZONE3_PASSWORD"),
    'zone4': os.getenv("ZONE4_PASSWORD"),
    'zone5': os.getenv("ZONE5_PASSWORD"),
    'qc':    os.getenv("QC_PASSWORD"),
    'admin': os.getenv("ADMIN_PASSWORD"),
    'dev':   os.getenv("DEV_PASSWORD"),
}

# Zones that can see all 5 zones and switch between them
SUPER_ZONES = {'admin', 'dev'}
# Zones that can use edit mode
EDIT_ZONES  = {'dev'}
# Warehouse zones available in zone switcher for super users
WAREHOUSE_ZONES = ('zone1', 'zone2', 'zone3', 'zone4', 'zone5', 'qc')
ZONE_USER_RESTRICTIONS = {
    'dev': 'mlo5',
    'admin': 'ink',
}
# ───────────────────────────────────────────────────────────────────

from flask import send_from_directory

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(app.static_folder, 'est.ico')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def zone_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login_page'))
        if not session.get('zone'):
            return redirect(url_for('zones_page'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET'])
def login_page():
    if session.get('logged_in'):
        if session.get('zone'):
            next_url = request.args.get('next', url_for('index'))
            return redirect(next_url)
        return redirect(url_for('zones_page'))
    return render_template('login.html')

@app.route('/register')
def register_page():
    if session.get('logged_in'):
        return redirect(url_for('index') if session.get('zone') else url_for('zones_page'))
    return render_template('register.html')

@app.route('/forgot-password')
def forgot_password_page():
    if session.get('logged_in'):
        return redirect(url_for('index') if session.get('zone') else url_for('zones_page'))
    return render_template('forgot_password.html')

@app.route('/api/captcha')
def api_captcha():
    a = secrets.randbelow(8) + 2
    b = secrets.randbelow(8) + 2
    token = secrets.token_hex(8)
    session['register_captcha'] = {'token': token, 'answer': str(a + b)}
    return jsonify({
        'token': token,
        'question': f'{a} + {b} = ?',
    })

@app.route('/api/register', methods=['POST'])
def api_register():
    data = request.get_json(silent=True) or {}
    full_name = str(data.get('full_name', '')).strip()
    username = _normalize_username(data.get('username', ''))
    email = str(data.get('email', '')).strip()
    phone = str(data.get('phone', '')).strip()
    job_title = str(data.get('job_title', '')).strip()
    gender = str(data.get('gender', '')).strip()
    birth_date = str(data.get('birth_date', '')).strip()
    privacy_accepted = bool(data.get('privacy_accepted'))
    password = str(data.get('password', '')).strip()
    confirm_password = str(data.get('confirm_password', '')).strip()
    security_question = str(data.get('security_question', '')).strip()
    security_answer = str(data.get('security_answer', '')).strip()
    captcha_answer = str(data.get('captcha_answer', '')).strip()
    captcha_token = str(data.get('captcha_token', '')).strip()

    required = [full_name, username, email, phone, job_title, gender, birth_date, password, confirm_password, security_question, security_answer, captcha_answer, captcha_token]
    if not all(required):
        return jsonify({'success': False, 'message': 'يرجى تعبئة جميع الحقول'}), 400
    reserved_usernames = {'admin', 'administrator', 'dev', 'developer', 'root', 'superadmin'}
    if username in reserved_usernames:
        return jsonify({'success': False, 'message': 'اسم المستخدم محجوز ولا يمكن التسجيل به'}), 400
    if len(username) < 5:
        return jsonify({'success': False, 'message': 'اسم المستخدم يجب أن يكون 5 أحرف على الأقل'}), 400
    if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return jsonify({'success': False, 'message': 'يرجى إدخال بريد إلكتروني صحيح'}), 400
    if not (re.search(r'[A-Za-z]', password) and re.search(r'\d', password)):
        return jsonify({'success': False, 'message': 'كلمة المرور يجب أن تحتوي أحرفاً وأرقاماً'}), 400
    if not privacy_accepted:
        return jsonify({'success': False, 'message': 'يجب الموافقة على سياسة الخصوصية وشروط الاستخدام'}), 400
    if password != confirm_password:
        return jsonify({'success': False, 'message': 'كلمة المرور وتأكيدها غير متطابقين'}), 400

    captcha = session.get('register_captcha') or {}
    if captcha.get('token') != captcha_token or str(captcha.get('answer', '')).strip() != captcha_answer:
        return jsonify({'success': False, 'message': 'التحقق الأمني غير صحيح'}), 400

    if _username_exists_everywhere(username):
        return jsonify({'success': False, 'message': 'اسم المستخدم مستخدم مسبقاً'}), 409

    password_hash = _hash_secret(password)
    answer_hash = _hash_secret(_normalize_text(security_answer))
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with _db_connect() as conn:
        conn.execute("""
            INSERT INTO registration_requests
            (full_name, username, email, phone, job_title, gender, birth_date, privacy_accepted, password_hash, security_question, security_answer_hash, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)
        """, (
            full_name,
            username,
            email,
            phone,
            job_title,
            gender,
            birth_date,
            1 if privacy_accepted else 0,
            password_hash,
            security_question,
            answer_hash,
            now,
        ))

    session.pop('register_captcha', None)
    return jsonify({
        'success': True,
        'message': 'تم إرسال طلب التسجيل بنجاح. بانتظار موافقة الأدمن.',
    })

@app.route('/api/password_reset/verify', methods=['POST'])
def api_password_reset_verify():
    data = request.get_json(silent=True) or {}
    username = _normalize_username(data.get('username', ''))
    security_question = str(data.get('security_question', '')).strip()
    security_answer = str(data.get('security_answer', '')).strip()

    user = _approved_db_user(username)
    if not user:
        return jsonify({'success': False, 'message': 'المستخدم غير موجود أو غير مفعل'}), 404
    if _normalize_text(user['security_question']) != _normalize_text(security_question):
        return jsonify({'success': False, 'message': 'السؤال الأمني غير صحيح'}), 401
    if not _verify_secret(_normalize_text(security_answer), user['security_answer_hash']):
        return jsonify({'success': False, 'message': 'الجواب الأمني غير صحيح'}), 401

    session['password_reset_username'] = user['username']
    return jsonify({'success': True, 'message': 'تم التحقق بنجاح'})

@app.route('/api/password_reset/complete', methods=['POST'])
def api_password_reset_complete():
    data = request.get_json(silent=True) or {}
    username = _normalize_username(session.get('password_reset_username') or data.get('username', ''))
    new_password = str(data.get('new_password', '')).strip()
    confirm_password = str(data.get('confirm_password', '')).strip()

    if not username:
        return jsonify({'success': False, 'message': 'انتهت جلسة الاسترجاع، أعد التحقق من البيانات'}), 400
    if not new_password or not confirm_password:
        return jsonify({'success': False, 'message': 'يرجى إدخال كلمة المرور الجديدة'}), 400
    if new_password != confirm_password:
        return jsonify({'success': False, 'message': 'كلمة المرور وتأكيدها غير متطابقين'}), 400

    password_hash = _hash_secret(new_password)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _db_connect() as conn:
        cur = conn.execute(
            "UPDATE users SET password_hash = ?, approved_at = COALESCE(approved_at, ?) WHERE lower(username) = ?",
            (password_hash, now, username),
        )
        if cur.rowcount == 0:
            return jsonify({'success': False, 'message': 'تعذر تحديث كلمة المرور'}), 404

    session.pop('password_reset_username', None)
    return jsonify({'success': True, 'message': 'تم تغيير كلمة المرور بنجاح'})

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/help')           # ← برا الدالة، بدون indentation
def help_page():
    return render_template('help.html')

@app.route('/zone3-qr')
@zone_required
def zone3_qr():
    return send_from_directory('static', 'zone3_qr.html')

@app.route('/login', methods=['POST'])
def do_login():
    ip = _get_ip()
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    next_url = data.get('next', '/zones')

    username_key = _normalize_username(username)
    # Check if password is correct FIRST — correct password bypasses lockout
    env_password = ENV_USERS.get(username_key)
    db_user = _approved_db_user(username)
    correct = (
        (env_password is not None and env_password == password)
        or (db_user is not None and _verify_secret(password, db_user['password_hash']))
    )

    if not correct and _is_locked_out(ip):
        remaining = _lockout_remaining(ip)
        mins = remaining // 60
        secs = remaining % 60
        return jsonify({
            'success': False,
            'locked': True,
            'remaining': remaining,
            'message': f'تم تجاوز عدد المحاولات. الرجاء الانتظار {mins}:{secs:02d}'
        }), 429

    if correct:
        if db_user is not None:
            suspended_remaining = _user_suspension_remaining(db_user)
            if suspended_remaining > 0:
                mins = suspended_remaining // 60
                secs = suspended_remaining % 60
                return jsonify({
                    'success': False,
                    'suspended': True,
                    'remaining': suspended_remaining,
                    'message': f'الحساب موقوف مؤقتاً. الوقت المتبقي {mins}:{secs:02d}'
                }), 403
        login_username = db_user['username'] if db_user is not None else username
        if not _is_single_login_exempt(login_username):
            uname = _normalize_username(login_username)
            with _active_sessions_lock:
                record = _active_user_sessions.get(uname)
                if _active_session_alive(record):
                    return jsonify({
                        'success': False,
                        'active_elsewhere': True,
                        'message': 'هذا المستخدم مسجل دخوله من جهاز آخر حالياً'
                    }), 409
                _active_user_sessions.pop(uname, None)
        _clear_attempts(ip)
        session['logged_in'] = True
        session['username']  = login_username
        if not _is_single_login_exempt(login_username):
            _register_active_session(login_username)
        # لو في صفحة QR scan منتظرة (بدون zone)، روّح عليها مباشرة
        qr_next = session.pop('qr_next', None)
        if qr_next and qr_next.startswith('/'):
            return jsonify({'success': True, 'redirect': qr_next})
        if not session.get('next_after_zone'):
            session['next_after_zone'] = next_url if next_url.startswith('/') else '/zones'
        session.pop('zone', None)
        return jsonify({'success': True, 'redirect': '/zones'})
    _record_failed_attempt(ip)
    return jsonify({'success': False, 'message': 'Incorrect username or password'}), 401

# ── ZONES PAGE ──────────────────────────────────────────────────────
@app.route('/zones')
@login_required
def zones_page():
    if session.get('zone'):
        return redirect(url_for('index'))
    return render_template('zones.html',
                           username=session.get('username', ''),
                           zones=ZONES)

@app.route('/api/zone_login', methods=['POST'])
@login_required
def api_zone_login():
    ip = _get_ip()
    data     = request.get_json(silent=True) or {}
    zone_id  = data.get('zone_id', '').strip()
    password = data.get('password', '').strip()

    zone = next((z for z in ZONES if z['id'] == zone_id), None)
    if not zone:
        return jsonify({'success': False, 'message': 'زون غير معروف'}), 400

    expected = ZONE_PASSWORDS.get(zone_id)
    correct = (expected and password == expected)

    if not correct and _is_locked_out(ip):
        remaining = _lockout_remaining(ip)
        mins = remaining // 60
        secs = remaining % 60
        return jsonify({'success': False, 'locked': True, 'remaining': remaining,
                        'message': f'تم تجاوز عدد المحاولات. انتظر {mins}:{secs:02d}'}), 429

    if not correct:
        _record_failed_attempt(ip)
        return jsonify({'success': False, 'message': 'Incorrect password'}), 401

    restricted_username = ZONE_USER_RESTRICTIONS.get(zone_id)
    current_username = str(session.get('username', '')).strip().lower()
    allowed_usernames = []
    if restricted_username:
        allowed_usernames.append(restricted_username.lower())
        if zone_id == 'admin':
            dev_user = ZONE_USER_RESTRICTIONS.get('dev')
            if dev_user:
                allowed_usernames.append(dev_user.lower())
    if restricted_username and current_username not in allowed_usernames:
        _record_failed_attempt(ip)
        return jsonify({'success': False, 'not_allowed': True, 'message': 'Access Denied'}), 403

    _clear_attempts(ip)
    session['zone']        = zone_id
    session['zone_name']   = zone['name']
    session['zone_label']  = zone['label']
    session['can_edit']    = zone_id in EDIT_ZONES
    session['is_super']    = zone_id in SUPER_ZONES
    session['login_time']  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if zone_id == 'qc':
        role = str(data.get('qc_role', '')).strip()
        session['qc_role'] = role if role in {'qc', 'labeling'} else 'qc'
    _record_login(session.get('username',''), zone_id, zone['label'], ip)
    # إذا في صفحة مطلوبة بعد اللوغن (مثل /scan)، حوّل عليها
    next_url = session.pop('next_after_zone', '/index')
    if zone_id == 'qc':
        next_url = '/qc-workflow'
    return jsonify({'success': True, 'redirect': next_url})

@app.route('/api/zone_access_check', methods=['POST'])
@login_required
def api_zone_access_check():
    """Check if the current user is allowed to even attempt a zone login."""
    data = request.get_json(silent=True) or {}
    zone_id = data.get('zone_id', '').strip()
    restricted_username = ZONE_USER_RESTRICTIONS.get(zone_id)
    current_username = str(session.get('username', '')).strip().lower()
    allowed_usernames = []
    if restricted_username:
        allowed_usernames.append(restricted_username.lower())
        if zone_id == 'admin':
            dev_user = ZONE_USER_RESTRICTIONS.get('dev')
            if dev_user:
                allowed_usernames.append(dev_user.lower())
    if restricted_username and current_username not in allowed_usernames:
        return jsonify({'allowed': False}), 200
    return jsonify({'allowed': True}), 200

@app.route('/api/switch_zone', methods=['POST'])
@zone_required
def api_switch_zone():
    """Super zones (admin/dev) can switch to any of the 5 warehouse zones."""
    if not session.get('is_super'):
        return jsonify({'success': False, 'message': 'غير مصرح'}), 403
    data    = request.get_json(silent=True) or {}
    zone_id = data.get('zone_id', '').strip()
    if zone_id not in ('zone1','zone2','zone3','zone4','zone5','qc'):
        return jsonify({'success': False, 'message': 'زون غير صحيح'}), 400
    zone = next((z for z in ZONES if z['id'] == zone_id), None)
    session['active_view_zone']      = zone_id
    session['active_view_zone_name'] = zone['name'] if zone else zone_id
    return jsonify({'success': True})

@app.route('/api/session_info')
@zone_required
def api_session_info():
    return jsonify({
        'username':         session.get('username', ''),
        'zone':             session.get('zone', ''),
        'zone_name':        session.get('zone_name', ''),
        'zone_label':       session.get('zone_label', ''),
        'can_edit':         session.get('can_edit', False),
        'is_super':         session.get('is_super', False),
        'active_view_zone': session.get('active_view_zone', session.get('zone', '')),
        'zones':            [z for z in ZONES if z['id'] not in SUPER_ZONES],
    })

@app.route('/api/profile')
@zone_required
def api_profile():
    """Return the current user's profile summary for the in-app profile modal."""
    username = session.get('username', '')
    zone_id = session.get('zone', '')
    active_view_zone = session.get('active_view_zone', zone_id)
    login_time = session.get('login_time', '')
    login_duration_seconds = None

    if login_time:
        try:
            login_dt = datetime.strptime(login_time, '%Y-%m-%d %H:%M:%S')
            login_duration_seconds = max(0, int((datetime.now() - login_dt).total_seconds()))
        except Exception:
            login_duration_seconds = None

    active_zone = next((z for z in ZONES if z['id'] == active_view_zone), None)
    allowed_zones = [z for z in ZONES if z['id'] not in SUPER_ZONES] if session.get('is_super') else [
        z for z in ZONES if z['id'] == zone_id
    ]

    with _log_lock:
        entries = _read_login_log()

    recent_logins = [
        e for e in reversed(entries)
        if str(e.get('username', '')).lower() == str(username).lower()
    ][:8]

    return jsonify({
        'username': username,
        'is_verified': str(username).lower() in VERIFIED_USERS,
        'zone': zone_id,
        'zone_name': session.get('zone_name', ''),
        'zone_label': session.get('zone_label', ''),
        'active_view_zone': active_view_zone,
        'active_view_zone_label': active_zone['label'] if active_zone else session.get('zone_label', ''),
        'is_super': bool(session.get('is_super', False)),
        'can_edit': bool(session.get('can_edit', False)),
        'login_time': login_time,
        'login_duration_seconds': login_duration_seconds,
        'permissions': {
            'can_edit': bool(session.get('can_edit', False)),
            'can_export': True,
            'can_print': True,
            'can_reports': True,
            'can_view_all_zones': bool(session.get('is_super', False)),
            'can_switch_zones': bool(session.get('is_super', False)),
        },
        'allowed_zones': allowed_zones,
        'recent_logins': recent_logins,
    })

@app.route('/api/verify_edit_password', methods=['POST'])
@zone_required
def verify_edit_password():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'message': 'غير مصرح لهذا الزون'}), 403
    data = request.get_json(silent=True) or {}
    password = data.get('password', '')
    if EDIT_PASSWORD and password == EDIT_PASSWORD:
        return jsonify({'success': True})
    return jsonify({'success': False}), 401

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('welcome'))

@app.route('/logout_zone')
@login_required
def logout_zone():
    """Go back to zone selection without full logout."""
    session.pop('zone', None)
    session.pop('zone_name', None)
    session.pop('zone_label', None)
    session.pop('can_edit', None)
    session.pop('is_super', None)
    session.pop('active_view_zone', None)
    session.pop('active_view_zone_name', None)
    return redirect(url_for('zones_page'))

MONTH_ORDER = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4,
    'May': 5, 'June': 6, 'July': 7, 'August': 8,
    'September': 9, 'October': 10, 'November': 11, 'December': 12
}
MONTH_AR = {
    'January': 'يناير', 'February': 'فبراير', 'March': 'مارس', 'April': 'أبريل',
    'May': 'مايو', 'June': 'يونيو', 'July': 'يوليو', 'August': 'أغسطس',
    'September': 'سبتمبر', 'October': 'أكتوبر', 'November': 'نوفمبر', 'December': 'ديسمبر'
}

# ── Column mapping (based on VBA: D=4,E=5,F=6,G=7,I=9,J=10,L=12,M=13) ──
# Sheet columns (1-based like Excel):
COL_DATE     = 1   # A - Date
COL_CATEGORY = 4   # D - Category (merged)
COL_TYPE     = 5   # E - Type (merged)
COL_COLOR    = 6   # F - Color
COL_SIZE     = 7   # G - Size
COL_BASIC    = 9   # I - Basic balance
COL_CURRENT  = 10  # J - Current balance
COL_IN       = 12  # L - IN
COL_OUT      = 13  # M - OUT
DATA_START_ROW = 7 # Data starts at row 7

# Log sheet columns
LOG_COL_TIME     = 1
LOG_COL_TYPE     = 2
LOG_COL_QTY      = 3
LOG_COL_BALANCE  = 4
LOG_COL_COLOR    = 5
LOG_COL_SIZE     = 6
LOG_COL_ITEMTYPE = 7
LOG_COL_CATEGORY = 8

def get_years_root():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(script_dir, 'zones'),
        os.path.join(os.path.dirname(script_dir), 'zones'),
    ]
    for c in candidates:
        if os.path.isdir(c):
            return c
    return None

def _validate_filepath(filepath, zone_id=None):
    """
    تحقق أن المسار داخل مجلد zones فقط (يمنع Path Traversal).
    لو zone_id محدد، يتحقق كذلك أن الملف داخل زون المستخدم.
    يرجع True لو المسار آمن، False لو فيه مشكلة.
    """
    root = get_years_root()
    if not root:
        return False
    try:
        real_root = os.path.realpath(root)
        real_file = os.path.realpath(filepath)
        if not real_file.startswith(real_root + os.sep):
            return False
        if zone_id:
            zone_root = os.path.realpath(os.path.join(root, zone_id))
            if not real_file.startswith(zone_root + os.sep):
                return False
        return True
    except Exception:
        return False

def get_base_path(year=None, zone_id=None):
    """Return path to a year folder inside a zone. zones/zone1/2026/"""
    root = get_years_root()
    if not root:
        return None
    if zone_id:
        zone_folder = os.path.join(root, zone_id)
    else:
        zone_folders = sorted([d for d in os.listdir(root) if os.path.isdir(os.path.join(root, d)) and d.startswith('zone')])
        if not zone_folders:
            return None
        zone_folder = os.path.join(root, zone_folders[0])
    if not os.path.isdir(zone_folder):
        return None
    years = sorted([d for d in os.listdir(zone_folder) if os.path.isdir(os.path.join(zone_folder, d)) and d.isdigit()])
    if not years:
        return None
    target = str(year) if year else years[0]
    candidate = os.path.join(zone_folder, target)
    return candidate if os.path.isdir(candidate) else None

def get_available_years():
    """Return sorted list of available year strings scanned across all zones."""
    root = get_years_root()
    if not root:
        return []
    years = set()
    for zone_folder in os.listdir(root):
        zone_path = os.path.join(root, zone_folder)
        if not os.path.isdir(zone_path):
            continue
        for d in os.listdir(zone_path):
            if os.path.isdir(os.path.join(zone_path, d)) and d.isdigit():
                years.add(d)
    return sorted(years)

def get_structure(year=None, zone_id=None):
    """Return file structure scoped to a zone: {year: {month: {fname: fpath}}}
    Disk layout: zones/zone1/2026/01-January/Other+.xlsm
    """
    root = get_years_root()
    if not root:
        return {}
    available_years = get_available_years()
    if not available_years:
        return {}
    target_years = [str(year)] if year else available_years
    if zone_id:
        zone_folders = [zone_id]
    else:
        zone_folders = sorted([d for d in os.listdir(root) if os.path.isdir(os.path.join(root, d)) and d.startswith('zone')])
    result = {}
    for yr in target_years:
        result[yr] = {}
        for zf in zone_folders:
            year_path = os.path.join(root, zf, yr)
            if not os.path.isdir(year_path):
                continue
            for month_folder in sorted(os.listdir(year_path), key=lambda m: MONTH_ORDER.get(m.split('-')[-1] if '-' in m else m, 99)):
                month_path = os.path.join(year_path, month_folder)
                if not os.path.isdir(month_path):
                    continue
                month_name = month_folder.split('-')[-1] if '-' in month_folder else month_folder
                files = {}
                for fname in ['Other+', 'Sacks']:
                    fpath = os.path.join(month_path, f'{fname}.xlsm')
                    if os.path.exists(fpath):
                        files[fname] = fpath
                if files:
                    if month_name not in result[yr]:
                        result[yr][month_name] = {}
                    result[yr][month_name].update(files)
    return result

def read_sheet_data(filepath, sheet_name):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None, []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if sheet_name == 'Log':
            if not rows:
                wb.close(); return None, []
            col_indices = [(ci, str(val).strip()) for ci, val in enumerate(rows[0])
                           if val is not None and str(val).strip()]
            if not col_indices:
                wb.close(); return None, []
            data_rows = []
            for row in rows[1:]:
                if not any(ci < len(row) and row[ci] is not None and str(row[ci]).strip() not in ('','None')
                           for ci,_ in col_indices):
                    continue
                rd = {}
                for ci, col_name in col_indices:
                    val = row[ci] if ci < len(row) else None
                    if isinstance(val, datetime): val = val.strftime('%Y-%m-%d %H:%M:%S')
                    elif val is not None:
                        val = str(val) if not isinstance(val,(int,float)) else val
                        if isinstance(val,str) and val.strip()=='None': val=None
                    rd[col_name] = val
                data_rows.append(rd)
            wb.close()
            return [h for _,h in col_indices], data_rows

        if sheet_name == 'Stocktaking':
            data_rows = []
            for row in rows:
                non_empty = [(ci,val) for ci,val in enumerate(row)
                             if val is not None and str(val).strip() not in ('','None')]
                if not non_empty: continue
                rd = {}
                for ci,val in non_empty:
                    col_label = f'Col {ci+1}'
                    if isinstance(val,datetime): val=val.strftime('%Y-%m-%d')
                    elif val is not None: val=str(val) if not isinstance(val,(int,float)) else val
                    rd[col_label]=val
                data_rows.append(rd)
            if not data_rows:
                wb.close(); return None,[]
            all_cols,seen=[],set()
            for r in data_rows:
                for k in r:
                    if k not in seen: all_cols.append(k); seen.add(k)
            wb.close()
            return all_cols, data_rows

        # Main inventory sheet
        header_idx = None
        for i, row in enumerate(rows):
            rv = [str(v).strip() if v else '' for v in row]
            if 'Date' in rv or 'التاريخ' in rv:
                header_idx = i; break
        if header_idx is None:
            wb.close(); return None,[]

        col_indices = [(ci,str(val).strip()) for ci,val in enumerate(rows[header_idx])
                       if val is not None and str(val).strip()]
        data_rows = []
        for row_idx, row in enumerate(rows[header_idx+1:], start=header_idx+2):
            if not any(ci<len(row) and row[ci] is not None and str(row[ci]).strip() not in ('','None')
                       for ci,_ in col_indices):
                continue
            rd={}
            for ci,col_name in col_indices:
                val=row[ci] if ci<len(row) else None
                if isinstance(val,datetime): val=val.strftime('%Y-%m-%d')
                elif val is not None:
                    val=str(val) if not isinstance(val,(int,float)) else val
                    if isinstance(val,str) and val.strip()=='None': val=None
                rd[col_name]=val
            # Store the actual Excel row number for editing
            rd['__row__'] = row_idx
            data_rows.append(rd)
        wb.close()
        headers = [h for _,h in col_indices]
        return headers, data_rows
    except Exception as _exc:
        _logger.error('read_sheet_data error — %s: %s', filepath, _exc)
        return None, []

# ═══════════════════════════════════════════════════════════════════
#  WRITE LOGIC  —  Replicates the VBA Worksheet_Change macro exactly
# ═══════════════════════════════════════════════════════════════════

def _get_cell_val(ws, row, col):
    """Get value from a cell, handling merged cells by reading the top-left."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    # If it's part of a merged range, get the master cell value
    for merge in ws.merged_cells.ranges:
        if cell.coordinate in merge:
            master = ws.cell(row=merge.min_row, column=merge.min_col)
            return master.value
    return None

def _find_last_balance(ws, target_row, color_value):
    """
    Walk backwards from target_row-1 looking for a row where Color matches.
    Falls back to Basic Balance of the target row if no prior match found.
    Returns (last_balance, found).
    """
    for i in range(target_row - 1, DATA_START_ROW - 1, -1):
        cell_color = ws.cell(row=i, column=COL_COLOR).value
        if cell_color == color_value:
            balance = ws.cell(row=i, column=COL_CURRENT).value
            try:
                return float(balance or 0), True
            except:
                return 0.0, True
    # No previous row with same color — use Basic Balance of THIS row
    basic = ws.cell(row=target_row, column=COL_BASIC).value
    try:
        return float(basic or 0), False
    except:
        return 0.0, False

def _append_log(ws_log, operation, qty, balance, color, size, item_type, category):
    """Append a row to the Log sheet."""
    lr = 1
    for row in ws_log.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value is not None:
                lr = cell.row
    lr += 1
    ws_log.cell(row=lr, column=LOG_COL_TIME).value     = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_log.cell(row=lr, column=LOG_COL_TYPE).value     = operation
    ws_log.cell(row=lr, column=LOG_COL_QTY).value      = qty
    ws_log.cell(row=lr, column=LOG_COL_BALANCE).value  = balance
    ws_log.cell(row=lr, column=LOG_COL_COLOR).value    = color
    ws_log.cell(row=lr, column=LOG_COL_SIZE).value     = size
    ws_log.cell(row=lr, column=LOG_COL_ITEMTYPE).value = item_type
    ws_log.cell(row=lr, column=LOG_COL_CATEGORY).value = category

# ── Flask routes ────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════
#  Stocktaking auto-recalculation
#  Parses LOOKUP(2,1/(Sheet!Col:Col=Val),Sheet!J:J) formulas and
#  computes the result in Python, then writes it back to the cell.
# ══════════════════════════════════════════════════════════════════
_LOOKUP_RE = re.compile(
    r'LOOKUP\s*\(\s*2\s*,\s*1\s*/\s*\(\s*(\w+)!\s*([A-Z]+)\s*:\s*\2\s*=\s*(.*?)\s*\)\s*,\s*\1!\s*([A-Z]+)\s*:\s*\4\s*\)',
    re.IGNORECASE
)

def _col_letter_to_idx(letter):
    """'A'->0, 'B'->1, ... 'J'->9  (0-based for row tuple indexing)"""
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1  # 0-based

def _recalc_stocktaking(wb):
    """
    Find the Stocktaking sheet in a workbook, parse every LOOKUP formula,
    evaluate it by scanning the referenced sheet, and write the computed
    value back.  Modifies wb in-place (caller must save).
    """
    if 'Stocktaking' not in wb.sheetnames:
        return

    ws_st = wb['Stocktaking']

    # Cache sheet rows for performance
    _sheet_rows_cache = {}
    def _get_rows(sheet_name):
        if sheet_name not in _sheet_rows_cache:
            if sheet_name in wb.sheetnames:
                _sheet_rows_cache[sheet_name] = list(
                    wb[sheet_name].iter_rows(min_row=DATA_START_ROW, values_only=True)
                )
            else:
                _sheet_rows_cache[sheet_name] = []
        return _sheet_rows_cache[sheet_name]

    for row in ws_st.iter_rows():
        for cell in row:
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                continue
            m = _LOOKUP_RE.search(cell.value)
            if not m:
                continue

            ref_sheet   = m.group(1)
            filter_col  = _col_letter_to_idx(m.group(2))   # e.g. F -> 5
            raw_val     = m.group(3).strip().strip('"')     # filter value
            result_col  = _col_letter_to_idx(m.group(4))   # e.g. J -> 9

            # Auto-cast numeric filter value
            try:
                filter_val = float(raw_val) if '.' in raw_val else int(raw_val)
            except ValueError:
                filter_val = raw_val  # keep as string

            # LOOKUP(2,1/(...)) = last matching row (walk forward, keep updating)
            last_result = None
            for data_row in _get_rows(ref_sheet):
                try:
                    cell_filter = data_row[filter_col]
                    # Normalize comparison: strip strings, cast numbers
                    if isinstance(filter_val, str):
                        match = (cell_filter is not None and
                                 str(cell_filter).strip() == filter_val)
                    else:
                        try:
                            match = float(cell_filter) == filter_val
                        except (TypeError, ValueError):
                            match = False

                    if match:
                        v = data_row[result_col]
                        if v is not None:
                            try:
                                last_result = float(v)
                            except (TypeError, ValueError):
                                last_result = v
                except IndexError:
                    continue

            # Write computed value (keep formula string intact, just overwrite value)
            cell.value = last_result if last_result is not None else cell.value

@app.route('/')
def welcome():
    return render_template('welcome.html')

@app.route('/index')
@zone_required
def index():
    structure = get_structure()
    available_years = get_available_years()
    zone        = session.get('zone', '')
    is_super    = session.get('is_super', False)
    can_edit    = session.get('can_edit', False)
    zone_label  = session.get('zone_label', '')
    username    = session.get('username', '')
    return render_template('index.html',
                           structure=structure,
                           available_years=available_years,
                           base_path=get_base_path() or 'Not found',
                           month_ar=MONTH_AR,
                           zone=zone,
                           zone_label=zone_label,
                           is_super=is_super,
                           is_dev=(zone == 'dev'),
                           can_edit=can_edit,
                           username=username,
                           login_time=session.get('login_time',''))

@app.route('/api/structure')
@zone_required
def api_structure():
    zone     = session.get('zone', '')
    is_super = session.get('is_super', False)
    # Super zones can request a specific view zone
    if is_super:
        view_zone = request.args.get('zone') or session.get('active_view_zone', 'zone1')
    else:
        view_zone = zone
    return jsonify(get_filtered_structure(view_zone))

@app.route('/api/years')
@zone_required
def api_years():
    return jsonify({'years': get_available_years()})

def get_filtered_structure(zone_id):
    """Return structure for a specific zone directly from disk."""
    return get_structure(zone_id=zone_id)

@app.route('/api/sheets')
@zone_required
def api_sheets():
    filepath = request.args.get('path')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'error': 'Access denied'}), 403
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames; wb.close()
        return jsonify({'sheets': sheets})
    except Exception:
        return jsonify({'error': 'Failed to read file'}), 500

@app.route('/api/data')
@zone_required
def api_data():
    filepath = request.args.get('path')
    sheet    = request.args.get('sheet','')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'error': 'Access denied'}), 403
    headers, rows = read_sheet_data(filepath, sheet)
    if headers is None:
        return jsonify({'headers':[],'rows':[],'count':0})
    return jsonify({'headers': headers, 'rows': rows, 'count': len(rows)})


# ══════════════════════════════════════════════════════════════════
#  NEW: /api/transaction  —  IN or OUT on the main inventory sheet
#
#  Body (JSON):
#    filepath  : full path to the .xlsm file
#    sheet     : sheet name (e.g. "Sheet1")
#    row       : Excel row number (integer, from __row__ field)
#    operation : "IN" or "OUT"
#    qty       : positive number
# ══════════════════════════════════════════════════════════════════
@app.route('/api/transaction', methods=['POST'])
@zone_required
def api_transaction():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح — يجب تفعيل وضع التعديل'}), 403

    data = request.get_json(silent=True) or {}
    filepath  = data.get('filepath', '')
    sheet     = data.get('sheet', '')
    row       = data.get('row')
    operation = data.get('operation', '').upper()
    qty_raw   = data.get('qty')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    if operation not in ('IN', 'OUT'):
        return jsonify({'success': False, 'error': 'operation must be IN or OUT'}), 400
    try:
        row = int(row)
        qty = float(qty_raw)
        if qty < 0:
            raise ValueError()
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid row or qty'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)

        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404

        ws      = wb[sheet]
        ws_log  = wb['Log'] if 'Log' in wb.sheetnames else None

        color_value    = _get_cell_val(ws, row, COL_COLOR)
        size_value     = _get_cell_val(ws, row, COL_SIZE)
        type_value     = _get_cell_val(ws, row, COL_TYPE)
        category_value = _get_cell_val(ws, row, COL_CATEGORY)
        basic_balance  = ws.cell(row=row, column=COL_BASIC).value or 0

        if not color_value or str(color_value).strip() in ('', 'None', 'null'):
            wb.close()
            return jsonify({'success': False,
                            'error': 'يجب تحديد اللون (Color) أولاً قبل إجراء أي عملية'}), 400

        last_balance, found = _find_last_balance(ws, row, color_value)
        if not found:
            try:
                last_balance = float(basic_balance)
            except Exception:
                last_balance = 0.0

        if operation == 'IN':
            new_balance = last_balance + qty
        else:
            new_balance = last_balance - qty

        ws.cell(row=row, column=COL_CURRENT).value = new_balance

        if ws_log:
            _append_log(ws_log, operation, qty, new_balance,
                        color_value, size_value, type_value, category_value)

        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()

        return jsonify({
            'success':     True,
            'new_balance': new_balance,
            'operation':   operation,
            'qty':         qty,
            'color':       color_value,
            'size':        size_value,
        })

    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء تنفيذ العملية'}), 500


# ══════════════════════════════════════════════════════════════════
#  NEW: /api/update_cell  —  Edit any plain cell directly
#
#  Body (JSON):
#    filepath : full path to the .xlsm file
#    sheet    : sheet name
#    row      : Excel row number
#    col_name : column header name (e.g. "Date", "Color")
#    value    : new value (string; numbers auto-cast)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/update_cell', methods=['POST'])
@zone_required
def api_update_cell():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403

    data      = request.get_json(silent=True) or {}
    filepath  = data.get('filepath', '')
    sheet     = data.get('sheet', '')
    row       = data.get('row')
    col_name  = data.get('col_name', '')
    value     = data.get('value', '')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    try:
        row = int(row)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid row'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404
        ws = wb[sheet]

        header_row_idx = None
        for i, row_cells in enumerate(ws.iter_rows(values_only=True), start=1):
            rv = [str(v).strip() if v else '' for v in row_cells]
            if 'Date' in rv or 'التاريخ' in rv:
                header_row_idx = i; break

        if header_row_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': 'Header row not found'}), 400

        col_idx = None
        for ci, cell in enumerate(ws[header_row_idx], start=1):
            if cell.value and str(cell.value).strip() == col_name:
                col_idx = ci; break

        if col_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': f'Column "{col_name}" not found'}), 400

        cast_value = value
        try:
            if '.' in str(value):
                cast_value = float(value)
            else:
                cast_value = int(value)
        except Exception:
            cast_value = value if value != '' else None

        ws.cell(row=row, column=col_idx).value = cast_value
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()

        return jsonify({'success': True, 'row': row, 'col': col_name, 'value': cast_value})

    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء تحديث الخلية'}), 500


# ══════════════════════════════════════════════════════════════════
#  /api/color_balance  —  Get last current balance for a Color
#
#  Query params: path, sheet, color, before_row (optional)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/color_balance')
@zone_required
def api_color_balance():
    filepath   = request.args.get('path', '')
    sheet      = request.args.get('sheet', '')
    color      = request.args.get('color', '')
    before_row = request.args.get('before_row', None)

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if not color or color.strip().lower() in ('', 'null', 'none'):
        return jsonify({'balance': None, 'found': False})

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'error': 'Access denied'}), 403

    try:
        before_row = int(before_row) if before_row else None
    except Exception:
        before_row = None

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'balance': None, 'found': False})
        ws = wb[sheet]

        last_balance = None
        last_row     = None
        rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))
        for idx, row_vals in enumerate(rows):
            actual_row = DATA_START_ROW + idx
            if before_row and actual_row >= before_row:
                break
            try:
                cell_color   = row_vals[COL_COLOR - 1]
                cell_current = row_vals[COL_CURRENT - 1]
            except IndexError:
                continue
            if cell_color and str(cell_color).strip() == color.strip():
                try:
                    val = float(cell_current) if cell_current is not None else None
                    if val is not None:
                        last_balance = val
                        last_row     = actual_row
                except Exception:
                    pass

        wb.close()
        return jsonify({'balance': last_balance, 'found': last_balance is not None, 'row': last_row})

    except Exception:
        return jsonify({'error': 'حدث خطأ أثناء قراءة الرصيد'}), 500


# ══════════════════════════════════════════════════════════════════
#  /api/set_opening_balance  —  Write Basic + Current in one call
#
#  Body: filepath, sheet, row, balance (number)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/set_opening_balance', methods=['POST'])
@zone_required
def api_set_opening_balance():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403

    data     = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet    = data.get('sheet', '')
    row      = data.get('row')
    balance  = data.get('balance')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    try:
        row     = int(row)
        balance = float(balance)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid row or balance'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        ws = wb[sheet]
        ws.cell(row=row, column=COL_BASIC).value   = balance
        ws.cell(row=row, column=COL_CURRENT).value = balance
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()
        return jsonify({'success': True, 'balance': balance})
    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء تعيين الرصيد الافتتاحي'}), 500


# ══════════════════════════════════════════════════════════════════
#  NEW: /api/add_row  —  Add a new item row to the sheet
#
#  Body (JSON):
#    filepath : full path to the .xlsm file
#    sheet    : sheet name
#    fields   : dict of { col_name: value, ... }
# ══════════════════════════════════════════════════════════════════
def _parse_dv_formula(formula1):
    """Parse a Data Validation formula1 string into a clean list of values."""
    if not formula1:
        return []
    s = formula1.strip().strip('"')
    items = [v.strip() for v in s.split(',') if v.strip() and v.strip().lower() not in ('null', 'none', '')]
    return items

def _col_letter_to_index(letter):
    """Convert Excel column letter(s) to 1-based index. E.g. 'F' -> 6."""
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result

@app.route('/api/options')
@zone_required
def api_options():
    """Read Data Validation lists from the Excel sheet for Color, Type, Size, Category."""
    filepath = request.args.get('path', '')
    sheet    = request.args.get('sheet', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'error': 'Access denied'}), 403

    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'colors': [], 'types': [], 'sizes': [], 'categories': []})
        ws = wb[sheet]

        header_col_map = {}
        for ci, cell in enumerate(ws[6], start=1):
            if cell.value and str(cell.value).strip():
                header_col_map[ci] = str(cell.value).strip()

        name_to_col = {v: k for k, v in header_col_map.items()}

        col_color    = name_to_col.get('Color',    COL_COLOR)
        col_type     = name_to_col.get('Type',     COL_TYPE)
        col_size     = name_to_col.get('Size',     COL_SIZE)
        col_category = name_to_col.get('Category', COL_CATEGORY)

        options = {'colors': [], 'types': [], 'sizes': [], 'categories': []}
        col_target = {
            col_color:    'colors',
            col_type:     'types',
            col_size:     'sizes',
            col_category: 'categories',
        }

        for dv in ws.data_validations.dataValidation:
            if dv.type != 'list' or not dv.formula1:
                continue
            try:
                first_ref = str(dv.sqref).split()[0]
                col_letters = ''.join(c for c in first_ref.split(':')[0] if c.isalpha())
                ci = _col_letter_to_index(col_letters)
            except Exception:
                continue

            key = col_target.get(ci)
            if key:
                vals = _parse_dv_formula(dv.formula1)
                options[key] = vals

        wb.close()
        return jsonify(options)
    except Exception:
        return jsonify({'error': 'حدث خطأ أثناء قراءة الخيارات'}), 500


@app.route('/api/clear_row', methods=['POST'])
@zone_required
def api_clear_row():
    """Clear all data cells in a given Excel row (does NOT delete the row itself)."""
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403

    data     = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet    = data.get('sheet', '')
    row      = data.get('row')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    try:
        row = int(row)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid row'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        ws = wb[sheet]
        for col in range(1, 14):
            ws.cell(row=row, column=col).value = None
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()
        return jsonify({'success': True})
    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء مسح الصف'}), 500


@app.route('/api/add_row', methods=['POST'])
@zone_required
def api_add_row():
    if not session.get('can_edit'):
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403

    data     = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet    = data.get('sheet', '')
    fields   = data.get('fields', {})

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    if not fields:
        return jsonify({'success': False, 'error': 'No fields provided'}), 400

    zone_id = None if session.get('is_super') else session.get('zone', '')
    if not _validate_filepath(filepath, zone_id):
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404
        ws = wb[sheet]

        header_row_idx = None
        for i, row_cells in enumerate(ws.iter_rows(values_only=True), start=1):
            rv = [str(v).strip() if v else '' for v in row_cells]
            if 'Date' in rv or 'التاريخ' in rv:
                header_row_idx = i; break

        if header_row_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': 'Header row not found'}), 400

        col_map = {}
        for ci, cell in enumerate(ws[header_row_idx], start=1):
            if cell.value:
                col_map[str(cell.value).strip()] = ci

        new_row = DATA_START_ROW
        for r in range(DATA_START_ROW, ws.max_row + 2):
            row_empty = True
            for ci in range(1, 14):
                if ws.cell(row=r, column=ci).value not in (None, ''):
                    row_empty = False
                    break
            if row_empty:
                new_row = r
                break

        for col_name, value in fields.items():
            if col_name in col_map:
                cast_value = value
                try:
                    if '.' in str(value):
                        cast_value = float(value)
                    else:
                        cast_value = int(value)
                except Exception:
                    cast_value = value if value != '' else None
                ws.cell(row=new_row, column=col_map[col_name]).value = cast_value

        if 'Date' not in fields and 'التاريخ' not in fields:
            date_col = col_map.get('Date') or col_map.get('التاريخ')
            if date_col:
                ws.cell(row=new_row, column=date_col).value = datetime.now().strftime('%Y-%m-%d')

        wb.save(filepath)
        wb.close()
        return jsonify({'success': True, 'new_row': new_row})

    except Exception:
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء إضافة الصف'}), 500



@app.route('/api/dashboard')
@zone_required
def api_dashboard():
    """Scan Excel files for the current dashboard scope and return live metrics."""
    zone_id = session.get('active_view_zone') or session.get('zone', '')
    is_super = session.get('is_super', False)
    requested_zone = request.args.get('zone', '').strip()

    root = get_years_root()
    if not root:
        return jsonify({'error': 'No data directory'}), 404

    files_to_scan = []
    available_dashboard_zones = [z for z in ZONES if z['id'] not in SUPER_ZONES]
    valid_zone_ids = {z['id'] for z in available_dashboard_zones}
    if is_super and requested_zone and requested_zone != 'all':
        if requested_zone not in valid_zone_ids:
            return jsonify({'error': 'Invalid dashboard zone'}), 400
        scan_zones = [requested_zone]
    elif is_super:
        scan_zones = [z['id'] for z in available_dashboard_zones]
    else:
        scan_zones = [zone_id]

    for zid in scan_zones:
        zone_path = os.path.join(root, zid)
        if not os.path.isdir(zone_path):
            continue
        for root_dir, dirs, files in os.walk(zone_path):
            for f in files:
                lf = f.lower()
                if f.startswith('~$'):
                    continue
                if lf in ('other+.xlsm', 'sacks.xlsm'):
                    files_to_scan.append((zid, os.path.join(root_dir, f)))

    total_items = 0
    total_in    = 0
    total_out   = 0
    zero_stock  = 0
    low_stock   = 0
    LOW_THRESHOLD = 10
    alerts      = []
    item_out    = {}   # name -> total out
    zone_in     = {}   # zone -> total in
    zone_out    = {}   # zone -> total out
    zone_items  = {}   # zone -> row count
    zone_files  = {}   # zone -> file count
    zone_item_out = {}  # zone -> item -> total out
    zone_low_zero = {}  # zone -> low/zero counters
    latest_files = []
    unreadable  = 0
    latest_mtime = None
    # Log sheet data: IN/OUT operations
    log_in_ops  = []   # list of {'time','qty','item','category','file','zone'}
    log_out_ops = []
    log_item_out = {}  # item_name -> total out from Log

    def to_number(value):
        if value in (None, ''):
            return 0
        try:
            return float(value)
        except Exception:
            try:
                text = str(value).replace(',', '').strip()
                return float(text) if text else 0
            except Exception:
                return 0

    def clean_header(value):
        return re.sub(r'\s+', ' ', str(value or '').strip()).lower()

    def build_header_map(row):
        aliases = {
            'color': {'color', 'colour', 'item', 'item name', 'name', 'الصنف', 'اللون'},
            'type': {'type', 'category', 'item type', 'النوع', 'الفئة'},
            'size': {'size', 'الحجم', 'المقاس'},
            'in': {'in', 'in qty', 'in quantity', 'qty in', 'الوارد'},
            'out': {'out', 'out qty', 'out quantity', 'qty out', 'الصادر'},
            'balance': {'current balance', 'current', 'balance', 'stock', 'الرصيد الحالي', 'الرصيد'},
        }
        header_map = {}
        normalized = [clean_header(c) for c in row]
        for key, names in aliases.items():
            for idx, header in enumerate(normalized):
                if header in names:
                    header_map[key] = idx
                    break
        return header_map

    def get_cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    for zid, fpath in files_to_scan:
        zone_meta = next((z for z in ZONES if z['id'] == zid), None)
        zone_label = zone_meta['name'] if zone_meta else zid
        zone_files[zone_label] = zone_files.get(zone_label, 0) + 1
        try:
            mtime = os.path.getmtime(fpath)
            latest_mtime = mtime if latest_mtime is None else max(latest_mtime, mtime)
            latest_files.append({
                'zone': zone_label,
                'file': os.path.basename(fpath),
                'path': os.path.relpath(fpath, root),
                'updated': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),
                'mtime': mtime,
            })
        except Exception:
            pass

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception:
            unreadable += 1
            continue

        # Read Log sheet for actual IN/OUT operations
        for log_sheet in wb.sheetnames:
            if log_sheet.lower() != 'log':
                continue
            try:
                ws_log = wb[log_sheet]
                log_rows_all = list(ws_log.iter_rows(values_only=True))
                if not log_rows_all:
                    continue
                log_headers = [str(v).strip() if v else '' for v in log_rows_all[0]]
                # Find column indices: Time(0), Type(1), Qty(2), Balance(3), Color(4/name), Size(5), ItemType(6), Category(7)
                # Based on LOG_COL_* constants
                lh = {h.lower(): i for i, h in enumerate(log_headers) if h}
                # Try by position fallback or header name
                ci_type = lh.get('type', lh.get('operation', 1))
                ci_qty  = lh.get('qty', lh.get('quantity', 2))
                ci_color = lh.get('color', lh.get('item', lh.get('name', 4)))
                ci_cat  = lh.get('category', lh.get('item type', 7))
                ci_size = lh.get('size', 5)
                for lrow in log_rows_all[1:]:
                    if not lrow or all(v is None for v in lrow):
                        continue
                    def lget(idx):
                        try: return lrow[idx] if idx < len(lrow) else None
                        except: return None
                    op  = str(lget(ci_type) or '').strip().upper()
                    qty = to_number(lget(ci_qty))
                    color = str(lget(ci_color) or '').strip()
                    cat   = str(lget(ci_cat) or '').strip()
                    size  = str(lget(ci_size) or '').strip()
                    item_label = ' - '.join(p for p in [cat, color, size] if p) or color or '—'
                    if op == 'IN' and qty:
                        log_in_ops.append({'qty': qty, 'item': item_label, 'zone': zone_label})
                    elif op == 'OUT' and qty:
                        log_out_ops.append({'qty': qty, 'item': item_label, 'zone': zone_label})
                        log_item_out[item_label] = log_item_out.get(item_label, 0) + qty
            except Exception as _le:
                pass

        for sheet_name in wb.sheetnames:
            if 'log' in sheet_name.lower():
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = None
            header_map = {}
            for i, row in enumerate(rows):
                candidate = build_header_map(row)
                if 'in' in candidate and 'out' in candidate and ('balance' in candidate or 'color' in candidate):
                    header_idx = i
                    header_map = candidate
                    break

            if header_idx is None:
                continue

            for row in rows[header_idx + 1:]:
                color_val = get_cell(row, header_map.get('color'))
                type_val = get_cell(row, header_map.get('type'))
                size_val = get_cell(row, header_map.get('size'))

                # Fallback to fixed column indices for merged cells
                # (openpyxl read_only=True returns None for non-master merged cells)
                def _fb(val, col_idx):
                    if val is None or str(val).strip() in ('', 'None'):
                        try: return row[col_idx - 1] if col_idx - 1 < len(row) else None
                        except: return None
                    return val
                color_val = _fb(color_val, COL_COLOR)
                type_val  = _fb(type_val,  COL_TYPE)
                size_val  = _fb(size_val,  COL_SIZE)
                # Category column as extra fallback for type
                if type_val is None or str(type_val).strip() in ('', 'None'):
                    try: type_val = row[COL_CATEGORY - 1] if COL_CATEGORY - 1 < len(row) else None
                    except: pass

                in_val = to_number(get_cell(row, header_map.get('in')))
                out_val = to_number(get_cell(row, header_map.get('out')))
                bal_cell = get_cell(row, header_map.get('balance'))
                # Fallback balance to fixed column
                if bal_cell is None or str(bal_cell).strip() in ('', 'None'):
                    try: bal_cell = row[COL_CURRENT - 1] if COL_CURRENT - 1 < len(row) else None
                    except: pass
                bal_val = None if bal_cell in (None, '') else to_number(bal_cell)

                meaningful = any(str(v or '').strip() for v in (color_val, type_val, size_val)) or in_val or out_val or bal_val not in (None, 0)
                if not meaningful:
                    continue
                total_items += 1
                zone_items[zone_label] = zone_items.get(zone_label, 0) + 1

                total_in  += in_val
                total_out += out_val
                zone_in[zone_label]  = zone_in.get(zone_label,  0) + in_val
                zone_out[zone_label] = zone_out.get(zone_label, 0) + out_val

                # Color is the primary identifier; size next, then type/category
                name_parts = [str(v).strip() for v in (color_val, size_val, type_val) if str(v or '').strip()]
                item_name = ' - '.join(name_parts) if name_parts else ''

                if item_name and out_val > 0:
                    item_out[item_name] = item_out.get(item_name, 0) + out_val
                    zone_bucket = zone_item_out.setdefault(zone_label, {})
                    zone_bucket[item_name] = zone_bucket.get(item_name, 0) + out_val

                if bal_val is not None:
                    if bal_val == 0:
                        zero_stock += 1
                        zone_low_zero.setdefault(zone_label, {'low': 0, 'zero': 0})['zero'] += 1
                        alerts.append({'name': item_name or '—', 'sheet': sheet_name,
                                       'balance': 0, 'level': 'danger'})
                    elif bal_val < LOW_THRESHOLD:
                        low_stock += 1
                        zone_low_zero.setdefault(zone_label, {'low': 0, 'zero': 0})['low'] += 1
                        alerts.append({'name': item_name or '—', 'sheet': sheet_name,
                                       'balance': bal_val, 'level': 'warn'})

        wb.close()

    top_items = sorted([{'name': k, 'out': v} for k, v in item_out.items()],
                       key=lambda x: x['out'], reverse=True)[:10]
    zone_consumption = {}
    for zname, items in zone_item_out.items():
        ranked = sorted(items.items(), key=lambda kv: kv[1], reverse=True)
        positive = [kv for kv in ranked if kv[1] > 0]
        zone_consumption[zname] = {
            'top': {'name': ranked[0][0], 'out': round(ranked[0][1], 2)} if ranked else None,
            'lowest': {'name': positive[-1][0], 'out': round(positive[-1][1], 2)} if positive else None,
            'top5': [{'name': name, 'out': round(value, 2)} for name, value in ranked[:5]],
            'moving_items': len(positive),
        }
    most_active_zone = None
    for zname in set(list(zone_in.keys()) + list(zone_out.keys())):
        movement = zone_in.get(zname, 0) + zone_out.get(zname, 0)
        if most_active_zone is None or movement > most_active_zone['movement']:
            most_active_zone = {'zone': zname, 'movement': round(movement, 2)}
    latest_files = sorted(latest_files, key=lambda f: f.get('mtime', 0), reverse=True)[:8]
    high_usage_threshold = max(100, (total_out / max(len(item_out), 1)) * 2) if item_out else 100
    high_usage_items = [
        {'name': item['name'], 'out': round(item['out'], 2)}
        for item in top_items
        if item.get('out', 0) >= high_usage_threshold
    ]
    latest_file_update = datetime.fromtimestamp(latest_mtime).strftime('%Y-%m-%d %H:%M:%S') if latest_mtime else ''
    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    scope = 'All zones' if len(scan_zones) > 1 else (next((z['name'] for z in ZONES if z['id'] == scan_zones[0]), scan_zones[0]) if scan_zones else zone_id)

    # Build log-based top items (from actual Log sheet operations)
    log_top_items = sorted(
        [{'name': k, 'out': round(v, 2)} for k, v in log_item_out.items()],
        key=lambda x: x['out'], reverse=True
    )[:10]
    log_total_in  = round(sum(op['qty'] for op in log_in_ops), 2)
    log_total_out = round(sum(op['qty'] for op in log_out_ops), 2)

    return jsonify({
        'total_items': total_items,
        'total_in':    round(total_in,  2),
        'total_out':   round(total_out, 2),
        'zero_stock':  zero_stock,
        'low_stock':   low_stock,
        'alerts':      alerts[:50],
        'top_items':   [{'name': item['name'], 'out': round(item['out'], 2)} for item in top_items],
        'zone_in':     {k: round(v, 2) for k, v in zone_in.items()},
        'zone_out':    {k: round(v, 2) for k, v in zone_out.items()},
        'zone_items':  zone_items,
        'zone_files':  zone_files,
        'zone_consumption': zone_consumption,
        'zone_low_zero': zone_low_zero,
        'latest_files': [{k: v for k, v in item.items() if k != 'mtime'} for item in latest_files],
        'most_active_zone': most_active_zone,
        'high_usage_items': high_usage_items,
        'high_usage_threshold': round(high_usage_threshold, 2),
        'file_count':  len(files_to_scan),
        'unreadable_files': unreadable,
        'latest_file_update': latest_file_update,
        'generated_at': generated_at,
        'scope': scope,
        'selected_zone': requested_zone if is_super and requested_zone else ('all' if is_super else zone_id),
        'dashboard_zones': [{'id': 'all', 'name': 'All zones'}] + [{'id': z['id'], 'name': z['name'], 'label': z['label']} for z in available_dashboard_zones] if is_super else [],
        # Log sheet data
        'log_top_items': log_top_items,
        'log_total_in':  log_total_in,
        'log_total_out': log_total_out,
        'log_ops_count': len(log_in_ops) + len(log_out_ops),
    })

@app.route('/api/login_log')
@zone_required
def api_login_log():
    """Return login history — dev only."""
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _log_lock:
        entries = _read_login_log()
    return jsonify({
        'entries': list(reversed(entries)),
        'total': len(entries),
        'log_file': _LOGIN_LOG_FILE,
    })

@app.route('/api/admin/pending_requests_count')
@zone_required
def api_admin_pending_requests_count():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    return jsonify({'count': _pending_request_count()})

@app.route('/api/admin/registration_requests')
@zone_required
def api_admin_registration_requests():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        rows = conn.execute(
            "SELECT * FROM registration_requests WHERE status = 'pending' ORDER BY created_at DESC, id DESC"
        ).fetchall()
    return jsonify({
        'count': len(rows),
        'requests': [
            {
                'id': row['id'],
                'full_name': row['full_name'],
                'username': row['username'],
                'email': row['email'],
                'phone': row['phone'],
                'job_title': row['job_title'],
                'gender': row['gender'] if 'gender' in row.keys() else '',
                'birth_date': row['birth_date'] if 'birth_date' in row.keys() else '',
                'privacy_accepted': bool(row['privacy_accepted'] if 'privacy_accepted' in row.keys() else 0),
                'security_question': row['security_question'],
                'created_at': row['created_at'],
            }
            for row in rows
        ],
    })

@app.route('/api/admin/registration_requests/<int:request_id>/approve', methods=['POST'])
@zone_required
def api_admin_approve_registration(request_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    admin_user = session.get('username', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _db_connect() as conn:
        row = conn.execute(
            "SELECT * FROM registration_requests WHERE id = ? AND status = 'pending'",
            (request_id,),
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
        exists = conn.execute(
            "SELECT 1 FROM users WHERE lower(username) = ?",
            (_normalize_username(row['username']),),
        ).fetchone()
        if exists:
            conn.execute(
                "UPDATE registration_requests SET status = 'approved', reviewed_at = ?, reviewed_by = ? WHERE id = ?",
                (now, admin_user, request_id),
            )
            return jsonify({'success': False, 'message': 'اسم المستخدم موجود مسبقاً'}), 409
        conn.execute(
            """
            INSERT INTO users
            (full_name, username, email, phone, job_title, gender, birth_date, privacy_accepted, password_hash, security_question, security_answer_hash, approved, is_admin, created_at, approved_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 0, ?, ?, ?)
            """,
            (
                row['full_name'],
                row['username'],
                row['email'],
                row['phone'],
                row['job_title'],
                row['gender'] if 'gender' in row.keys() else '',
                row['birth_date'] if 'birth_date' in row.keys() else '',
                row['privacy_accepted'] if 'privacy_accepted' in row.keys() else 0,
                row['password_hash'],
                row['security_question'],
                row['security_answer_hash'],
                row['created_at'],
                now,
                admin_user,
            ),
        )
        conn.execute(
            "UPDATE registration_requests SET status = 'approved', reviewed_at = ?, reviewed_by = ? WHERE id = ?",
            (now, admin_user, request_id),
        )
    return jsonify({'success': True, 'message': 'تمت الموافقة على الطلب'})

@app.route('/api/admin/registration_requests/<int:request_id>/reject', methods=['POST'])
@zone_required
def api_admin_reject_registration(request_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    admin_user = session.get('username', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _db_connect() as conn:
        row = conn.execute(
            "SELECT id FROM registration_requests WHERE id = ? AND status = 'pending'",
            (request_id,),
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
        conn.execute(
            "UPDATE registration_requests SET status = 'rejected', reviewed_at = ?, reviewed_by = ? WHERE id = ?",
            (now, admin_user, request_id),
        )
    return jsonify({'success': True, 'message': 'تم رفض الطلب'})

@app.route('/api/admin/registered_users')
@zone_required
def api_admin_registered_users():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        rows = conn.execute(
            "SELECT id, full_name, username, email, phone, job_title, gender, birth_date, privacy_accepted, security_question, password_hash, security_answer_hash, approved_at, created_at, suspended_until, suspended_by, suspended_at FROM users WHERE approved = 1 ORDER BY approved_at DESC, id DESC"
        ).fetchall()
    return jsonify({
        'count': len(rows),
        'db_file': AUTH_DB_FILE,
        'users': [
            {
                'id': row['id'],
                'full_name': row['full_name'],
                'username': row['username'],
                'email': row['email'],
                'phone': row['phone'],
                'job_title': row['job_title'],
                'gender': row['gender'] if 'gender' in row.keys() else '',
                'birth_date': row['birth_date'] if 'birth_date' in row.keys() else '',
                'privacy_accepted': bool(row['privacy_accepted'] if 'privacy_accepted' in row.keys() else 0),
                'security_question': row['security_question'],
                'password_stored_as': 'one_way_hash' if row['password_hash'] else '',
                'security_answer_stored_as': 'one_way_hash' if row['security_answer_hash'] else '',
                'approved_at': row['approved_at'],
                'created_at': row['created_at'],
                'suspended_until': row['suspended_until'],
                'suspended_by': row['suspended_by'],
                'suspended_at': row['suspended_at'],
            }
            for row in rows
        ],
    })

@app.route('/api/admin/registered_users/export.xlsx')
@zone_required
def api_admin_export_registered_users():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        rows = conn.execute("""
            SELECT id, full_name, username, email, phone, job_title, gender, birth_date, privacy_accepted,
                   security_question, password_hash, security_answer_hash,
                   approved, is_admin, created_at, approved_at, created_by,
                   suspended_until, suspended_by, suspended_at
            FROM users
            WHERE approved = 1
            ORDER BY approved_at DESC, id DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registered Users'
    headers = [
        'ID', 'Full Name', 'Username', 'Email', 'Phone', 'Job Title', 'Gender', 'Birth Date', 'Privacy Accepted',
        'Password', 'Security Question', 'Security Answer',
        'Password Hash', 'Security Answer Hash',
        'Approved', 'Is Admin', 'Created At', 'Approved At', 'Created By',
        'Suspended Until', 'Suspended By', 'Suspended At',
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row['id'], row['full_name'], row['username'], row['email'], row['phone'], row['job_title'],
            row['gender'] if 'gender' in row.keys() else '',
            row['birth_date'] if 'birth_date' in row.keys() else '',
            'Yes' if (row['privacy_accepted'] if 'privacy_accepted' in row.keys() else 0) else 'No',
            'Not recoverable - stored as one-way hash',
            row['security_question'],
            'Not recoverable - stored as one-way hash',
            row['password_hash'], row['security_answer_hash'],
            row['approved'], row['is_admin'], row['created_at'], row['approved_at'], row['created_by'],
            row['suspended_until'], row['suspended_by'], row['suspended_at'],
        ])
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAF7')
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or '')))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = 'registered_users_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
@app.route('/api/admin/registered_users/<int:user_id>/suspend', methods=['POST'])
@zone_required
def api_admin_suspend_user(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    minutes = int(data.get('minutes') or 0)
    if minutes <= 0 or minutes > 43200:
        return jsonify({'success': False, 'message': 'مدة الإيقاف يجب أن تكون بين دقيقة و 30 يوم'}), 400
    now = datetime.now()
    until = now + timedelta(minutes=minutes)
    admin_user = session.get('username', '')
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        if _is_single_login_exempt(row['username']):
            return jsonify({'success': False, 'message': 'لا يمكن إيقاف حساب الأدمن أو الديف'}), 400
        conn.execute(
            "UPDATE users SET suspended_until = ?, suspended_by = ?, suspended_at = ? WHERE id = ?",
            (until.strftime('%Y-%m-%d %H:%M:%S'), admin_user, now.strftime('%Y-%m-%d %H:%M:%S'), user_id),
        )
    _clear_active_session(row['username'])
    return jsonify({'success': True, 'message': 'تم إيقاف المستخدم مؤقتاً', 'suspended_until': until.strftime('%Y-%m-%d %H:%M:%S')})

@app.route('/api/admin/registered_users/<int:user_id>/unsuspend', methods=['POST'])
@zone_required
def api_admin_unsuspend_user(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        cur = conn.execute(
            "UPDATE users SET suspended_until = NULL, suspended_by = NULL, suspended_at = NULL WHERE id = ?",
            (user_id,),
        )
    if cur.rowcount == 0:
        return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
    return jsonify({'success': True, 'message': 'تم إلغاء الإيقاف'})

@app.route('/api/admin/registered_users/<int:user_id>/password', methods=['POST'])
@zone_required
def api_admin_reset_user_password(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    new_password = str(data.get('new_password', '')).strip()
    confirm_password = str(data.get('confirm_password', '')).strip()
    if not new_password:
        return jsonify({'success': False, 'message': 'يرجى إدخال كلمة المرور الجديدة'}), 400
    if new_password != confirm_password:
        return jsonify({'success': False, 'message': 'كلمة المرور وتأكيدها غير متطابقين'}), 400
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (_hash_secret(new_password), user_id))
    _clear_active_session(row['username'])
    return jsonify({'success': True, 'message': 'تم تغيير كلمة مرور المستخدم'})

@app.route('/api/admin/registered_users/<int:user_id>/security', methods=['POST'])
@zone_required
def api_admin_reset_user_security(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json(silent=True) or {}
    security_question = str(data.get('security_question', '')).strip()
    security_answer = str(data.get('security_answer', '')).strip()
    if not security_question or not security_answer:
        return jsonify({'success': False, 'message': 'يرجى إدخال السؤال والجواب الأمني'}), 400
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        conn.execute(
            "UPDATE users SET security_question = ?, security_answer_hash = ? WHERE id = ?",
            (security_question, _hash_secret(_normalize_text(security_answer)), user_id),
        )
    return jsonify({'success': True, 'message': 'تم تغيير السؤال والجواب الأمني'})

@app.route('/api/admin/registered_users/<int:user_id>', methods=['DELETE'])
@zone_required
def api_admin_delete_user(user_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _db_connect() as conn:
        row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'المستخدم غير موجود'}), 404
        if _is_single_login_exempt(row['username']):
            return jsonify({'success': False, 'message': 'لا يمكن حذف حساب الأدمن أو الديف'}), 400
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    _clear_active_session(row['username'])
    return jsonify({'success': True, 'message': 'تم حذف الحساب بالكامل'})

@app.route('/api/alert_count')
@zone_required
def api_alert_count():
    """Quick scan: return number of zero-stock items for the badge."""
    zone_id  = session.get('active_view_zone') or session.get('zone', '')
    is_super = session.get('is_super', False)
    root = get_years_root()
    if not root:
        return jsonify({'zero': 0})
    scan_zones = [z['id'] for z in ZONES if z['id'] not in SUPER_ZONES] if is_super else [zone_id]
    zero = 0
    for zid in scan_zones:
        zone_path = os.path.join(root, zid)
        if not os.path.isdir(zone_path):
            continue
        for rdir, dirs, files in os.walk(zone_path):
            for f in files:
                if not f.lower().endswith(('.xlsx','.xlsm','.xls')):
                    continue
                try:
                    wb = openpyxl.load_workbook(os.path.join(rdir,f), data_only=True, read_only=True)
                    for sname in wb.sheetnames:
                        if 'log' in sname.lower(): continue
                        ws = wb[sname]
                        rows = list(ws.iter_rows(values_only=True))
                        hdr_idx = None
                        headers = []
                        for i,row in enumerate(rows):
                            rv=[str(c).strip() if c else '' for c in row]
                            if 'Current Balance' in rv:
                                hdr_idx=i; headers=rv; break
                        if hdr_idx is None: continue
                        ci_bal = headers.index('Current Balance')
                        for row in rows[hdr_idx+1:]:
                            if all(c is None or str(c).strip()=='' for c in row): continue
                            try:
                                if ci_bal < len(row) and float(row[ci_bal] or 1) == 0:
                                    zero += 1
                            except: pass
                    wb.close()
                except: pass
    return jsonify({'zero': zero})

# ══════════════════════════════════════════════════════════════════
#  QR SCANNER — صفحة المسح + API يرجع آخر رصيد من ملف الإكسيل
# ══════════════════════════════════════════════════════════════════

# خريطة SKU → (اسم الشيت، اسم اللون في الملف)
SKU_MAP = {
    'CHKN-BLU-001': ('Chicken', 'Blue'),
    'CHKN-BRN-001': ('Chicken', 'Brown'),
    'CHKN-DBL-001': ('Chicken', 'Dark Blue'),
    'CHKN-DGR-001': ('Chicken', 'Dark Green'),
    'CHKN-GRY-001': ('Chicken', 'Gray'),
    'CHKN-GRN-001': ('Chicken', 'Green'),
    'CHKN-LBL-001': ('Chicken', 'Light Blue'),
    'CHKN-ORG-001': ('Chicken', 'Orange'),
    'CHKN-ORG-002': ('Chicken', 'Orange (70*45)'),
    'CHKN-PRP-001': ('Chicken', 'Purple'),
    'CHKN-RED-001': ('Chicken', 'Red'),
    'CHKN-YLW-001': ('Chicken', 'Yellow'),
}

SKU_NAMES_AR = {
    'CHKN-BLU-001': 'أزرق',
    'CHKN-BRN-001': 'بني',
    'CHKN-DBL-001': 'أزرق غامق',
    'CHKN-DGR-001': 'أخضر غامق',
    'CHKN-GRY-001': 'رمادي',
    'CHKN-GRN-001': 'أخضر',
    'CHKN-LBL-001': 'أزرق فاتح',
    'CHKN-ORG-001': 'برتقالي',
    'CHKN-ORG-002': 'برتقالي 70×45',
    'CHKN-PRP-001': 'بنفسجي',
    'CHKN-RED-001': 'أحمر',
    'CHKN-YLW-001': 'أصفر',
}

SKU_HEX = {
    'CHKN-BLU-001': '#3b82f6',
    'CHKN-BRN-001': '#92400e',
    'CHKN-DBL-001': '#1e3a8a',
    'CHKN-DGR-001': '#14532d',
    'CHKN-GRY-001': '#6b7280',
    'CHKN-GRN-001': '#16a34a',
    'CHKN-LBL-001': '#7dd3fc',
    'CHKN-ORG-001': '#f97316',
    'CHKN-ORG-002': '#fb923c',
    'CHKN-PRP-001': '#7c3aed',
    'CHKN-RED-001': '#dc2626',
    'CHKN-YLW-001': '#eab308',
}

def _find_latest_sacks_file():
    """يبحث عن أحدث ملف Sacks.xlsm بناءً على اسم السنة والشهر."""
    root = get_years_root()
    if not root:
        return None
    best_path = None
    best_key  = (-1, -1)
    month_num = {
        'january':1,'february':2,'march':3,'april':4,
        'may':5,'june':6,'july':7,'august':8,
        'september':9,'october':10,'november':11,'december':12
    }
    for zone_folder in os.listdir(root):
        zone_path = os.path.join(root, zone_folder)
        if not os.path.isdir(zone_path):
            continue
        for year_folder in os.listdir(zone_path):
            year_path = os.path.join(zone_path, year_folder)
            if not os.path.isdir(year_path) or not year_folder.isdigit():
                continue
            year_int = int(year_folder)
            for month_folder in os.listdir(year_path):
                month_path = os.path.join(year_path, month_folder)
                sacks_path = os.path.join(month_path, 'Sacks.xlsm')
                if not os.path.isfile(sacks_path):
                    continue
                parts = month_folder.split('-')
                month_int = int(parts[0]) if parts[0].isdigit() else                             month_num.get(parts[-1].lower(), 0)
                key = (year_int, month_int)
                if key > best_key:
                    best_key  = key
                    best_path = sacks_path
    return best_path

def _get_last_balance(sheet_name, color_name):
    """يفتح أحدث Sacks.xlsm ويرجع آخر Current Balance للون المحدد."""
    filepath = _find_latest_sacks_file()
    if not filepath:
        return None, None, None

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None, None, None

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # إيجاد صف الهيدر
        hdr_idx = None
        headers = []
        for i, row in enumerate(rows):
            rv = [str(c).strip() if c else '' for c in row]
            if 'Current Balance' in rv:
                hdr_idx = i
                headers = rv
                break
        if hdr_idx is None:
            return None, None, None

        ci_color   = headers.index('Color')           if 'Color'           in headers else None
        ci_balance = headers.index('Current Balance') if 'Current Balance' in headers else None
        ci_date    = headers.index('Date')            if 'Date'            in headers else None

        if ci_color is None or ci_balance is None:
            return None, None, None

        last_balance = None
        last_date    = None

        for row in rows[hdr_idx + 1:]:
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            row_color = str(row[ci_color]).strip() if ci_color < len(row) and row[ci_color] else ''
            if row_color.lower() != color_name.lower():
                continue
            bal = row[ci_balance] if ci_balance < len(row) else None
            if bal is not None and str(bal).strip() not in ('', 'None', '/'):
                try:
                    last_balance = float(bal)
                    if ci_date and ci_date < len(row) and row[ci_date]:
                        d = row[ci_date]
                        last_date = d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d)[:10]
                except (ValueError, TypeError):
                    pass

        return last_balance, last_date, filepath

    except Exception as e:
        _logger.error(f'QR scan error: {e}')
        return None, None, None


STOCK_QR_FILE_MAP = {
    'OTHER': 'Other+.xlsm',
    'SACKS': 'Sacks.xlsm',
}

STOCK_QR_HEX = {
    'BLUE': '#3b82f6', 'DARK BLUE': '#1e3a8a', 'LIGHT BLUE': '#7dd3fc',
    'GREEN': '#16a34a', 'DARK GREEN': '#14532d', 'RED': '#dc2626',
    'ORANGE': '#f97316', 'PURPLE': '#7c3aed', 'BROWN': '#92400e',
    'YELLOW': '#eab308', 'GRAY': '#6b7280', 'WHITE': '#e5e7eb',
    'BLACK': '#111827',
}

def _guess_stock_hex(name):
    text = str(name or '').upper()
    for key, color in STOCK_QR_HEX.items():
        if key in text:
            return color
    return '#1a3a5c'

def _find_month_folder(year_path, month_code):
    if not os.path.isdir(year_path):
        return None
    prefix = f'{int(month_code):02d}-'
    for folder in os.listdir(year_path):
        full = os.path.join(year_path, folder)
        if os.path.isdir(full) and folder.startswith(prefix):
            return folder
    return None

def _parse_stocktaking_sku(sku):
    m = re.fullmatch(r'STK-(ZONE[0-9]+)-(\d{4})-(\d{2})-(OTHER|SACKS)-R(\d{3})-C(\d{3})', sku or '')
    if not m:
        return None
    zone_id, year, month_code, file_key, row_s, col_s = m.groups()
    return {
        'zone_id': zone_id.lower(),
        'year': year,
        'month_code': month_code,
        'file_key': file_key,
        'row': int(row_s),
        'col': int(col_s),
    }

def _stocktaking_category(ws, row, col, item_name):
    item_text = str(item_name or '').strip()
    for merged in getattr(ws, 'merged_cells', []).ranges:
        if merged.min_row <= 1 <= merged.max_row and merged.min_col <= col <= merged.max_col:
            value = ws.cell(merged.min_row, merged.min_col).value
            if value is not None and str(value).strip() and str(value).strip() != item_text:
                return str(value).strip()
    for c in range(col, 0, -1):
        value = ws.cell(1, c).value
        if value is not None and str(value).strip() and str(value).strip() != item_text:
            return str(value).strip()
    for r in range(row - 1, 0, -1):
        value = ws.cell(r, col).value
        if value is None or str(value).strip() == '':
            continue
        if isinstance(value, (int, float)):
            continue
        text = str(value).strip()
        if text != item_text:
            return text
    return ''

def _stocktaking_scan_result(sku):
    meta = _parse_stocktaking_sku(sku)
    if not meta:
        return None
    root = get_years_root()
    if not root:
        return {'found': False, 'error': 'تعذر تحديد مجلد zones'}
    filename = STOCK_QR_FILE_MAP.get(meta['file_key'])
    year_path = os.path.join(root, meta['zone_id'], meta['year'])
    month_folder = _find_month_folder(year_path, meta['month_code'])
    if not month_folder or not filename:
        return {'found': False, 'error': 'ملف الشهر غير موجود'}
    filepath = os.path.join(year_path, month_folder, filename)
    if not os.path.isfile(filepath):
        return {'found': False, 'error': 'ملف الإكسل غير موجود'}
    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
        if 'Stocktaking' not in wb.sheetnames:
            wb.close()
            return {'found': False, 'error': 'شيت Stocktaking غير موجود'}
        ws = wb['Stocktaking']
        item_name = ws.cell(meta['row'], meta['col']).value
        balance = ws.cell(meta['row'] + 1, meta['col']).value
        category = _stocktaking_category(ws, meta['row'], meta['col'], item_name)
        wb.close()
        if item_name is None or str(item_name).strip() == '':
            return {'found': False, 'error': 'الصنف غير موجود داخل Stocktaking'}
        try:
            balance_value = float(balance)
            balance_out = int(balance_value) if balance_value.is_integer() else balance_value
        except (TypeError, ValueError, AttributeError):
            balance_out = balance if balance not in (None, '') else 0
        modified = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%Y-%m-%d %H:%M')
        return {
            'found': True,
            'sku': sku,
            'nameAr': str(item_name).strip(),
            'category': category or meta['file_key'].title(),
            'color': str(item_name).strip(),
            'hex': _guess_stock_hex(item_name),
            'balance': balance_out,
            'date': modified,
            'file': filename,
            'zone': meta['zone_id'],
            'year': meta['year'],
            'month': month_folder,
        }
    except Exception as e:
        _logger.error(f'Stocktaking QR scan error: {e}')
        return {'found': False, 'error': 'تعذر قراءة بيانات Stocktaking'}


@app.route('/scan')
def scan_page():
    """صفحة مسح QR — تتطلب تسجيل دخول (مستخدم + زون).
    لو ما دخل، يحوله للوغن وبعدها يرجعه هنا تلقائياً."""
    if not session.get('logged_in'):
        return redirect(url_for('login_page') + '?next=/scan')
    if not session.get('zone'):
        return redirect(url_for('zones_page'))
    return render_template('scan.html')



@app.route('/qrscan')
def qrscan_page():
    """صفحة مسح QR — عامة بدون تسجيل دخول."""
    sku = request.args.get('sku', '').strip().upper()
    return render_template('qrscan.html', sku=sku)


@app.route('/api/qrscan/<sku>')
def api_qrscan(sku):
    """يرجع آخر رصيد للصنف — يتطلب login فقط بدون zone."""
    sku = sku.strip().upper()
    stocktaking_result = _stocktaking_scan_result(sku)
    if stocktaking_result is not None:
        status = 200 if stocktaking_result.get('found') else 404
        return jsonify(stocktaking_result), status
    if sku not in SKU_MAP:
        return jsonify({'found': False, 'error': f'"{sku}" غير مسجل في النظام'}), 404
    sheet_name, color_name = SKU_MAP[sku]
    balance, date, filepath = _get_last_balance(sheet_name, color_name)
    if balance is None:
        return jsonify({'found': False, 'error': 'تعذر قراءة الملف أو لا توجد بيانات'}), 500
    return jsonify({
        'found':    True,
        'sku':      sku,
        'nameAr':   SKU_NAMES_AR.get(sku, sku),
        'category': 'شوالات الجاج',
        'hex':      SKU_HEX.get(sku, '#6b7280'),
        'balance':  int(balance),
        'date':     date or '—',
    })


@app.route('/about')
def about_page():
    """صفحة About Us — لا تتطلب تسجيل دخول."""
    return render_template('about.html')



@app.route('/privacy')
def privacy_page():
    """Privacy Policy page - public."""
    return render_template('privacy.html')

@app.route('/terms')
def terms_page():
    """Terms of Use page - public."""
    return render_template('terms.html')

@app.route('/for-more')
def for_more_page():
    """Developer / For More page - public."""
    return render_template('formore.html')



@app.route('/api/contact', methods=['POST'])
def api_contact():
    data = request.get_json(silent=True) or {}
    name = str(data.get('name', '')).strip()
    phone = str(data.get('phone', '')).strip()
    message = str(data.get('message', '')).strip()
    if not name or not phone or not message:
        return jsonify({'success': False, 'message': 'Missing required fields'}), 400
    with _data_lock:
        items = _read_json_list(CONTACT_MESSAGES_FILE)
        item = {
            'id': _next_json_id(items),
            'name': name,
            'phone': phone,
            'email': str(data.get('email', '')).strip(),
            'department': str(data.get('department', '')).strip(),
            'message': message,
            'status': 'new',
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ip': _get_ip(),
        }
        items.append(item)
        _write_json_list(CONTACT_MESSAGES_FILE, items)
    return jsonify({'success': True, 'id': item['id']})

@app.route('/api/admin/contact_messages')
@zone_required
def api_contact_messages():
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    items = sorted(_read_json_list(CONTACT_MESSAGES_FILE), key=lambda x: x.get('id', 0), reverse=True)
    return jsonify({'count': len([x for x in items if x.get('status') == 'new']), 'messages': items})

@app.route('/api/admin/contact_messages/<int:message_id>/read', methods=['POST'])
@zone_required
def api_contact_message_read(message_id):
    if session.get('zone') != 'dev':
        return jsonify({'error': 'غير مصرح'}), 403
    with _data_lock:
        items = _read_json_list(CONTACT_MESSAGES_FILE)
        for item in items:
            if int(item.get('id', 0)) == message_id:
                item['status'] = 'read'
                item['read_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                break
        _write_json_list(CONTACT_MESSAGES_FILE, items)
    return jsonify({'success': True})

@app.route('/qc-workflow')
@zone_required
def qc_workflow_page():
    if session.get('zone') != 'qc':
        return redirect(url_for('index'))
    return render_template('qc.html', qc_role=session.get('qc_role', 'qc'), username=session.get('username', ''))

@app.route('/api/qc/submissions', methods=['GET', 'POST'])
@zone_required
def api_qc_submissions():
    if session.get('zone') != 'qc':
        return jsonify({'error': 'غير مصرح'}), 403
    if request.method == 'GET':
        items = sorted(_read_json_list(QC_SUBMISSIONS_FILE), key=lambda x: x.get('id', 0), reverse=True)
        return jsonify({'items': items, 'role': session.get('qc_role', 'qc')})
    if session.get('qc_role') != 'labeling':
        return jsonify({'success': False, 'message': 'Only Labeling Assistant can submit photos'}), 403
    photo = request.files.get('photo')
    note = request.form.get('note', '').strip()
    if not photo:
        return jsonify({'success': False, 'message': 'Photo is required'}), 400
    os.makedirs(QC_UPLOAD_DIR, exist_ok=True)
    ext = os.path.splitext(photo.filename or '')[1].lower()
    if ext not in {'.jpg', '.jpeg', '.png', '.webp'}:
        ext = '.jpg'
    filename = f"qc_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(4)}{ext}"
    photo.save(os.path.join(QC_UPLOAD_DIR, filename))
    with _data_lock:
        items = _read_json_list(QC_SUBMISSIONS_FILE)
        item = {
            'id': _next_json_id(items),
            'image_url': '/static/qc_uploads/' + filename,
            'note': note,
            'status': 'pending',
            'created_by': session.get('username', ''),
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'reviewed_by': '',
            'reviewed_at': '',
            'review_note': '',
        }
        items.append(item)
        _write_json_list(QC_SUBMISSIONS_FILE, items)
    return jsonify({'success': True, 'item': item})

@app.route('/api/qc/submissions/<int:item_id>/status', methods=['POST'])
@zone_required
def api_qc_submission_status(item_id):
    if session.get('zone') != 'qc' or session.get('qc_role') != 'qc':
        return jsonify({'success': False, 'message': 'Only QC can review submissions'}), 403
    data = request.get_json(silent=True) or {}
    status = str(data.get('status', '')).strip().lower()
    if status not in {'approved', 'rejected', 'pending'}:
        return jsonify({'success': False, 'message': 'Invalid status'}), 400
    with _data_lock:
        items = _read_json_list(QC_SUBMISSIONS_FILE)
        found = False
        for item in items:
            if int(item.get('id', 0)) == item_id:
                item['status'] = status
                item['reviewed_by'] = session.get('username', '')
                item['reviewed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                item['review_note'] = str(data.get('review_note', '')).strip()
                found = True
                break
        if not found:
            return jsonify({'success': False, 'message': 'Not found'}), 404
        _write_json_list(QC_SUBMISSIONS_FILE, items)
    return jsonify({'success': True})

@app.route('/api/dashboard/excel_status')
@zone_required
def api_dashboard_excel_status():
    root = get_years_root()
    if not root:
        return jsonify({'connected': False, 'message': 'فشل قراءة الملف', 'last_update': None})
    latest_path = None
    latest_mtime = 0
    for rdir, dirs, files in os.walk(root):
        for f in files:
            if f.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                p = os.path.join(rdir, f)
                try:
                    m = os.path.getmtime(p)
                    if m > latest_mtime:
                        latest_mtime = m
                        latest_path = p
                except Exception:
                    pass
    if not latest_path:
        return jsonify({'connected': False, 'message': 'فشل قراءة الملف', 'last_update': None})
    minutes = max(0, int((datetime.now() - datetime.fromtimestamp(latest_mtime)).total_seconds() // 60))
    return jsonify({'connected': True, 'message': 'متصل', 'file': os.path.basename(latest_path), 'minutes_ago': minutes, 'last_update': datetime.fromtimestamp(latest_mtime).strftime('%Y-%m-%d %H:%M:%S')})


@app.route('/api/scan/<sku>')
@zone_required
def api_scan(sku):
    """يرجع آخر رصيد للصنف من ملف الإكسيل الأحدث."""
    sku = sku.strip().upper()
    stocktaking_result = _stocktaking_scan_result(sku)
    if stocktaking_result is not None:
        status = 200 if stocktaking_result.get('found') else 404
        return jsonify(stocktaking_result), status
    if sku not in SKU_MAP:
        return jsonify({'found': False, 'error': f'"{sku}" غير مسجل في النظام'}), 404

    sheet_name, color_name = SKU_MAP[sku]
    balance, date, filepath = _get_last_balance(sheet_name, color_name)

    if balance is None:
        return jsonify({'found': False, 'error': 'تعذّر قراءة الملف أو لا توجد بيانات'}), 500

    return jsonify({
        'found':    True,
        'sku':      sku,
        'nameAr':   SKU_NAMES_AR.get(sku, sku),
        'category': 'شوالات الجاج',
        'color':    color_name,
        'hex':      SKU_HEX.get(sku, '#6b7280'),
        'balance':  int(balance),
        'date':     date or '—',
    })


# ══════════════════════════════════════════════════════════════════
#  REPORTS — list & serve Excel files from /reports folder
# ══════════════════════════════════════════════════════════════════
REPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reports')

REPORTS_ALLOWED = ('.xlsx', '.xlsm', '.xls', '.csv',
                   '.docx', '.doc', '.dotx',
                   '.pptx', '.ppt', '.pps',
                   '.pdf')

@app.route('/api/reports')
@zone_required
def api_reports():
    """Return a list of all supported report files inside the reports/ folder."""
    if not os.path.isdir(REPORTS_DIR):
        return jsonify({'files': []})
    files = [
        f for f in sorted(os.listdir(REPORTS_DIR))
        if f.lower().endswith(REPORTS_ALLOWED)
    ]
    return jsonify({'files': files})

@app.route('/reports/file/<path:filename>')
@zone_required
def download_report(filename):
    """Serve a report file directly (for PDF, Word, PowerPoint)."""
    safe_name = os.path.basename(filename)
    filepath  = os.path.join(REPORTS_DIR, safe_name)
    if not os.path.isfile(filepath):
        return 'File not found', 404
    if not safe_name.lower().endswith(REPORTS_ALLOWED):
        return 'File type not allowed', 403
    from flask import send_file
    return send_file(filepath, as_attachment=False)

@app.route('/reports/print/<path:filename>')
@zone_required
def print_report(filename):
    """Convert an Excel report to a printable HTML page."""
    safe_name = os.path.basename(filename)
    filepath  = os.path.join(REPORTS_DIR, safe_name)
    if not os.path.isfile(filepath):
        return 'File not found', 404

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return f'Could not open file: {e}', 500

    # Build one HTML table per sheet
    sheets_html = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data:
            continue

        # Find actual used columns range
        max_col = max(
            (sum(1 for c in row if c is not None and str(c).strip() != '') for row in rows_data),
            default=1
        )

        table = f'<h2 class="sheet-title">{sheet_name}</h2><table>'
        for ri, row in enumerate(rows_data):
            # Skip completely empty rows
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            tag = 'th' if ri == 0 else 'td'
            cells = ''.join(
                f'<{tag}>{("" if (c is None or str(c).strip() == "") else str(c))}</{tag}>'
                for c in row
            )
            table += f'<tr>{cells}</tr>'
        table += '</table>'
        sheets_html.append(table)

    wb.close()

    html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <title>{safe_name}</title>
  <style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family: Arial, sans-serif; font-size: 11px; padding: 16px; background:#fff; color:#000; }}
    .sheet-title {{ font-size:14px; font-weight:700; margin: 18px 0 6px; color:#1a3a5c; border-bottom:2px solid #1a3a5c; padding-bottom:4px; }}
    table {{ border-collapse: collapse; width:100%; margin-bottom: 24px; page-break-inside: avoid; }}
    th, td {{ border: 1px solid #bbb; padding: 5px 8px; text-align: center; white-space: nowrap; }}
    th {{ background: #1a3a5c; color: #fff; font-size:11px; }}
    tr:nth-child(even) td {{ background: #f5f7fa; }}
    @media print {{
      body {{ padding:8px; }}
      .sheet-title {{ margin-top:10px; }}
      @page {{ margin: 1cm; size: landscape; }}
    }}
  </style>
</head>
<body>
  <div style="text-align:center;margin-bottom:14px;">
    <strong style="font-size:15px;">{safe_name.rsplit('.',1)[0]}</strong>
    <span style="font-size:11px;color:#666;margin-left:10px;">Printed: {datetime.now().strftime('%Y-%m-%d %H:%M')}</span>
  </div>
  {''.join(sheets_html)}
  <script>window.onload = function(){{ window.print(); }};</script>
</body>
</html>"""

    from flask import Response
    return Response(html, mimetype='text/html')

# ══════════════════════════════════════════════════════════════════
#  VISIT COUNTER (file-based)
# ══════════════════════════════════════════════════════════════════
_COUNTER_FILE = os.path.join(os.path.dirname(__file__), 'visit_counter.json')
_counter_lock = threading.Lock()

def _load_counter():
    try:
        with open(_COUNTER_FILE, 'r') as f:
            return json.load(f)
    except Exception:
        return {'total': 0, 'today': 0, 'date': ''}

def _save_counter(data):
    with open(_COUNTER_FILE, 'w') as f:
        json.dump(data, f)

@app.route('/api/track_visit', methods=['POST'])
def api_track_visit():
    today = datetime.now().strftime('%Y-%m-%d')
    with _counter_lock:
        data = _load_counter()
        if data.get('date') != today:
            data['today'] = 0
            data['date']  = today
        data['total'] = data.get('total', 0) + 1
        data['today'] = data.get('today', 0) + 1
        _save_counter(data)
    return jsonify({'ok': True})

@app.route('/api/stats')
def api_stats():
    today = datetime.now().strftime('%Y-%m-%d')
    with _counter_lock:
        data = _load_counter()
    if data.get('date') != today:
        data['today'] = 0
    return jsonify({
        'total': data.get('total', 0),
        'today': data.get('today', 0),
    })

# ══════════════════════════════════════════════════════════════════
#  AI CHAT PROXY
# ══════════════════════════════════════════════════════════════════
@app.route('/api/ai-chat', methods=['POST'])
def ai_chat():
    data = request.get_json(silent=True) or {}
    api_key = os.getenv('GROQ_API_KEY')
    if not api_key:
        return jsonify({'error': 'AI not configured'}), 503

    # Convert Anthropic-style messages to OpenAI-compatible format for Groq
    messages = data.get('messages', [])
    system_prompt = data.get('system', '')

    groq_messages = []
    if system_prompt:
        groq_messages.append({'role': 'system', 'content': system_prompt})
    groq_messages.extend(messages)

    groq_payload = {
        'model': 'llama-3.1-8b-instant',  # Free Groq model
        'max_tokens': data.get('max_tokens', 1000),
        'messages': groq_messages,
    }

    try:
        res = _requests.post(
            'https://api.groq.com/openai/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            },
            json=groq_payload,
            timeout=30
        )
        groq_data = res.json()
        # Convert Groq response to Anthropic-style format so the frontend works as-is
        if 'choices' in groq_data:
            reply_text = groq_data['choices'][0]['message']['content']
            anthropic_style = {
                'content': [{'type': 'text', 'text': reply_text}]
            }
            return jsonify(anthropic_style), 200
        return jsonify(groq_data), res.status_code
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    try:
        from flaskwebgui import FlaskUI
        ui = FlaskUI(
            app=app,
            server='flask',
            width=1280,
            height=800,
            port=3049,
            fullscreen=False,
        )
        ui.run()
    except ImportError:
        import webbrowser, threading, time
        def _open():
            time.sleep(1.2)
            webbrowser.open('http://127.0.0.1:3049')
        threading.Thread(target=_open, daemon=True).start()
        app.run(host='127.0.0.1', port=3049, debug=False)
