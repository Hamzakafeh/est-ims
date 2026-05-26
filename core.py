"""Core utilities and shared state for EST-iMs."""
from dotenv import load_dotenv
load_dotenv()
import os
import io
import sys
import json
import re
import sqlite3
import hashlib
import secrets
import warnings
import threading
import shutil
import requests as _requests
from datetime import datetime, timedelta
from functools import wraps
from flask import session, request, jsonify, redirect, url_for

warnings.filterwarnings('ignore')

# ── Brute-force protection (in-memory, resets on restart) ──────────
from collections import defaultdict
import time as _time

_login_attempts = defaultdict(list)
_MAX_ATTEMPTS = 5
_LOCKOUT_SECONDS = 300
_lockout_until = {}

_LOGIN_LOG_FILE = os.getenv(
    'LOGIN_LOG_FILE',
    os.path.join(os.getenv('RENDER_DISK_PATH', os.path.dirname(os.path.abspath(__file__))), 'login_log.json')
)
_log_lock = threading.Lock()
_country_cache = {}
_verified_usernames = {'mlo5'}
_active_user_sessions = {}
_active_sessions_lock = threading.Lock()
_ACTIVE_SESSION_TTL = 8 * 60 * 60

APP_DIR = os.path.dirname(os.path.abspath(__file__))
LEGACY_AUTH_DB_FILE = os.path.join(APP_DIR, 'auth.sqlite3')
AUTH_DB_FILE = os.getenv(
    'AUTH_DB_FILE',
    os.path.join(os.getenv('RENDER_DISK_PATH', APP_DIR), 'auth.sqlite3')
)

DATA_STORE_DIR = os.getenv('RENDER_DISK_PATH', APP_DIR)
CONTACT_MESSAGES_FILE = os.path.join(DATA_STORE_DIR, 'contact_messages.json')
QC_SUBMISSIONS_FILE = os.path.join(DATA_STORE_DIR, 'qc_submissions.json')
QC_UPLOAD_DIR = os.path.join(APP_DIR, 'static', 'qc_uploads')
_data_lock = threading.Lock()

_SECRET_KEY = os.getenv('SECRET_KEY')
_EDIT_PASSWORD = os.getenv('EDIT_PASSWORD')
EDIT_PASSWORD = _EDIT_PASSWORD

# ── Zone configuration ──────────────────────────────────────────────
ZONES = [
    {'id': 'zone1', 'name': 'Zone 1', 'label': 'زون 1',        'icon': '🏭'},
    {'id': 'zone2', 'name': 'Zone 2', 'label': 'زون 2',        'icon': '🏭'},
    {'id': 'zone3', 'name': 'Zone 3', 'label': 'Packaging',    'icon': '🏭'},
    {'id': 'zone4', 'name': 'Zone 4', 'label': 'زون 4',        'icon': '🏭'},
    {'id': 'zone5', 'name': 'Zone 5', 'label': 'زون 5',        'icon': '🏭'},
    {'id': 'qc',    'name': 'QC',     'label': 'Quality Control','icon': '🔬'},
    {'id': 'admin', 'name': 'Admin',  'label': 'Administration', 'icon': '🏢'},
    {'id': 'dev',   'name': 'Dev',    'label': 'Dev',            'icon': '💻'},
]

ZONE_PASSWORDS = {
    'zone1': os.getenv('ZONE1_PASSWORD'),
    'zone2': os.getenv('ZONE2_PASSWORD'),
    'zone3': os.getenv('ZONE3_PASSWORD'),
    'zone4': os.getenv('ZONE4_PASSWORD'),
    'zone5': os.getenv('ZONE5_PASSWORD'),
    'qc':    os.getenv('QC_PASSWORD'),
    'admin': os.getenv('ADMIN_PASSWORD'),
    'dev':   os.getenv('DEV_PASSWORD'),
}

SUPER_ZONES = {'admin', 'dev'}
EDIT_ZONES = {'dev'}
WAREHOUSE_ZONES = ('zone1', 'zone2', 'zone3', 'zone4', 'zone5', 'qc')
ZONE_USER_RESTRICTIONS = {
    'dev': 'mlo5',
    'admin': 'ink',
}
ZONE_ALLOWED_USERS = {
    'qc': [],
}

MONTH_ORDER = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4,
    'May': 5, 'June': 6, 'July': 7, 'August': 8,
    'September': 9, 'October': 10, 'November': 11, 'December': 12
}
MONTH_AR = {
    'January': 'يناير', 'February': 'فبراير', 'March': 'مارس', 'April': 'أبريل',
    'May': 'مايو', 'June': 'يونيو', 'July': 'يوليو', 'August': 'أغسطس',
    'September': 'سبتمبر', 'October': 'أكتوبر', 'November': 'نوفمبر', 'December': 'ديسمبر'
}

# ── Stocktaking / QR helpers ─────────────────────────────────────────
SKU_MAP = {
    'CHKN-BLU-001': ('Chicken', 'Blue'),
    'CHKN-BRN-001': ('Chicken', 'Brown'),
    'CHKN-DBL-001': ('Chicken', 'Dark Blue'),
    'CHKN-DGR-001': ('Chicken', 'Dark Green'),
    'CHKN-GRY-001': ('Chicken', 'Gray'),
    'CHKN-GRN-001': ('Chicken', 'Green'),
    'CHKN-LBL-001': ('Chicken', 'Light Blue'),
    'CHKN-ORG-001': ('Chicken', 'Orange'),
    'CHKN-ORG-002': ('Chicken', 'Orange (70*45)'),
    'CHKN-PRP-001': ('Chicken', 'Purple'),
    'CHKN-RED-001': ('Chicken', 'Red'),
    'CHKN-YLW-001': ('Chicken', 'Yellow'),
}

SKU_NAMES_AR = {
    'CHKN-BLU-001': 'أزرق',
    'CHKN-BRN-001': 'بني',
    'CHKN-DBL-001': 'أزرق غامق',
    'CHKN-DGR-001': 'أخضر غامق',
    'CHKN-GRY-001': 'رمادي',
    'CHKN-GRN-001': 'أخضر',
    'CHKN-LBL-001': 'أزرق فاتح',
    'CHKN-ORG-001': 'برتقالي',
    'CHKN-ORG-002': 'برتقالي 70×45',
    'CHKN-PRP-001': 'بنفسجي',
    'CHKN-RED-001': 'أحمر',
    'CHKN-YLW-001': 'أصفر',
}

SKU_HEX = {
    'CHKN-BLU-001': '#3b82f6', 'CHKN-BRN-001': '#92400e', 'CHKN-DBL-001': '#1e3a8a',
    'CHKN-DGR-001': '#14532d', 'CHKN-GRY-001': '#6b7280', 'CHKN-GRN-001': '#16a34a',
    'CHKN-LBL-001': '#7dd3fc', 'CHKN-ORG-001': '#f97316', 'CHKN-ORG-002': '#fb923c',
    'CHKN-PRP-001': '#7c3aed', 'CHKN-RED-001': '#dc2626', 'CHKN-YLW-001': '#eab308',
}

STOCK_QR_FILE_MAP = {
    'OTHER': 'Other+.xlsm',
    'SACKS': 'Sacks.xlsm',
}

STOCK_QR_HEX = {
    'BLUE': '#3b82f6', 'DARK BLUE': '#1e3a8a', 'LIGHT BLUE': '#7dd3fc',
    'GREEN': '#16a34a', 'DARK GREEN': '#14532d', 'RED': '#dc2626',
    'ORANGE': '#f97316', 'PURPLE': '#7c3aed', 'BROWN': '#92400e',
    'YELLOW': '#eab308', 'GRAY': '#6b7280', 'WHITE': '#e5e7eb',
    'BLACK': '#111827',
}


def _get_ip():
    return request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip()


def _is_locked_out(ip):
    now = _time.time()
    if ip in _lockout_until and now >= _lockout_until[ip]:
        _lockout_until.pop(ip, None)
        _login_attempts.pop(ip, None)
    if ip in _lockout_until:
        return True
    _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < _LOCKOUT_SECONDS]
    if len(_login_attempts[ip]) >= _MAX_ATTEMPTS:
        _lockout_until[ip] = now + _LOCKOUT_SECONDS
        return True
    return False


def _lockout_remaining(ip):
    now = _time.time()
    until = _lockout_until.get(ip, 0)
    return max(0, int(until - now))


def _record_failed_attempt(ip):
    _login_attempts[ip].append(_time.time())


def _clear_attempts(ip):
    _login_attempts.pop(ip, None)
    _lockout_until.pop(ip, None)


def _is_single_login_exempt(username):
    uname = _normalize_username(username)
    privileged = {_normalize_username(v) for v in ZONE_USER_RESTRICTIONS.values() if v}
    return uname in privileged


def _active_session_alive(record):
    return bool(record) and (_time.time() - float(record.get('last_seen', 0))) < _ACTIVE_SESSION_TTL


def _register_active_session(username):
    token = secrets.token_hex(16)
    uname = _normalize_username(username)
    with _active_sessions_lock:
        _active_user_sessions[uname] = {'token': token, 'last_seen': _time.time()}
    session['session_token'] = token


def _clear_active_session(username=None):
    explicit_username = username is not None
    uname = _normalize_username(username or session.get('username', ''))
    token = session.get('session_token')
    if not uname:
        return
    with _active_sessions_lock:
        record = _active_user_sessions.get(uname)
        if explicit_username or not token or (record and record.get('token') == token):
            _active_user_sessions.pop(uname, None)


def _validate_active_session():
    if not session.get('logged_in'):
        return True
    username = session.get('username', '')
    if _is_single_login_exempt(username):
        return True
    uname = _normalize_username(username)
    token = session.get('session_token')
    with _active_sessions_lock:
        record = _active_user_sessions.get(uname)
        if not record:
            return False
        if record.get('token') != token:
            return False
        if not _active_session_alive(record):
            _active_user_sessions.pop(uname, None)
            return False
        record['last_seen'] = _time.time()
    return True

VERIFIED_USERS = _verified_usernames


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('auth.login_page'))
        return f(*args, **kwargs)
    return decorated


def zone_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('auth.login_page'))
        if not session.get('zone'):
            return redirect(url_for('zones.zones_page'))
        return f(*args, **kwargs)
    return decorated


def _read_login_log():
    try:
        with open(_LOGIN_LOG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


def _write_login_log(entries):
    log_dir = os.path.dirname(_LOGIN_LOG_FILE)
    if log_dir:
        os.makedirs(log_dir, exist_ok=True)
    with open(_LOGIN_LOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(entries[-500:], f, ensure_ascii=False)


def _ip_country(ip):
    if not ip or ip.startswith(('127.', '10.', '192.168.', '172.16.', '172.17.', '172.18.', '172.19.', '172.20.', '172.21.', '172.22.', '172.23.', '172.24.', '172.25.', '172.26.', '172.27.', '172.28.', '172.29.', '172.30.', '172.31.')):
        return 'Local / Private'
    if ip in _country_cache:
        return _country_cache[ip]
    try:
        res = _requests.get(f'https://ipapi.co/{ip}/json/', timeout=2)
        data = res.json() if res.ok else {}
        country = data.get('country_name') or data.get('country') or 'Unknown'
    except Exception:
        country = 'Unknown'
    _country_cache[ip] = country
    return country


def _record_login(username, zone_id, zone_label, ip):
    with _log_lock:
        entries = _read_login_log()
        entries.append({
            'username': username,
            'zone_id': zone_id,
            'zone_label': zone_label,
            'ip': ip,
            'country': _ip_country(ip),
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })
        _write_login_log(entries)


def _prepare_auth_db_file():
    db_dir = os.path.dirname(AUTH_DB_FILE)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    if os.path.abspath(AUTH_DB_FILE) != os.path.abspath(LEGACY_AUTH_DB_FILE):
        if not os.path.exists(AUTH_DB_FILE) and os.path.exists(LEGACY_AUTH_DB_FILE):
            shutil.copy2(LEGACY_AUTH_DB_FILE, AUTH_DB_FILE)


def _normalize_username(value):
    return str(value or '').strip().lower()

ENV_USERS = {
    _normalize_username(k): v
    for k, v in {
        os.getenv('USER1'): os.getenv('PASS1'),
        os.getenv('USER2'): os.getenv('PASS2'),
        os.getenv('USER3'): os.getenv('PASS3'),
        os.getenv('USER4'): os.getenv('PASS4'),
    }.items()
    if k
}


def _normalize_text(value):
    return re.sub(r'\s+', ' ', str(value or '').strip()).lower()


def _hash_secret(value, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        'sha256',
        str(value or '').encode('utf-8'),
        bytes.fromhex(salt),
        200_000,
    ).hex()
    return f'pbkdf2_sha256${salt}${digest}'


def _verify_secret(value, stored):
    if not stored:
        return False
    try:
        algo, salt, digest = str(stored).split('$', 2)
    except ValueError:
        return False
    if algo != 'pbkdf2_sha256':
        return False
    test = hashlib.pbkdf2_hmac(
        'sha256',
        str(value or '').encode('utf-8'),
        bytes.fromhex(salt),
        200_000,
    ).hex()
    return test == digest


# ── Database backend (SQLite local / PostgreSQL on Render) ─────────
_DATABASE_URL = os.getenv('DATABASE_URL', '')
if _DATABASE_URL.startswith('postgres://'):
    _DATABASE_URL = 'postgresql://' + _DATABASE_URL[len('postgres://'):]
_USE_PG = bool(_DATABASE_URL)

if _USE_PG:
    import psycopg2
    import psycopg2.extras


class _PGCursor:
    """Wraps psycopg2 cursor to match sqlite3 cursor API."""
    __slots__ = ('_cur',)

    def __init__(self, cur):
        self._cur = cur

    def fetchone(self):
        return self._cur.fetchone()

    def fetchall(self):
        return self._cur.fetchall()

    @property
    def rowcount(self):
        return self._cur.rowcount


class _PGConn:
    """Wraps psycopg2 connection to match sqlite3 connection API used throughout the codebase."""
    __slots__ = ('_conn',)

    def __init__(self, conn):
        self._conn = conn

    def execute(self, sql, params=()):
        cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(re.sub(r'\?', '%s', sql), params if params else None)
        return _PGCursor(cur)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self._conn.commit()
        else:
            self._conn.rollback()
        self._conn.close()
        return False


def _db_connect():
    if _USE_PG:
        return _PGConn(psycopg2.connect(_DATABASE_URL))
    conn = sqlite3.connect(AUTH_DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def _init_auth_db():
    if not _USE_PG:
        _prepare_auth_db_file()

    _id_col  = 'id BIGSERIAL PRIMARY KEY' if _USE_PG else 'id INTEGER PRIMARY KEY AUTOINCREMENT'
    _add_col = 'ADD COLUMN IF NOT EXISTS' if _USE_PG else 'ADD COLUMN'

    with _db_connect() as conn:
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS users (
                {_id_col},
                full_name TEXT NOT NULL,
                username TEXT NOT NULL UNIQUE,
                email TEXT NOT NULL,
                phone TEXT NOT NULL,
                job_title TEXT NOT NULL,
                gender TEXT,
                birth_date TEXT,
                privacy_accepted INTEGER NOT NULL DEFAULT 0,
                password_hash TEXT NOT NULL,
                security_question TEXT NOT NULL,
                security_answer_hash TEXT NOT NULL,
                approved INTEGER NOT NULL DEFAULT 0,
                is_admin INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                approved_at TEXT,
                created_by TEXT,
                suspended_until TEXT,
                suspended_by TEXT,
                suspended_at TEXT
            )
        """)
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS registration_requests (
                {_id_col},
                full_name TEXT NOT NULL,
                username TEXT NOT NULL,
                email TEXT NOT NULL,
                phone TEXT NOT NULL,
                job_title TEXT NOT NULL,
                gender TEXT,
                birth_date TEXT,
                privacy_accepted INTEGER NOT NULL DEFAULT 0,
                password_hash TEXT NOT NULL,
                security_question TEXT NOT NULL,
                security_answer_hash TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL,
                reviewed_at TEXT,
                reviewed_by TEXT,
                review_note TEXT
            )
        """)
        for column_sql in (
            f"ALTER TABLE users {_add_col} gender TEXT",
            f"ALTER TABLE users {_add_col} birth_date TEXT",
            f"ALTER TABLE users {_add_col} privacy_accepted INTEGER NOT NULL DEFAULT 0",
            f"ALTER TABLE users {_add_col} suspended_until TEXT",
            f"ALTER TABLE users {_add_col} suspended_by TEXT",
            f"ALTER TABLE users {_add_col} suspended_at TEXT",
            f"ALTER TABLE registration_requests {_add_col} gender TEXT",
            f"ALTER TABLE registration_requests {_add_col} birth_date TEXT",
            f"ALTER TABLE registration_requests {_add_col} privacy_accepted INTEGER NOT NULL DEFAULT 0",
        ):
            if _USE_PG:
                conn.execute(column_sql)
            else:
                try:
                    conn.execute(column_sql)
                except sqlite3.OperationalError:
                    pass


def _username_in_env(username):
    uname = _normalize_username(username)
    return uname in ENV_USERS


def _get_db_user(username):
    uname = _normalize_username(username)
    if not uname:
        return None
    with _db_connect() as conn:
        return conn.execute(
            "SELECT * FROM users WHERE lower(username) = ?",
            (uname,),
        ).fetchone()


def _username_exists_everywhere(username):
    uname = _normalize_username(username)
    if not uname:
        return False
    if uname in ENV_USERS:
        return True
    with _db_connect() as conn:
        user = conn.execute("SELECT 1 FROM users WHERE lower(username) = ?", (uname,)).fetchone()
        if user:
            return True
        req = conn.execute(
            "SELECT 1 FROM registration_requests WHERE lower(username) = ? AND status = 'pending'",
            (uname,),
        ).fetchone()
        return bool(req)


def _approved_db_user(username):
    user = _get_db_user(username)
    if user and int(user['approved'] or 0) == 1:
        return user
    return None


def _user_suspension_remaining(user):
    if not user:
        return 0
    until = user['suspended_until'] if 'suspended_until' in user.keys() else None
    if not until:
        return 0
    try:
        dt = datetime.strptime(until, '%Y-%m-%d %H:%M:%S')
    except Exception:
        return 0
    remaining = int((dt - datetime.now()).total_seconds())
    return max(0, remaining)


def _pending_request_count():
    with _db_connect() as conn:
        row = conn.execute("SELECT COUNT(*) AS c FROM registration_requests WHERE status = 'pending'").fetchone()
    return int(row['c'] if row else 0)


def _read_json_list(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _write_json_list(path, items):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def _next_json_id(items):
    return (max([int(x.get('id', 0) or 0) for x in items] or [0]) + 1)


def get_years_root():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(script_dir, 'zones'),
        os.path.join(os.path.dirname(script_dir), 'zones'),
    ]
    for c in candidates:
        if os.path.isdir(c):
            return c
    return None


def _validate_filepath(filepath, zone_id=None):
    root = get_years_root()
    if not root:
        return False
    try:
        real_root = os.path.realpath(root)
        real_file = os.path.realpath(filepath)
        if not real_file.startswith(real_root + os.sep):
            return False
        if zone_id:
            zone_root = os.path.realpath(os.path.join(root, zone_id))
            if not real_file.startswith(zone_root + os.sep):
                return False
        return True
    except Exception:
        return False


def get_base_path(year=None, zone_id=None):
    root = get_years_root()
    if not root:
        return None
    if zone_id:
        zone_folder = os.path.join(root, zone_id)
    else:
        zone_folders = sorted([d for d in os.listdir(root) if os.path.isdir(os.path.join(root, d)) and d.startswith('zone')])
        if not zone_folders:
            return None
        zone_folder = os.path.join(root, zone_folders[0])
    if not os.path.isdir(zone_folder):
        return None
    years = sorted([d for d in os.listdir(zone_folder) if os.path.isdir(os.path.join(zone_folder, d)) and d.isdigit()])
    if not years:
        return None
    target = str(year) if year else years[0]
    candidate = os.path.join(zone_folder, target)
    return candidate if os.path.isdir(candidate) else None


def get_available_years():
    root = get_years_root()
    if not root:
        return []
    years = set()
    for zone_folder in os.listdir(root):
        zone_path = os.path.join(root, zone_folder)
        if not os.path.isdir(zone_path):
            continue
        for d in os.listdir(zone_path):
            if os.path.isdir(os.path.join(zone_path, d)) and d.isdigit():
                years.add(d)
    return sorted(years)


def get_structure(year=None, zone_id=None):
    root = get_years_root()
    if not root:
        return {}
    available_years = get_available_years()
    if not available_years:
        return {}
    target_years = [str(year)] if year else available_years
    if zone_id:
        zone_folders = [zone_id]
    else:
        zone_folders = sorted([d for d in os.listdir(root) if os.path.isdir(os.path.join(root, d)) and d.startswith('zone')])
    result = {}
    for yr in target_years:
        result[yr] = {}
        for zf in zone_folders:
            year_path = os.path.join(root, zf, yr)
            if not os.path.isdir(year_path):
                continue
            for month_folder in sorted(os.listdir(year_path), key=lambda m: MONTH_ORDER.get(m.split('-')[-1] if '-' in m else m, 99)):
                month_path = os.path.join(year_path, month_folder)
                if not os.path.isdir(month_path):
                    continue
                month_name = month_folder.split('-')[-1] if '-' in month_folder else month_folder
                files = {}
                for fname in ['Other+', 'Sacks']:
                    fpath = os.path.join(month_path, f'{fname}.xlsm')
                    if os.path.exists(fpath):
                        files[fname] = fpath
                if files:
                    if month_name not in result[yr]:
                        result[yr][month_name] = {}
                    result[yr][month_name].update(files)
    return result


def read_sheet_data(filepath, sheet_name):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None, []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if sheet_name == 'Log':
            if not rows:
                wb.close(); return None, []
            col_indices = [(ci, str(val).strip()) for ci, val in enumerate(rows[0]) if val is not None and str(val).strip()]
            if not col_indices:
                wb.close(); return None, []
            data_rows = []
            for row in rows[1:]:
                if not any(ci < len(row) and row[ci] is not None and str(row[ci]).strip() not in ('','None') for ci,_ in col_indices):
                    continue
                rd = {}
                for ci, col_name in col_indices:
                    val = row[ci] if ci < len(row) else None
                    if isinstance(val, datetime): val = val.strftime('%Y-%m-%d %H:%M:%S')
                    elif val is not None:
                        val = str(val) if not isinstance(val,(int,float)) else val
                        if isinstance(val,str) and val.strip()=='None': val=None
                    rd[col_name] = val
                data_rows.append(rd)
            wb.close()
            return [h for _,h in col_indices], data_rows

        if sheet_name == 'Stocktaking':
            data_rows = []
            for row in rows:
                non_empty = [(ci,val) for ci,val in enumerate(row) if val is not None and str(val).strip() not in ('','None')]
                if not non_empty: continue
                rd = {}
                for ci,val in non_empty:
                    col_label = f'Col {ci+1}'
                    if isinstance(val,datetime): val=val.strftime('%Y-%m-%d')
                    elif val is not None: val=str(val) if not isinstance(val,(int,float)) else val
                    rd[col_label]=val
                data_rows.append(rd)
            if not data_rows:
                wb.close(); return None,[ ]
            all_cols,seen=[],set()
            for r in data_rows:
                for k in r:
                    if k not in seen: all_cols.append(k); seen.add(k)
            wb.close()
            return all_cols, data_rows

        header_idx = None
        for i, row in enumerate(rows):
            rv = [str(v).strip() if v else '' for v in row]
            if 'Date' in rv or 'التاريخ' in rv:
                header_idx = i; break
        if header_idx is None:
            wb.close(); return None, []

        col_indices = [(ci,str(val).strip()) for ci,val in enumerate(rows[header_idx]) if val is not None and str(val).strip()]
        data_rows = []
        for row_idx, row in enumerate(rows[header_idx+1:], start=header_idx+2):
            if not any(ci<len(row) and row[ci] is not None and str(row[ci]).strip() not in ('','None') for ci,_ in col_indices):
                continue
            rd={}
            for ci,col_name in col_indices:
                val=row[ci] if ci<len(row) else None
                if isinstance(val,datetime): val=val.strftime('%Y-%m-%d')
                elif val is not None:
                    val=str(val) if not isinstance(val,(int,float)) else val
                    if isinstance(val,str) and val.strip()=='None': val=None
                rd[col_name]=val
            rd['__row__'] = row_idx
            data_rows.append(rd)
        wb.close()
        headers = [h for _,h in col_indices]
        return headers, data_rows
    except Exception as _exc:
        return None, []


def _get_cell_val(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for merge in ws.merged_cells.ranges:
        if cell.coordinate in merge:
            master = ws.cell(row=merge.min_row, column=merge.min_col)
            return master.value
    return None


def _find_last_balance(ws, target_row, color_value):
    for i in range(target_row - 1, 6, -1):
        cell_color = ws.cell(row=i, column=COL_COLOR).value
        if cell_color == color_value:
            balance = ws.cell(row=i, column=COL_CURRENT).value
            try:
                return float(balance or 0), True
            except Exception:
                return 0.0, True
    basic = ws.cell(row=target_row, column=COL_BASIC).value
    try:
        return float(basic or 0), False
    except Exception:
        return 0.0, False


def _append_log(ws_log, operation, qty, balance, color, size, item_type, category):
    lr = 1
    for row in ws_log.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value is not None:
                lr = cell.row
    lr += 1
    ws_log.cell(row=lr, column=LOG_COL_TIME).value     = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_log.cell(row=lr, column=LOG_COL_TYPE).value     = operation
    ws_log.cell(row=lr, column=LOG_COL_QTY).value      = qty
    ws_log.cell(row=lr, column=LOG_COL_BALANCE).value  = balance
    ws_log.cell(row=lr, column=LOG_COL_COLOR).value    = color
    ws_log.cell(row=lr, column=LOG_COL_SIZE).value     = size
    ws_log.cell(row=lr, column=LOG_COL_ITEMTYPE).value = item_type
    ws_log.cell(row=lr, column=LOG_COL_CATEGORY).value = category


def _col_letter_to_idx(letter):
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1


_LOOKUP_RE = re.compile(
    r'LOOKUP\s*\(\s*2\s*,\s*1\s*/\s*\(\s*(\w+)!\s*([A-Z]+)\s*:\s*\2\s*=\s*(.*?)\s*\)\s*,\s*\1!\s*([A-Z]+)\s*:\s*\4\s*\)',
    re.IGNORECASE
)


def _recalc_stocktaking(wb):
    if 'Stocktaking' not in wb.sheetnames:
        return
    ws_st = wb['Stocktaking']
    _sheet_rows_cache = {}

    def _get_rows(sheet_name):
        if sheet_name not in _sheet_rows_cache:
            if sheet_name in wb.sheetnames:
                _sheet_rows_cache[sheet_name] = list(
                    wb[sheet_name].iter_rows(min_row=DATA_START_ROW, values_only=True)
                )
            else:
                _sheet_rows_cache[sheet_name] = []
        return _sheet_rows_cache[sheet_name]

    for row in ws_st.iter_rows():
        for cell in row:
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                continue
            m = _LOOKUP_RE.search(cell.value)
            if not m:
                continue
            ref_sheet = m.group(1)
            filter_col = _col_letter_to_idx(m.group(2))
            raw_val = m.group(3).strip().strip('"')
            result_col = _col_letter_to_idx(m.group(4))
            try:
                filter_val = float(raw_val) if '.' in raw_val else int(raw_val)
            except ValueError:
                filter_val = raw_val
            last_result = None
            for data_row in _get_rows(ref_sheet):
                try:
                    cell_filter = data_row[filter_col]
                    if isinstance(filter_val, str):
                        match = (cell_filter is not None and str(cell_filter).strip() == filter_val)
                    else:
                        try:
                            match = float(cell_filter) == filter_val
                        except (TypeError, ValueError):
                            match = False
                    if match:
                        v = data_row[result_col]
                        if v is not None:
                            try:
                                last_result = float(v)
                            except (TypeError, ValueError):
                                last_result = v
                except IndexError:
                    continue
            cell.value = last_result if last_result is not None else cell.value

# ── Excel column definitions ────────────────────────────────────────
COL_DATE     = 1
COL_CATEGORY = 4
COL_TYPE     = 5
COL_COLOR    = 6
COL_SIZE     = 7
COL_BASIC    = 9
COL_CURRENT  = 10
COL_IN       = 12
COL_OUT      = 13
DATA_START_ROW = 7
LOG_COL_TIME     = 1
LOG_COL_TYPE     = 2
LOG_COL_QTY      = 3
LOG_COL_BALANCE  = 4
LOG_COL_COLOR    = 5
LOG_COL_SIZE     = 6
LOG_COL_ITEMTYPE = 7
LOG_COL_CATEGORY = 8


def _parse_dv_formula(formula1):
    if not formula1:
        return []
    s = formula1.strip().strip('"')
    items = [v.strip() for v in s.split(',') if v.strip() and v.strip().lower() not in ('null', 'none', '')]
    return items


def _col_letter_to_index(letter):
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def _find_latest_sacks_file():
    root = get_years_root()
    if not root:
        return None
    best_path = None
    best_key = (-1, -1)
    month_num = {
        'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,'july':7,'august':8,
        'september':9,'october':10,'november':11,'december':12
    }
    for zone_folder in os.listdir(root):
        zone_path = os.path.join(root, zone_folder)
        if not os.path.isdir(zone_path):
            continue
        for year_folder in os.listdir(zone_path):
            year_path = os.path.join(zone_path, year_folder)
            if not os.path.isdir(year_path) or not year_folder.isdigit():
                continue
            year_int = int(year_folder)
            for month_folder in os.listdir(year_path):
                month_path = os.path.join(year_path, month_folder)
                sacks_path = os.path.join(month_path, 'Sacks.xlsm')
                if not os.path.isfile(sacks_path):
                    continue
                parts = month_folder.split('-')
                month_int = int(parts[0]) if parts[0].isdigit() else month_num.get(parts[-1].lower(), 0)
                key = (year_int, month_int)
                if key > best_key:
                    best_key = key
                    best_path = sacks_path
    return best_path


def _get_last_balance(sheet_name, color_name):
    filepath = _find_latest_sacks_file()
    if not filepath:
        return None, None, None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close(); return None, None, None
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        hdr_idx = None
        headers = []
        for i, row in enumerate(rows):
            rv = [str(c).strip() if c else '' for c in row]
            if 'Current Balance' in rv:
                hdr_idx = i
                headers = rv
                break
        if hdr_idx is None:
            return None, None, None
        ci_color   = headers.index('Color')           if 'Color'           in headers else None
        ci_balance = headers.index('Current Balance') if 'Current Balance' in headers else None
        ci_date    = headers.index('Date')            if 'Date'            in headers else None
        if ci_color is None or ci_balance is None:
            return None, None, None
        last_balance = None
        last_date = None
        for row in rows[hdr_idx + 1:]:
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            row_color = str(row[ci_color]).strip() if ci_color < len(row) and row[ci_color] else ''
            if row_color.lower() != color_name.lower():
                continue
            bal = row[ci_balance] if ci_balance < len(row) else None
            if bal is not None and str(bal).strip() not in ('', 'None', '/'):
                try:
                    last_balance = float(bal)
                    if ci_date and ci_date < len(row) and row[ci_date]:
                        d = row[ci_date]
                        last_date = d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d)[:10]
                except (ValueError, TypeError):
                    pass
        return last_balance, last_date, filepath
    except Exception:
        return None, None, None


def _parse_stocktaking_sku(sku):
    m = re.fullmatch(r'STK-(ZONE[0-9]+)-(\d{4})-(\d{2})-(OTHER|SACKS)-R(\d{3})-C(\d{3})', sku or '')
    if not m:
        return None
    zone_id, year, month_code, file_key, row_s, col_s = m.groups()
    return {
        'zone_id': zone_id.lower(),
        'year': year,
        'month_code': month_code,
        'file_key': file_key,
        'row': int(row_s),
        'col': int(col_s),
    }


def _stocktaking_category(ws, row, col, item_name):
    item_text = str(item_name or '').strip()
    for merged in getattr(ws, 'merged_cells', []).ranges:
        if merged.min_row <= 1 <= merged.max_row and merged.min_col <= col <= merged.max_col:
            value = ws.cell(merged.min_row, merged.min_col).value
            if value is not None and str(value).strip() and str(value).strip() != item_text:
                return str(value).strip()
    for c in range(col, 0, -1):
        value = ws.cell(1, c).value
        if value is not None and str(value).strip() and str(value).strip() != item_text:
            return str(value).strip()
    for r in range(row - 1, 0, -1):
        value = ws.cell(r, col).value
        if value is None or str(value).strip() == '':
            continue
        if isinstance(value, (int, float)):
            continue
        text = str(value).strip()
        if text != item_text:
            return text
    return ''


def _stocktaking_scan_result(sku):
    meta = _parse_stocktaking_sku(sku)
    if not meta:
        return None
    root = get_years_root()
    if not root:
        return {'found': False, 'error': 'تعذر تحديد مجلد zones'}
    filename = STOCK_QR_FILE_MAP.get(meta['file_key'])
    year_path = os.path.join(root, meta['zone_id'], meta['year'])
    month_folder = _find_month_folder(year_path, meta['month_code'])
    if not month_folder or not filename:
        return {'found': False, 'error': 'ملف الشهر غير موجود'}
    filepath = os.path.join(year_path, month_folder, filename)
    if not os.path.isfile(filepath):
        return {'found': False, 'error': 'ملف الإكسيل غير موجود'}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
        if 'Stocktaking' not in wb.sheetnames:
            wb.close()
            return {'found': False, 'error': 'شيت Stocktaking غير موجود'}
        ws = wb['Stocktaking']
        item_name = ws.cell(meta['row'], meta['col']).value
        balance = ws.cell(meta['row'] + 1, meta['col']).value
        category = _stocktaking_category(ws, meta['row'], meta['col'], item_name)
        wb.close()
        if item_name is None or str(item_name).strip() == '':
            return {'found': False, 'error': 'الصنف غير موجود داخل Stocktaking'}
        try:
            balance_value = float(balance)
            balance_out = int(balance_value) if balance_value.is_integer() else balance_value
        except (TypeError, ValueError, AttributeError):
            balance_out = balance if balance not in (None, '') else 0
        modified = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%Y-%m-%d %H:%M')
        return {
            'found': True,
            'sku': sku,
            'nameAr': str(item_name).strip(),
            'category': category or meta['file_key'].title(),
            'color': str(item_name).strip(),
            'hex': _guess_stock_hex(item_name),
            'balance': balance_out,
            'date': modified,
            'file': filename,
            'zone': meta['zone_id'],
            'year': meta['year'],
            'month': month_folder,
        }
    except Exception:
        return {'found': False, 'error': 'تعذر قراءة بيانات Stocktaking'}


def _guess_stock_hex(name):
    text = str(name or '').upper()
    for key, color in STOCK_QR_HEX.items():
        if key in text:
            return color
    return '#1a3a5c'


def _find_month_folder(year_path, month_code):
    if not os.path.isdir(year_path):
        return None
    prefix = f'{int(month_code):02d}-'
    for folder in os.listdir(year_path):
        full = os.path.join(year_path, folder)
        if os.path.isdir(full) and folder.startswith(prefix):
            return folder
    return None


def _send_push_notification(subscription_info, title, body, url='/qc-workflow', tag='qc-update'):
    try:
        from pywebpush import webpush, Vapid
        private_key_pem = os.getenv('VAPID_PRIVATE_KEY', '')
        email = os.getenv('VAPID_CLAIMS_EMAIL', 'admin@example.com')
        if not private_key_pem:
            return False
        # Restore newlines if Render flattened them
        pem = private_key_pem.replace('\\n', '\n')
        vapid = Vapid.from_string(private_key=pem)
        payload = json.dumps({
            'title': title,
            'body':  body,
            'url':   url,
            'tag':   tag,
            'icon':  '/static/icons/icon-192.png',
            'badge': '/static/icons/icon-192.png',
        })
        webpush(
            subscription_info=subscription_info,
            data=payload,
            vapid_private_key=vapid,
            vapid_claims={'sub': f'mailto:{email}'},
        )
        return True
    except Exception:
        return False


_push_subs_lock = threading.Lock()


def _read_push_subs():
    try:
        with open(os.path.join(DATA_STORE_DIR, 'push_subscriptions.json'), 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _write_push_subs(data):
    os.makedirs(os.path.dirname(os.path.join(DATA_STORE_DIR, 'push_subscriptions.json')) or '.', exist_ok=True)
    with open(os.path.join(DATA_STORE_DIR, 'push_subscriptions.json'), 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)


def get_firebase_config():
    """Return Firebase client config dict from env vars."""
    return {
        'apiKey':            os.getenv('FIREBASE_API_KEY', ''),
        'authDomain':        os.getenv('FIREBASE_AUTH_DOMAIN', ''),
        'databaseURL':       os.getenv('FIREBASE_DATABASE_URL', ''),
        'projectId':         os.getenv('FIREBASE_PROJECT_ID', ''),
        'storageBucket':     os.getenv('FIREBASE_STORAGE_BUCKET', ''),
        'messagingSenderId': os.getenv('FIREBASE_MESSAGING_ID', ''),
        'appId':             os.getenv('FIREBASE_APP_ID', ''),
    }


def _firebase_set_user_status(username, status, message=''):
    """Write user_status/{username} in Firebase Realtime DB — non-blocking.
    Used to notify a logged-in user in real-time when their account is deleted or suspended.
    Requires FIREBASE_DATABASE_URL and FIREBASE_DB_SECRET env vars.
    """
    def _do():
        try:
            db_url = os.getenv('FIREBASE_DATABASE_URL', '').rstrip('/')
            db_secret = os.getenv('FIREBASE_DB_SECRET', '')
            if not db_url or not db_secret:
                return
            # Firebase keys cannot contain . # $ [ ] /
            safe_key = re.sub(r'[.#$\[\]/]', '_', username)
            url = f"{db_url}/user_status/{safe_key}.json?auth={db_secret}"
            payload = {'status': status, 'message': message, 'ts': _time.time()}
            _requests.put(url, json=payload, timeout=8)
        except Exception:
            pass
    threading.Thread(target=_do, daemon=True).start()


def _verify_recaptcha(token):
    """Verify a reCAPTCHA v2 token with Google. Returns (ok: bool, error_msg: str)."""
    from urllib.request import urlopen
    from urllib.parse import urlencode
    secret = os.getenv('RECAPTCHA_SECRET_KEY', '')
    if not secret:
        return False, 'الكابتشا غير مفعّلة على الخادم'
    if not token:
        return False, 'يرجى إتمام التحقق من الكابتشا'
    try:
        payload = urlencode({'secret': secret, 'response': token}).encode()
        with urlopen('https://www.google.com/recaptcha/api/siteverify', data=payload, timeout=10) as f:
            result = json.loads(f.read())
    except Exception as e:
        return False, f'تعذر التحقق من الكابتشا: {e}'
    if result.get('success'):
        return True, ''
    return False, 'فشل التحقق من الكابتشا، حاول مجدداً'


def _firebase_clear_user_status(username):
    """Remove user_status/{username} from Firebase (e.g. on unsuspend)."""
    def _do():
        try:
            db_url = os.getenv('FIREBASE_DATABASE_URL', '').rstrip('/')
            db_secret = os.getenv('FIREBASE_DB_SECRET', '')
            if not db_url or not db_secret:
                return
            safe_key = re.sub(r'[.#$\[\]/]', '_', username)
            url = f"{db_url}/user_status/{safe_key}.json?auth={db_secret}"
            _requests.delete(url, timeout=8)
        except Exception:
            pass
    threading.Thread(target=_do, daemon=True).start()


_COUNTER_FILE = os.path.join(APP_DIR, 'visit_counter.json')
_counter_lock = threading.Lock()


def _load_counter():
    try:
        with open(_COUNTER_FILE, 'r') as f:
            return json.load(f)
    except Exception:
        return {'total': 0, 'today': 0, 'date': ''}


def _save_counter(data):
    with open(_COUNTER_FILE, 'w') as f:
        json.dump(data, f)


def create_app():
    from flask import Flask
    app = Flask(__name__, static_folder='static')
    app.config['TEMPLATES_AUTO_RELOAD'] = True
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    app.config['PERMANENT_SESSION_LIFETIME'] = 28800
    return app


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('auth.login_page'))
        return f(*args, **kwargs)
    return decorated


def zone_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('auth.login_page'))
        if not session.get('zone'):
            return redirect(url_for('zones.zones_page'))
        return f(*args, **kwargs)
    return decorated

_init_auth_db()
